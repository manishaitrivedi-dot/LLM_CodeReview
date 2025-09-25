import os
import ast
import csv
from abc import abstractmethod
from io import StringIO
import numpy as np
import pandas as pd
import qubles.core.functions.datetools as datetools
from datetime import datetime
from json import loads, dumps
from sys import exc_info
from threading import current_thread, RLock
from inspect import getfullargspec, getsourcelines
from csv import writer as csv_writer
from collections import defaultdict
from contextlib import closing
from copy import deepcopy
from re import escape, search, sub, subn
from types import MethodType
from utils.path import normalize
from qubles.core.classes.index import Index
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from qubles.core.exceptions.datalib_exceptions import (
    TemplateError,
    BuilderError,
    DataLibCSVError,
    UnregisteredFieldError,
    GracefulUnregisteredFieldError,
    ControlPropertyError,
    MultiBuildError,
    RefLibExpectedError,
    DynamicBuilderSyntaxError,
)
from qubles.core.exceptions.libaddress_exceptions import UnresolvedDomainError
from qubles.core.classes.daterange import DateRange, make_index
from qubles.core.functions.common import (
    is_user_data,
    isnull,
    is_quble,
    is_screen,
    missing_val_by_dtype,
    is_rootlib,
)
from qubles.util.messenger import Messenger
from qubles.util import strftime
from qubles.util.hashing import make_hash
from qubles.io.util import libtypes
from qubles.io.util.table_utils import multi_col_info_custom_writer, drop_table
from qubles.io.util.cache.redis import Redis
from qubles.io.util.all_permutations import all_permutations
from qubles.io.util.recorder import DependencyRecorder
from qubles.io.util.libaddress import LibAddress
from qubles.io.util.classproperty import classproperty
from qubles.io.util.properties import (
    MissingProperty,
    NPModes,
    Properties,
    validate_property_metadata,
)
from qubles.io.snowflake.core import execute, generate_random_table_name
from qubles.io.snowflake.lib import lib_select
from qubles.io.snowflake.qbl import snowflake_get_quble_timestamp
from qubles.io.snowflake.env import (
    get_lib,
    set_lib,
    get_parent_id,
    snowflake_persist_quble,
    read_last_updated,
    sort_field_index,
)
from qubles.io.snowflake.prop import (
    sql_set_definition,
    definition_select,
    properties_field_select,
)
from api.models import Activity, Definition
from qubles.core.jinja.jinja_utils import dquote_dot
from qubles.util.logger import init_logger
from viewer.api.utils.app.users import get_usernames
from viewer.middleware.context_middleware import context_username

_logger = init_logger(__name__)

missing_date = missing_val_by_dtype("datetime")

DATALIB_PROPERTY_METADATA = {
    "access_grace": {
        "display_name": "Access Grace",
        "default_value": None,
        "allowed_values": (
            [None] + ["_".join(x) for x in all_permutations(["key", "read"])]
        ),
        "description": (
            "Content access error handing.\n"
            "key: Gracefully return None when an unsupported field is requested\n"
            "read: Gracefully return None on file access failure"
        ),
    },
    "access_mode": {
        "display_name": "Access Mode",
        "default_value": "virtual",
        "allowed_values": (
            [None]
            + ["_".join(x) for x in all_permutations(["virtual", "read", "build"])]
        ),
        "description": (
            "Controls how DataLibs access content.\n"
            "virtual: Fields can only be accessed from memory\n"
            "read: Fields can only be accessed from a file or memory\n"
            "build: Fields can only be accessed from a builder method or memory"
        ),
    },
    "auto_redefine": {
        "display_name": "Auto Redefine",
        "default_value": False,
        "validators": "bool",
        "description": (
            "Invokes a field 'redefine' event (i.e. resets a field's 'redefn_date' "
            "property) when certain field-specific properties (such as file I/O "
            "credentials) are changed"
        ),
    },
    "auto_register": {
        "display_name": "Auto Register",
        "default_value": False,
        "validators": "bool",
        "description": (
            "If an unregistered field is requested, register it instead of throwing "
            "an exception"
        ),
    },
    "dates_keyspace": {
        "display_name": "Dates Keyspace",
        "default_value": "Dates",
        "validators": "str",
        "description": ("Primary dates keyspace"),
    },
    "field_category": {
        "display_name": "Field Category",
        "allowed_values": [None, "data", "lib"],
        "description": ("Marks a field as a data object or library object."),
    },
    "fieldspace": {
        "display_name": "Fieldspace",
        "default_value": "Fields",
        "validators": "str",
        "description": (
            "When representing a DataLib as a Quble, 'fieldspace' will be the name "
            "given to the fields dimension"
        ),
    },
    "file_type": {
        "display_name": "File Type",
        "default_value": "SNOWFLAKE",
        "allowed_values": [
            None,
            "LIBADDRESS",
            "SNOWFLAKE",
            "QDB",
        ],
        "description": ("File format for read access"),
    },
    # NOTE: Please coordinate changes/additions to format options with files:
    # viewer/api/utils/quble/grid.py and viewer/app/src/utils/format.js
    "format": {
        "display_name": "format",
        "allowed_values": [
            None,
            "bool",
            "boolean",
            "str",
            "date",
            "fiscaldate",
            "int",
            "float",  # default-precision
            "number",  # low-precision
            "precise",  # high-precision
            "midprecise",  # mid-precision
            "percent",  # percent [1.0=100%]
            "percent100",  # percent [100.0=100%]
            "scientific",  # scientific notation
            "list",
            "local",
            "lcl",
            "prc",
            "rep",
            "dom",
        ],
        "validators": "str",
        "description": ("Formatting option for display"),
    },
    "filename": {
        "display_name": "Filename",
        "validators": "str",
        "description": ("Filename for read access"),
    },
    "id_keyspace": {
        "display_name": "ID Keyspace",
        "default_value": "BB_GID",
        "validators": "str",
        "description": ("Primary identifier keyspace"),
    },
    "native_property_mode": {
        "display_name": "Native Property Mode",
        "default_value": "secondary",
        "allowed_values": [None, "exclusive", "primary", "secondary"],
        "description": (
            "Determines how/when native property are invoked.\n"
            "None: Native property will not be consulted\n"
            "exclusive: Native property will be exclusively called\n"
            "primary: Native property called first, then fldspec property\n"
            "secondary: Fldspec property called first, then native property"
        ),
    },
    "path": {
        "display_name": "File Path",
        "validators": "str",
        "description": ("File path for read access"),
    },
    "store_mode": {
        "display_name": "Store Mode",
        "default_value": "virtual",
        "allowed_values": (
            [None]
            + ["_".join(x) for x in all_permutations(["virtual", "write", "touch"])]
        ),
        "description": (
            "Controls how DataLibs access content.\n"
            "virtual: Fields are only stored in memory\n"
            "write: Fields are only stored in file\n"
            'touch: Deprecated. Use "write".'
        ),
    },
}

FIELD_FRESH_TIME_OVERRIDE_FN_DICT = {
    "NOW": datetime.now,
    "TODAY_BEGIN": datetools.todayBegin,
    "WEEKDAY_BEGIN": datetools.recentWeekdayBegin,
    "WEEKDAY": datetools.recentWeekdayBegin,
    "PREV_WEEKDAY_BEGIN": datetools.prevWeekdayBegin,
    "PREV_WEEKDAY": datetools.prevWeekdayBegin,
    "MONDAY_BEGIN": datetools.recentMondayBegin,
    "TUESDAY_BEGIN": datetools.recentTuesdayBegin,
    "WEDNESDAY_BEGIN": datetools.recentWednesdayBegin,
    "THURSDAY_BEGIN": datetools.recentThursdayBegin,
    "FRIDAY_BEGIN": datetools.recentFridayBegin,
    "SATURDAY_BEGIN": datetools.recentSaturdayBegin,
    "SUNDAY_BEGIN": datetools.recentSundayBegin,
    "2ND_FRIDAY_BEGIN": datetools.recent2ndFridayBegin,
    "2ND_FRIDAY3_BEGIN": datetools.recent2ndFriday_Plus3DaysBegin,
    "3RD_FRIDAY_BEGIN": datetools.recent3rdFridayBegin,
    "3RD_FRIDAY10_BEGIN": datetools.recent3rdFriday_Plus10DaysBegin,
    "CBOM_BEGIN": datetools.recentCBOMBegin,
    "CBOM1_BEGIN": datetools.recentCBOM_Plus1DayBegin,
    "CBOM2_BEGIN": datetools.recentCBOM_Plus2DaysBegin,
    "CBOM3_BEGIN": datetools.recentCBOM_Plus3DaysBegin,
    "CBOM4_BEGIN": datetools.recentCBOM_Plus4DaysBegin,
    "CBOM5_BEGIN": datetools.recentCBOM_Plus5DaysBegin,
    "CBOM10_BEGIN": datetools.recentCBOM_Plus10DaysBegin,
    "CBOM15_BEGIN": datetools.recentCBOM_Plus15DaysBegin,
    "CBOM20_BEGIN": datetools.recentCBOM_Plus20DaysBegin,
    "CEOM_BEGIN": datetools.recentCEOMBegin,
    "CEOM": datetools.recentCEOMBegin,
    "EOM": datetools.recentEOMBegin,
    "W@MON": datetools.recentMondayBegin,
    "W@TUE": datetools.recentTuesdayBegin,
    "W@WED": datetools.recentWednesdayBegin,
    "W@THU": datetools.recentThursdayBegin,
    "W@FRI": datetools.recentFridayBegin,
    "W@SAT": datetools.recentSaturdayBegin,
    "W@SUN": datetools.recentSundayBegin,
}


class DataLib(object):
    """Core data management class.

    A DataLib is a collection of addressable objects. The DataLib class is the
    primary mechanism for managing content I/O. Every addressable object in the data
    environment that is not a data object (i.e. Quble) is a DataLib or subclass of
    DataLib.

    DataLibs are comprised of two main components: **properties** and **fields**.

    **Fields** are named objects in the collection. For example, a DataLib may have
    two data fields: "Open Prices" and "Close Prices". The data objects associated
    with these fields can be access with the syntax ``lib['Open Prices']`` and
    ``lib['Close Prices']``.

    The resolution of the objects is determined by the DataLib's **properties**.
    Properties contain all the configuration for a DataLib. Ultimately, properties are
    just a flexible key-value metadata store. The behavior of each property is
    defined within the DataLib's logic. For example, one property on the base DataLib
    class is *file_type*. A user can execute ``lib.set_property('file_type', 'CSV')``
    to specify that the DataLib should resolve its objects by reading CSV files.

    :type fields: iterable of strings
    :param fields:
        List of fields to register on the library.

    :type default_properties: dict
    :param default_properties:
        Default property assignments. Uses the structure:
        ``default_properties[property name] -> property value.``

    :type fldspec_properties: dict
    :param fldspec_properties:
        Field-specific property assignments. Uses the structure:
        ``fldspec_properties[property name][field name] -> property value.``

    :type property_metadata: dict
    :param property_metadata:
        Metadata about properties supported by this DataLib. Uses the structure:
        ``property_metadata[property name][metadata name] -> metadata value``

    :type address: LibAddress
    :param address:
        Instantiate the DataLib with an address.

    :type time_stamp: str or datetime
    :param time_stamp:
        The update timestamp of the DataLib. Used for freshness checking. Specify
        a specific datetime or use "now" to set the timestamp to the current time.

    :param \*\*kwargs:
        Alternative way to pass in default property values. For example, instead
        of ``DataLib(default_properties={'file_type': 'CSV'}`` you can do
        ``DataLib(file_type='CSV')``. Note that values defined in the
        *default_properties* keyword argument take precedence.

    """

    ALL_FIELDS_KEYWORD = "<all>"
    REDIS_RESIDENCE_TIMESTAMP_KEY = "Residence last evaluated"
    REDIS_FILE_TIMESTAMP_KEY = "File last evaluated"
    _dependency_recorders = {}
    _dependency_recorders_lock = RLock()

    def __getstate__(self):
        _logger.critical("DataLib serialized")

    @classproperty
    def dependency_recorder(cls):
        thread_id = current_thread().ident

        with cls._dependency_recorders_lock:
            if thread_id not in cls._dependency_recorders:
                cls._dependency_recorders[thread_id] = DependencyRecorder()

            dep_recorder = cls._dependency_recorders[thread_id]
        return dep_recorder

    CONTROL_PROPERTY_LIST = []

    FIELD_PROPERTY_PROHIBITION_LIST = (
        "auto_register",
        "fieldspace",
    )

    DEFAULT_PROPERTY_PROHIBITION_LIST = (
        "field_defn_time",
        "field_dependencies",
        "field_fresh_time",
        "field_post_time",
        "field_time_stamp",
        "is_fresh_field",
        "is_stale_field",
    )

    NATIVE_PROPERTY_NONSTORE_LIST = (
        "address_as_file_path",
        "field_fresh_time",
        "field_post_time",
        "field_time_stamp",
        "is_fresh_field",
        "is_stale_field",
        "property_dep_hash_current",
        "field_dependencies_current",
    )

    NATIVE_PROPERTY_AUTOSTORE_LIST = (
        "field_dependencies",
        "procurement_duration_last",
        "property_dep_dict_last",
        "property_dep_hash_last",
        "property_dependencies",
        "value_type_last",
        "value_fx_last",
        "field_valuespaces_last",
        "field_col_types_last",
        "field_val_fxs_last",
    )

    REDEFINE_PROPERTY_LIST = (
        "file_type",
        "filename",
        "path",
    )

    # Hard coded properties that should not be considered native properties, even if they are callable
    NON_NATIVE_PROPERTY_LIST = ["field_type"]

    # The following properties do not affect results and therefore do not need to be recorded.
    PROPERTY_RECORDER_PROHIBITION_LIST = (
        "auto_redefine",
        "auto_register",
        "field_category",
        "field_dependencies",
        "field_fresh_time_override",
        "procurement_duration_last",
        "property_dep_hash_current",
        "property_dep_dict_last",
        "property_dep_hash_last",
        "property_dependencies",
        "field_dependencies_current",
        "description",
        "field_type",
    )

    PROCUREMENT_PROPERTY_RECORDINGS_LIST = {
        "property_dependencies",
        "procurement_duration_last",
        "property_dep_dict_last",
        "property_dep_hash_last",
        "field_dependencies",
        "field_dependencies_current",
        "property_dep_hash_current",
        "dependent_fields",
        "value_type_last",
        "value_fx_last",
        "field_valuespaces_last",
        "field_col_types_last",
        "field_val_fxs_last",
        "build_tree_last",
    }

    def __init__(
        self,
        fields=None,
        default_properties=None,
        fldspec_properties=None,
        property_metadata=None,
        address=None,
        time_stamp="now",
        **kwargs,
    ):
        # Prevent property validation when loading from source, otherwise an existing
        # bad value will prevent the user from loading the lib and fixing the value.
        # Piggyback on the default properties because many DataLib subclasses don't
        # pass **kwargs to super(), and it would be a lot of work to update every
        # DataLib subclass to take a ``validate_properties`` keyword argument.
        if default_properties and "validate_properties" in default_properties:
            validate_properties = default_properties["validate_properties"]
            del default_properties["validate_properties"]
        else:
            validate_properties = True

        # Newer snowflake libs store field ids as well as references, and pass this in as a dict
        if isinstance(fields, dict):
            self.add_fields(list(fields.keys()))
            self.field_to_id_map = fields
        else:
            self.add_fields(fields)
            self.field_to_id_map = {}

        # Collect property metadata
        # --------------------------
        all_property_metadata = self.initial_property_metadata() or {}
        property_metadata = property_metadata or {}

        validate_property_metadata(all_property_metadata)
        validate_property_metadata(property_metadata)

        all_property_metadata.update(deepcopy(property_metadata))

        # Collect property values
        # ------------------------
        all_default_properties = {}
        all_fldspec_properties = {}

        all_default_properties["timestamp"] = datetime.now()
        for prop_name, metadata in all_property_metadata.items():
            if "default_value" in metadata:
                all_default_properties[prop_name] = metadata["default_value"]

        for prop_name, prop_value in kwargs.items():
            if isinstance(prop_value, dict):
                all_fldspec_properties[prop_name] = prop_value
            else:
                all_default_properties[prop_name] = prop_value

        for prop_name, prop_value in (default_properties or {}).items():
            all_default_properties[prop_name] = prop_value

        for prop_name, field_props in (fldspec_properties or {}).items():
            if prop_name not in all_fldspec_properties:
                all_fldspec_properties[prop_name] = {}

            for field_name, prop_value in field_props.items():
                all_fldspec_properties[prop_name][field_name] = prop_value

        # If we have a default builder assignment, default to field_category="data"
        if "builder" in all_default_properties:
            all_default_properties["field_category"] = "data"

        # Instantiate the Properties object
        # ----------------------------------
        self.properties = Properties(
            self,
            self.NATIVE_PROPERTY_NONSTORE_LIST,
            self.NATIVE_PROPERTY_AUTOSTORE_LIST,
            self.NON_NATIVE_PROPERTY_LIST,
            self.DEFAULT_PROPERTY_PROHIBITION_LIST,
            default_properties=all_default_properties,
            fldspec_properties=all_fldspec_properties,
            native_properties=self.initial_native_properties(),
            property_metadata=all_property_metadata,
            validate=validate_properties,
        )

        # Initialize remaining attributes
        # --------------------------------
        self._reflibs_cache = None
        self.address = address
        self._name = None
        if time_stamp == "now":
            self.time_stamp = datetime.now()
        else:
            self.time_stamp = time_stamp

        # Dynamic builder method definitions are exec'd into the ``dynamic_builders``
        # dict and the definitions themselves are stored in ``dynamic_builder_defns``
        self.dynamic_builders = {}
        self.dynamic_builder_defns = {}

        # Flags this DataLib as temporary. Temporary libs may have an address, but
        # they are blocked from being committed to residence or file. The only way
        # to access a temporary lib is to have access to the in-memory object.
        self.is_temporary = False
        self.residence = {}

        # Temporary attribute for attach_or_sync_shared_data_props pending rootlib refactor
        self.synchronized = False

    def initial_property_metadata(self):
        return deepcopy(DATALIB_PROPERTY_METADATA)

    def merge_property_metadata(self, metadata1, metadata2):
        if not isinstance(metadata1, dict):
            return metadata2

        if not isinstance(metadata2, dict):
            return metadata1

        new_metadata = deepcopy(metadata1 or {})

        for prop_name, metadata in (metadata2 or {}).items():
            if prop_name not in new_metadata:
                new_metadata[prop_name] = {}

            for metadata_name, metadata_value in metadata.items():
                new_metadata[prop_name][metadata_name] = metadata_value

        return new_metadata

    def initial_native_properties(self):
        return {
            "is_derived_field": self.is_derived_field,
            "is_lib_field": self.is_lib_field,
            "is_reflib_field": self.is_reflib_field,
            "field_dependencies": self.field_dependencies,
            "property_dependencies": self.property_dependencies,
            "address_as_file_path": self.address_as_file_path,
            "property_dep_hash_current": self.property_dep_hash_current,
            "field_dependencies_current": self.field_dependencies_current,
            "field_keyspaces_last": self.field_keyspaces_last,
            "field_val_types_last": self.field_val_types_last,
            "field_key_types_last": self.field_key_types_last,
        }

    @property
    def name(self):
        if self._name:
            return self._name
        if not self.address or len(self.address) == 0:
            raise ValueError("Ill defined name for library, no name/address assigned.")
        else:
            return self.address[-1]

    @name.setter
    def name(self, value):
        self._name = value

    @property
    def num_frames(self):
        return self.properties.num_frames

    def add_frame(self):
        self.properties.add_frame()

    def pop_frame(self):
        self.properties.pop_frame()

    def pop_all_frames(self):
        self.properties.pop_all_frames()

    def unfreeze_property(self, property_name):
        self.properties.unfreeze_property(property_name)

    def unfreeze_all_properties(self):
        self.properties.unfreeze_all_properties()

    def id_to_fieldref(self, id):
        return next(
            (key for key, value in self.field_to_id_map.items() if value == id), None
        )

    @property
    def has_dynamic_builders(self):
        """Returns True if the library has at least one field using a dynamic
        builder method. A dynamic builder method is a builder whose definition
        is stored in a file outside the library.
        """
        return len(self.fields()) > 0 and (
            self.properties.is_default_property("definition")
            or self.properties.is_fldspec_property("definition")
        )

    def update_timestamp(self):
        self.set_property("last_updated", datetime.now())

    def add_dynamic_builder(self, fieldref, code="<defn>"):
        """Add the dynamic builder method for the given field.

        Dynamic builder methods are builders whose definitions are stored in filesoutside the library.
        These methods are added using the following two properties:
                `builder`: the name of the builder method
                `definition`: the file path of the external definition file

        If the definition for this field has not changed and the builder has already been added, this method will be a no-op.

        :param fieldref: The name of the field
        :type fieldref: str

        :param code (OPTIONAL): The code to be added dynamically
                                DEFAULTS to <defn>, in which case we extract the corresponding defn from persistence
        :type code: str
        """
        if not self.has_dynamic_builders:
            return

        builder = self.get_property(
            "builder", fieldref, grace=True, resolve_templates=True
        )

        if code == "<defn>":
            defn, _ = self.get_definition(fieldref)
        else:
            defn = code

        if not builder or not defn:
            raise BuilderError("Cannot locate the builder/defn")

        # NOTE: Eventually we can pare-down imports to only Quble-related modules
        setup = [
            "import datetime",
            "from collections import defaultdict, OrderedDict",
            "",
            "import numpy as np",
            "",
            "from qubles.core.classes.daterange import DateRange, generate_range, join_indexes, join_multi_index, indexfromstr, indexfromstrlist, make_index",
            "from qubles.core.classes.timer import Timer",
            "from qubles.core.classes.index import Index",
            "from qubles.core.quble import Quble",
            "from qubles.io.base.rootlib import RootLib",
            "",
            "root_lib = RootLib()",
            "",
            "root_lib.add_frame()",
            "",
        ]

        teardown = [
            "root_lib.pop_frame()",
        ]

        new_defn = "\n".join(
            [f"def {builder}(self, field):"]
            + [f"    {line}" for line in setup]
            + [""]
            + ["    try:"]
            + ["##### ____BEGIN_USER_DEFINITION____ #####"]
            + [f"        {line}" for line in defn.split("\n")]
            + ["##### ____END_USER_DEFINITION_____ #####"]
            + ["    finally:"]
            + [f"        {line}" for line in teardown]
        )

        current_defn = self.dynamic_builder_defns.get(builder)

        if current_defn and new_defn == current_defn:
            return

        try:
            exec(new_defn, self.dynamic_builders)
        except Exception:
            _, exc_obj, _ = exc_info()[:]
            # These are user-introduced field definitions, so if there is a syntax
            # error make sure it doesn't bring the while library crashing down.
            new_defn = "\n".join(
                [
                    f"def {builder}(self):",
                    "    from qubles.io.base.datalib import DynamicBuilderSyntaxError",
                    '    raise DynamicBuilderSyntaxError("""{}""", {})'.format(
                        new_defn.replace('"', r"\""),
                        exc_obj.lineno,
                    ),
                ]
            )

            exec(new_defn, self.dynamic_builders)

        self.dynamic_builder_defns[builder] = new_defn
        method = MethodType(self.dynamic_builders[builder], self)

        setattr(self, builder, method)

    def add_fields(self, fields):
        if fields is None:
            field_index_add = Index()
        elif isinstance(fields, list) or isinstance(fields, tuple):
            field_index_add = make_index(fields)
        elif isinstance(fields, Index) or isinstance(fields, DateRange):
            field_index_add = fields.copy()
        else:
            raise TypeError("Bad fields arg: list, tuple (of strings) expected")

        if not hasattr(self, "field_index") or (self.field_index is None):
            self.field_index = field_index_add
        if (field_index_add is not None) and (len(field_index_add) > 0):
            self.field_index = self.field_index.union(field_index_add)

    def resolve_field_dependencies(self):
        """Resolve field dependencies to LibAddress."""
        if (
            not self.properties.is_fldspec_property("field_dependencies")
            or len(self.properties.fields_with_property("field_dependencies")) == 0
        ):
            return

        for field1 in self.properties.fields_with_property("field_dependencies"):
            original_dependencies = self.properties.get_fldspec_property(
                "field_dependencies", field1
            )
            # list/tuple original dependencies
            # -------------------------------------
            if isinstance(original_dependencies, (list, tuple)):
                converted_dependencies = []
                for dependency1 in original_dependencies:
                    if dependency1 is None:
                        continue

                    # Convert string to LibAddress
                    # -------------------------------
                    if isinstance(dependency1, str):
                        if dependency1[0:1] == "*":
                            dependency1 = LibAddress(dependency1[1:])
                        elif dependency1[0:5].upper() == "ROOT:":
                            dependency1 = LibAddress(dependency1[5:])
                        else:
                            dependency1 = LibAddress(dependency1, base_domain=self)
                    # Convert list/tuple to LibAddress
                    # -----------------------------------
                    elif isinstance(dependency1, list) or isinstance(
                        dependency1, tuple
                    ):
                        if (len(dependency1) > 0) and isinstance(dependency1[0], str):
                            if dependency1[0][0:1] == "*":
                                dependency1[0] = dependency1[1:]
                                dependency1 = LibAddress(dependency1)
                            elif dependency1[0][0:5].upper() == "ROOT:":
                                dependency1[0] = dependency1[5:]
                                dependency1 = LibAddress(dependency1)
                            else:
                                dependency1 = LibAddress(dependency1, base_domain=self)
                        else:
                            dependency1 = LibAddress(dependency1, base_domain=self)

                    converted_dependencies.append(dependency1)

                self.properties.set(
                    "field_dependencies", converted_dependencies, field=field1
                )

    def copy(self, virtual_only=False, new_fields=None, retain_address=False):
        """Returns a deep copy of a library."""
        from qubles.io.base.rootlib import RootLib

        if isinstance(self, RootLib):
            raise TypeError("cannot copy the RootLib")

        if retain_address:
            new_lib = self.__class__(
                address=self.address
            )  # <-- Preserves potential parent class
        else:
            new_lib = self.__class__()  # <-- Preserves potential parent class

        # Initialize new_lib.field_index...
        # --------------------------------------
        if new_fields is None:
            new_lib.field_index = deepcopy(self.field_index)
        elif isinstance(new_fields, Index) or isinstance(new_fields, DateRange):
            new_lib.field_index = deepcopy(new_fields)
        elif isinstance(new_fields, tuple) or isinstance(new_fields, list):
            new_lib.field_index = make_index(
                new_fields, freq_hint=self.field_index.freq
            )
        elif np.isscalar(self.new_fields):
            new_lib.field_index = make_index(
                [new_fields], freq_hint=self.field_index.freq
            )
        else:
            raise TypeError(
                "Invalid new_fields arg: None or Index or DateRange or tuple or list expected"
            )

        # Build default properties...
        # -----------------------------
        new_lib.properties = self.properties.copy(new_lib)

        # Update remaining DataLib properties...
        # -----------------------------------------
        if not retain_address:
            new_lib.address = None  # <-- DO NOT WANT TO INHERIT self.address

        if virtual_only:
            new_lib.set_property("path", None, fieldref=self.ALL_FIELDS_KEYWORD)
            new_lib.set_property("filename", None, fieldref=self.ALL_FIELDS_KEYWORD)
            new_lib.set_property("access_mode", None, fieldref=self.ALL_FIELDS_KEYWORD)
            new_lib.set_property(
                "store_mode", "virtual", fieldref=self.ALL_FIELDS_KEYWORD
            )

        # Store new_virtual_values when appropriate...
        # -----------------------------------------------
        if virtual_only:
            for field in new_lib.field_index:
                if field in self.field_index:
                    residence_key = (field, self.property_dep_hash_current(field))
                    new_lib._assign_to_residence(residence_key, self.get(field).copy())

        # Establish new_lib's time_stamp accordingly...
        # -----------------------------------------------
        if self.field_index == new_lib.field_index:
            new_lib.time_stamp = self.time_stamp
        else:
            new_lib.touch()

        return new_lib

    def _exists_in_residence(self, residence_key) -> bool:
        return residence_key in self.residence

    def _assign_to_residence(self, residence_key, value):
        from qubles.io.util.multi_build_lib import MultiBuildLib

        if residence_key is None:
            raise ValueError(
                f"Failure assigning value to residence for library: {self}. Residence key was None"
            )

        self.residence[residence_key] = value

    # ================================= Field Methods =================================

    @property
    def registry(self):
        return self.fields()

    def dir(self, max_fields=None):
        from qubles.io.base.rootlib import RootLib

        output = []
        field_no = None
        index_delimiter = RootLib().get_control("index_delimiter")
        for field_no, field in enumerate(self.field_index):
            if max_fields is None:
                pass
            elif field_no >= max_fields:
                output.append("...")
                break
            output_add = f"{field}"
            if self.is_lib_field(field):
                output_add += index_delimiter
            output.append(output_add)

        return "\n".join(output)

    def remove_derived_field_files(self):
        """
        Remove any files associated with the derived fields (if file(s) exist)
        """
        return self.remove_field_file(self.derived_fields())

    def remove_field_file(self, field):
        """
        Remove any files associated with the specified field(s) (if file(s) exist)
        """
        if field is None:
            return
        elif type(field) in (list, tuple, Index, DateRange):
            pass
        else:
            field = [field]

        for fld1 in field:
            # Validate field
            # ------------------
            if fld1 not in self.field_index:
                continue

            # Get path
            # -----------
            path = normalize(self.get_property("path", fld1, grace=True))
            if path is None:
                continue

            # Get filename
            # --------------
            filename = self.get_property("filename", fld1, grace=True)
            if filename is None:
                continue

            # Try to remove the speciifed field's file (if it exists)
            # --------------------------------------------------------
            full_filename = os.path.join(path, filename)
            if os.path.isfile(full_filename):
                os.remove(full_filename)

    def derived_fields(
        self, inclusions=None, exclusions=None, honor_none_filtering=True
    ):
        # Update inclusions...
        if inclusions is None:
            inclusions2 = {}
        elif not isinstance(inclusions, dict):
            raise TypeError("Invalid inclusions arg: dict expected")
        else:
            inclusions2 = deepcopy(
                inclusions
            )  # <-- Do not want to contaminate inclusions list on exit

        # Add to inclusions: 'is_derived_field' = True
        # ------------------------------------------------------
        inclusions2["is_derived_field"] = True

        return self.fields(
            inclusions=inclusions2,
            exclusions=exclusions,
            honor_none_filtering=honor_none_filtering,
        )

    def source_fields(
        self, inclusions=None, exclusions=None, honor_none_filtering=True
    ):
        # Update exclusions...
        if exclusions is None:
            exclusions2 = {}
        elif not isinstance(exclusions, dict):
            raise TypeError("Invalid exclusions arg: dict expected")
        else:
            exclusions2 = deepcopy(
                exclusions
            )  # <-- Do not want to contaminate inclusions list on exit

        # Add to exclusions: 'is_derived_field' = True
        # ------------------------------------------------------
        exclusions2["is_derived_field"] = True

        return self.fields(
            inclusions=inclusions,
            exclusions=exclusions2,
            honor_none_filtering=honor_none_filtering,
        )

    def touch_derived_field_defn_times(
        self, inclusions=None, exclusions=None, honor_none_filtering=True
    ):
        """
        Sets 'field_defn_time' property of all derived fields to current time
        This exercise has the effect of indirectly making all derived fields 'stale'
        """
        derived_fields = self.derived_fields(
            inclusions=inclusions,
            exclusions=exclusions,
            honor_none_filtering=honor_none_filtering,
        )
        new_defn_time = datetime.now()
        for field in derived_fields:
            self.set_property("field_defn_time", new_defn_time, field)

    def _reflibs(self, recursive=False):
        """
        Generator that iterates through reflibs
        (returns a sub-library reference at each iteration)
        """
        # Loop through reflibs
        # -----------------------
        for field in self.field_index:
            field_category = self.get_property(
                "field_category", field, grace=True, suppress_recording=True
            )  # <-- Do not want the reflib establishment to affect property recordings
            if field_category == "data":
                pass  # <-- Trying to handle case when field_category is None or field_category != 'data'
            elif self.is_reflib_field(field):
                lib = self.get(field, suppress_recording=True, log=False)
                yield lib
                if recursive:
                    for sublib in lib.reflibs(recursive=recursive):
                        yield sublib

    def sublibs(self, recursive=False):
        """
        Generator that iterates through sublibs
        (returns a sub-library reference at each iteration)
        """
        # Loop through sublibs
        # -----------------------
        for field in self.field_index:
            if self.is_lib_field(field):
                lib = self.get(field)
                yield lib
                if recursive:
                    for sublib in lib.sublibs(recursive=recursive):
                        yield sublib

    def change_sublibs_default_property(
        self,
        property_name,
        new_property_value,
        initiate_property=False,
        recursive=False,
    ):
        """
        Changes property value for the specified property name within all embedded sub-libraries
        NOTE: Does not alter property values for the current library!!

        initiate_property=False: Does not initiate a new property name within any sub-library that was not originally present
        initiate_property=True: Will initiate new property name within sub-libraries if not originally persent
        """
        for sublib in self.sublibs(recursive=recursive):
            if initiate_property or sublib.has_property(property_name):
                sublib.set_property(property_name, new_property_value)

    def lib_search(self, field, recursive=True, search_root=False, grace=False):
        """
        Identifies the (sub)library that contains the specifies field.
        """
        from qubles.io.base.rootlib import RootLib

        root_lib = RootLib()
        if field in self.field_index:
            return self

        for lib in self.sublibs(recursive=recursive):
            if field in lib.field_index:
                return lib

        if (self != root_lib) and search_root:
            for lib in root_lib.sublibs(recursive=True):
                if field in lib.field_index:
                    return lib

        if grace:
            return None
        else:
            raise ValueError(f"field:{field} not found in sublibs")

    def content(self, recursive=False):
        """
        Generator that iterates through data content within a library
        """
        for terminus in self.termini(recursive=recursive):
            yield terminus.pickup()

    def termini(self, recursive=False):
        """
        Generator that iterates thorugh all the (sub)fields
        and returns the associated LibAddress at each iteration
        """
        # Loop through fields
        # -----------------------
        for field in self.field_index:
            new_address = LibAddress(field, self)
            field_category = new_address.pickup(request_type="field_category")
            if field_category is None:
                pass
            elif field_category != "lib":
                yield new_address
            elif recursive:
                lib = new_address.pickup()
                for terminus in lib.termini(recursive=recursive):
                    yield terminus

    def fields(self, inclusions=None, exclusions=None, honor_none_filtering=True):
        """
        **Provides iterable container of fields that satisfy the inclusion/exclusion filters provided...**

        :param inclusions:  dictionary, where the keys indicate property names to be filtered on,
                            and dictionary values identify associated property values for fields to be INCLUDED
        :type inclusions: dict or None

        :param exclusions:  dictionary, where the keys indicate property names to be filtered on,
                            and dictionary values identify associated property values for fields to be EXCLUDED
        :type exclusions: dict or None

        :param honor_none_filtering: True (default): filter values of None are supported/applied
                                     False: filter values of None are NOT unsupported/not applied/ignored

        :type honor_none_filtering: bool
        """
        # ------------------------------------------
        # Handle trivial (unconditional) case...
        # ------------------------------------------
        if (inclusions is None) and (exclusions is None):
            return self.field_index  # <-- shallow copy OK
        # --------------------------------
        # Handle conditional case...
        # --------------------------------
        else:
            all_properties = self.property_list()
            field_list = []
            for field in self.field_index:
                filtering_state = (
                    True  # <-- Initialize filtering_state for current field
                )

                # Apply inclusion filtering (if applicable)...
                # ------------------------------------------------
                if (inclusions is not None) and (len(list(inclusions.keys())) > 0):
                    for inclusion_property, inclusion_value in inclusions.items():
                        if (inclusion_value is None) and not honor_none_filtering:
                            continue

                        if inclusion_property == "field":
                            property_value = field
                        elif (inclusion_property not in all_properties) and not hasattr(
                            self, inclusion_property
                        ):
                            raise ValueError(
                                "Unsupported (and non-native) inclusion property/key: {0}".format(
                                    inclusion_property
                                )
                            )
                        else:
                            property_value = self.get_property(
                                inclusion_property, field, grace=True
                            )

                        if isinstance(inclusion_value, list) or isinstance(
                            inclusion_value, tuple
                        ):
                            if property_value not in inclusion_value:
                                filtering_state = False

                        elif property_value != inclusion_value:
                            filtering_state = False

                        if not filtering_state:
                            break

                # Apply exclusion filtering (if applicable)...
                # ---------------------------------------------
                if (
                    filtering_state
                    and (exclusions is not None)
                    and (len(list(exclusions.keys())) > 0)
                ):
                    for exclusion_property, exclusion_value in exclusions.items():
                        if (exclusion_value is None) and not honor_none_filtering:
                            continue

                        if exclusion_property == "field":
                            property_value = field
                        elif (exclusion_property not in all_properties) and not hasattr(
                            self, exclusion_property
                        ):
                            raise ValueError(
                                "Unsupported (and non-native) exclusion property/key: {0}".format(
                                    exclusion_property
                                )
                            )
                        else:
                            property_value = self.get_property(
                                exclusion_property, field, grace=True
                            )

                        if isinstance(exclusion_value, list) or isinstance(
                            exclusion_value, tuple
                        ):
                            if property_value in exclusion_value:
                                filtering_state = False

                        elif property_value == exclusion_value:
                            filtering_state = False

                        if not filtering_state:
                            break

                # Append candidate field to list
                # (if all filtering logic was satisfied)
                # ----------------------------------------
                if filtering_state:
                    field_list.append(field)

            # --------------------------------------------------------------------
            # NOTE: At this point, fields is a list of conditional fields
            # --------------------------------------------------------------------
            return make_index(field_list, freq_hint=self.field_index.freq)

    def build_all(self, inclusions=None, exclusions=None, honor_none_filtering=True):
        """
        Iteratively calls self.get(<field>) for all fields that pass the specified filters
        """
        fields = self.derived_fields(
            inclusions=inclusions,
            exclusions=exclusions,
            honor_none_filtering=honor_none_filtering,
        )
        for field in fields:
            self.get(field)

    def contains_field(self, field, recursive=False):
        if field in self.field_index:
            return True
        elif recursive:
            address = LibAddress(field)
            if address.min_field is not None:
                return True
            else:
                return False
        else:
            return False

    @property
    def local_address(self):
        if not hasattr(self, "address") or (self.address is None):
            return None
        else:
            return self.address

    @property
    def parent_lib(self):
        return None if self.address is None else self.address.min_domain

    def build_method(self, field, build_access_check=True):
        """
        Searches the library (and possibly up the predecessor tree)
        to find the most recent build function and the child_hyper from the build function
        Returns tuple (build_method,child_hyper)
        """
        # ----------------------------------------------------------------------------------------------
        # Develop notes:
        #   1) Turned off traverse feature & arg (can be handled as a default property of the library)
        #   2) Turned off the aggr_op builder logic (this functionality now handled within MultiFactorModel class
        #   ==> This method now only returns the associated build_method (no longer returns a tuple)
        #   ==> This method now no longer requests 'aggrop' or 'fieldspace' properties
        # ----------------------------------------------------------------------------------------------
        lib = self
        builder = None

        # Handle trivial case
        # ----------------------
        if (
            lib is None
            or field is None
            or (build_access_check and (lib.build_access(field, grace=True) != True))
        ):
            return builder

        defn, def_type = self.get_definition(field)
        if def_type == "python":
            self.add_dynamic_builder(field, code=defn)

        # First, consult 'builder' property (if available)
        # ---------------------------------------------------
        builder_property = lib.get_property(
            "builder", field, grace=True, resolve_templates=True
        )

        # Otherwise, see if build property ifs a dictionary (keyed by field)
        # ------------------------------------------------------------------
        if (
            (builder_property is None)
            and hasattr(lib, "builders")
            and isinstance(lib.builders, dict)
        ):
            if isinstance(lib.builders, defaultdict) or (field in lib.builders):
                builder_property = lib.builders[field]

        # Process builder property (if available)
        # -------------------------------------------
        if builder_property is not None:
            if isinstance(builder_property, str):
                if not hasattr(lib, builder_property):
                    raise AttributeError(
                        "Error resolving lib.get_property('builders',{0})={1} (string)...No associated method: lib.{2}".format(
                            field, builder_property, builder_property
                        )
                    )
                else:
                    builder = lib.__getattribute__(builder_property)
            elif hasattr(builder_property, "__call__"):
                builder = builder_property
            else:
                raise AttributeError(
                    "Error resolving lib.get_property('builders',{0})...string or ...No associated method: lib.{1}".format(
                        field, builder_property
                    )
                )
        # Otherwise, see if there is a (callable) method
        # with with a name that matches the field
        # ----------------------------------------------
        elif isinstance(field, str) and hasattr(lib, field):
            builder = lib.__getattribute__(field)

        return builder

    def num_fields(self, inclusions=None, exclusions=None, honor_none_filtering=True):
        return len(
            self.fields(
                inclusions=inclusions,
                exclusions=exclusions,
                honor_none_filtering=honor_none_filtering,
            )
        )

    def show_fields(
        self,
        inclusions=None,
        exclusions=None,
        honor_none_filtering=True,
    ):
        for field in self.fields(
            inclusions=inclusions,
            exclusions=exclusions,
            honor_none_filtering=honor_none_filtering,
        ):
            print(field)

    def is_derived_field(self, fieldref, build_access_check=True):
        # Handle list/tuple case...
        # ----------------------------
        if isinstance(fieldref, list) or isinstance(fieldref, tuple):
            is_derived = []
            for fref in fieldref:
                is_derived.append(self.is_derived_field(fieldref=fref))
            return is_derived

        # Procure lib that contains fieldname (may be an embedded lib)
        # --------------------------------------------------------------
        lib = self
        field = fieldref
        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        # Return the derived state...
        # -----------------------------
        result = False
        if (
            self.get_property("file_type", field) == "LIBADDRESS"
            or self.get_property("field_category", field, grace=True) == "lib"
        ):
            return result

        builder = lib.build_method(field, build_access_check=build_access_check)
        if builder is not None:
            result = True

        return result

    def is_linked_field(self, fieldref):
        return (
            self.get_property("access_mode", fieldref, grace=True) is not None
        ) and (self.get_property("file_type", fieldref, grace=True) == "LIBADDRESS")

    def is_data_field(self, fieldref):
        is_lib_result = self.is_lib_field(fieldref)
        if is_lib_result is None:
            return None
        else:
            return not is_lib_result

    def is_lib_field(self, fieldref, **kwds_args):
        """
        Tests whether a fieldname's category == 'lib'  <==> is this field an instance of a DataLib
        [Could be a library derived from DataLib Class]

        Output: True/False/None (None indicating that the related field value is None)
        """
        result = False
        field_category = self.get_property("field_category", fieldref)
        if field_category not in (None, "data", "lib"):
            raise ValueError(
                f"Unsupported field_category: {field_category}, fieldref: {fieldref}"
            )
        elif field_category is None:
            result = None
        elif field_category == "data":
            result = False
        else:
            result = True

        return result

    def is_reflib_field(self, fieldref, **kwds_args):
        """
        Tests whether a fieldname is an instance of a RefLib
        [Could be a library derived from RefLib Class]

        Output: True/False/None (None indicating that the related field value is None)
        """
        from qubles.io.ref.reflib import RefLib

        instance_sample = self.get(
            fieldref, request_type="field_instance_sample", log=False
        )  # <-- Adopting this approach to align/leverage logic in DataLib.get() method
        result = False
        if instance_sample is None:
            result = None
        elif isinstance(instance_sample, RefLib):
            result = True
        else:
            result = False

        return result

    # ================================== Link Methods =================================

    def get_reflibs(self, recursive=False):
        """
        If necessary, initializes self._reflibs_cache as a list of underlying RefLibs (or fields for reflibs)
        Then returns the self._reflibs_cache on all subsequent calls
        """
        if self._reflibs_cache is None:

            # Set up a temporary variable to store reference libraries retrieved from properties
            reflibs = []
            for reflib in self._reflibs(recursive=recursive):
                reflibs.append(reflib)

            # If we were able to pull reference libraries from properties, update the cache
            if len(reflibs) > 0:
                self._reflibs_cache = reflibs
            # If we were not able to pull any reference libraries from properties,
            # return an empty list, but don't cache the result.
            else:
                return reflibs

        return self._reflibs_cache

    def get_src_keyspace(self, src_keyspace):
        """
        Given a scalar, list or tuple of candidate src_keyspace(s), returns tuple pair: (src_ks, reflib)
        specifying applicable src_ks (scalar) and reflib (Reference Library) where it was found respectively
        ----------------------------------------------------------------------------------------------------------------------
        If no src_keypace is supported by any registered Reference Library, the method will return (None, None)
        """
        from qubles.io.ref.reflib import RefLib

        # NOTE: get_reflibs() will take advantage of self._reflibs_cache
        for reflib in self.get_reflibs():
            if (reflib is not None) and not isinstance(reflib, DataLib):
                # Double check validity of reflib
                if reflib in self.field_index:
                    reflib = self.get(reflib, suppress_recording=True, log=False)
                else:
                    _logger.warning(f"reflib:{reflib} not present in self.field_index")

            if reflib is None:
                continue
            elif not isinstance(reflib, RefLib):
                raise RefLibExpectedError

            # NOTE: .get_src_keyspace() is overloaded in RefLib Class...
            # see <RefLib>.get_src_keyspace() for functionality
            src_ks = reflib.get_src_keyspace(src_keyspace)
            if src_ks is not None:
                return (src_ks, reflib)

        return (None, None)

    def ks_maps_with_reflib(self, src_keyspace, tgt_keyspace=None):
        from qubles.io.ref.reflib import RefLib

        # Check self first (if applicable)
        if isinstance(self, RefLib):
            keyspace_maps = self.ks_maps(src_keyspace, tgt_keyspace)
            if keyspace_maps is not None:
                return keyspace_maps, self

        # Then check fields of self
        # NOTE: get_reflibs() will take advantage of self._reflibs_cache
        for reflib in self.get_reflibs():
            # Double check validity of reflib
            if (reflib is not None) and not isinstance(reflib, DataLib):
                if reflib in self.field_index:
                    reflib = self.get(reflib, suppress_recording=True, log=False)
                else:
                    _logger.warning(f"reflib:{reflib} not present in self.field_index")

            if reflib is None:
                continue
            elif not isinstance(reflib, RefLib):
                raise RefLibExpectedError

            # NOTE: .map_id() is overloaded in RefLib Class...
            # see <RefLib>.map_id() for functionality
            keyspace_maps = reflib.ks_maps(src_keyspace, tgt_keyspace)
            if keyspace_maps is not None:
                return keyspace_maps, reflib

        return None, None

    def ks_maps(self, src_keyspace, tgt_keyspace=None):
        """
        Given src_keyspace and tgt_keyspace (optional), looks through all registered RefLibs and returns ks_map
        If src_keyspace is a tuple or list, then a single src_keyspace will be determined and returned
        If tgt_keyspace = None, then the associated RefLib.get_property("tgt_keyspace") is invoked for a given security master
        -----------------------------------------------------------------------------------------------------
        To obtain the post-resolved src_keyspace & tgt_keyspace, refer to ks_map.keyspaces[0] and ks_map.valuespace of the resultant ks_map accordingly
        To obtain the reflib_name that was used, the user should call self.ks_map_with_reflib() with the same args
        -----------------------------------------------------------------------------------------------------
        NOTE: method ks_map is overloaded within the RefLib class
        """
        (keyspace_maps, reflib) = self.ks_maps_with_reflib(
            src_keyspace=src_keyspace, tgt_keyspace=tgt_keyspace
        )
        return keyspace_maps

    def ks_map(self, src_keyspace, tgt_keyspace=None):
        (keyspace_map, reflib) = self.ks_map_with_reflib(
            src_keyspace=src_keyspace, tgt_keyspace=tgt_keyspace
        )
        return keyspace_map

    def ks_map_with_reflib(self, src_keyspace, tgt_keyspace=None):
        """
        Given src_keyspace and tgt_keyspace (optional), looks through all registered RefLibs
        and returns ks_maps (dictionary) keyed on respective intermediary tgt_keyspaces

        If src_keyspace is a tuple or list, then a single src_keyspace will be determined and returned
        If tgt_keyspace = None, then the associated RefLib.get_property("tgt_keyspace") is invoked for a given security master
        -----------------------------------------------------------------------------------------------------
        To obtain the post-resolved src_keyspace & tgt_keyspace, refer to ks_map.keyspaces[0] and ks_map.valuespace of the resultant ks_map accordingly
        To obtain the reflib_name that was used, the user should call self.ks_map_with_reflib() with the same args
        -----------------------------------------------------------------------------------------------------
        NOTE: method ks_maps is overloaded within the RefLib class
        """
        (keyspace_maps, reflib) = self.ks_maps_with_reflib(
            src_keyspace=src_keyspace, tgt_keyspace=tgt_keyspace
        )
        if keyspace_maps is None:
            return (None, None)
        elif not isinstance(keyspace_maps, dict):
            raise TypeError("RootLib().ks_maps() yeilded non-dictionary")
        elif len(keyspace_maps) == 0:
            return (None, None)
        elif len(keyspace_maps) == 1:
            first_tgt_keyspace = list(keyspace_maps.keys())[0]
            return keyspace_maps[first_tgt_keyspace]
        else:
            final_tgt_keyspace = list(keyspace_maps.keys())[-1]
            raise ValueError(
                f"Multiple keyspace maps required for translate:{src_keyspace} -> {final_tgt_keyspace}"
            )

    def apply_reflib_multilink(
        self,
        qubles,
        keyspace_aliases="<keyspace_aliases>",
        link_check=False,
        link_dupe_grace=True,
        impose_tgt_keyspace=False,
        reverse_iteration=False,
        fiscal_keyspace="Fiscal",
        vantage_keyspace="Vantage",
    ):
        """
        First investigates the application of keyspace aliases
        (if provided either directly or indirectly)
        and then seeks to apply appropriate
        keyspace mappings translation (e.g., security master maps)
        to resolve to a common keyspace (e.g., security identifier)
        across multiple Qubles (tuple, list or dictionary of Qubles)

        Returns a tuple set of the translated Qubles
        which should share a common keyspace when possible

        References registered 'reference libraries'
        (e.g., Security Masters) to identify
        translatable keyspaces in the original Qubles.
        -----------------------------------------------------
        If all Qubles contain CONSISTENT keyspaces,
        then the original qubles are returned unchanged (shallow_copy)
        -----------------------------------------------------
        If the Qubles contain INCONSISTENT keyspaces,
        then they are resolved ("remapped") to a common/target keyspace

        If impose_tgt_keyspace=False: common/target keyspace
        will be determined as first occurance of a keyspace across
        qubles list/dictionary that is supported by a given reflib

        If impose_tgt_keyspace=True:
           common keyspace will prioritize the
           RefLib.get_property('tgt_keyspace) (if available)

        NOTE: tgt_keyspace derived from applicable
        RefLib.get_property('tgt_keyspace) (if not None)
        otherwise from applicable src_keyspace1 from qbl
        (earliest qubles in list/dictionary)
        -----------------------------------------------------
        When no linking/translation applies,
        shallow copies (references) of the Qubles are returned

        If impose_tgt_keyspace=True: will possibly translate to reflib's tgt_keyspace
        even if a single keyspace is involved (not linking of multiple keyspaces)

        :type qubles: list/tuple/dict of Qubles (or a single Quble instance)
        :param qubles: The Qubles to be reconciled
        """
        from qubles.io.ref.reflib import RefLib

        linking_still_needed = True  # <-- Initialization

        # Validate  q1 & q2 args...
        # ----------------------------
        if qubles is None or is_quble(qubles):
            return qubles

        new_qubles = {}
        if isinstance(qubles, dict):
            for k in qubles:
                # new_qubles is a dictionary with keys as originally provided
                new_qubles[k] = qubles[k]
        elif is_quble(qubles):
            return qubles
        elif isinstance(qubles, (list, tuple)):
            for k in range(len(qubles)):
                # new_qubles is a dictionary with integers as keys
                new_qubles[k] = qubles[k]
        else:
            raise TypeError(
                f"Invalid qubles arg:{qubles}...dictionary/list/tuple expected"
            )

        num_qubles = len(qubles)
        non_trivial_quble_count = 0

        for qbl in list(new_qubles.values()):
            if qbl is None:
                continue
            elif not is_quble(qbl):
                raise TypeError("None or Quble elements expected")
            elif qbl.is_undefined:
                continue
            elif qbl.is_empty:
                # TODO: Investigate/ponder if we should exclude this case?
                continue
            else:
                non_trivial_quble_count += 1

        if non_trivial_quble_count <= 1:
            # Simply return the original qubles arg (dictionary or list)
            linking_still_needed = False

        # =============================================================
        # SECTION #1: FIRST, HANDLE NON-PIT CASES (Fiscal -> Calendar)
        # AS WELL AS PIT CASES (Fiscal + Vantage -> Calendar)
        # INVOLVING 'Fiscal' (AND 'Vantage') keyspaces IN PRESENCE OF
        # calendar time-keyspaces not in ('Fiscal', 'Vantage')
        # =============================================================

        # ============================ START: SECTION #1 ==========================

        if linking_still_needed:
            calendar_time_keyspace = None  # <-- Initialization
            qkeys_with_fiscal_keyspace = []  # <-- Initialization
            for q1key, q1 in new_qubles.items():
                if q1 is None:
                    continue
                elif not is_quble(q1):
                    continue
                elif q1.is_undefined:
                    continue
                else:
                    # Is fiscal_keyspace present in q1.keyspaces?
                    if fiscal_keyspace in q1.keyspaces:
                        qkeys_with_fiscal_keyspace.append(q1key)

                    ftks = q1.first_time_keyspace()

                    # Is calendar_time_keyspace present in q1.keyspaces?
                    # 'Fiscal' (and 'Vantage') are reserved keywords for keyspaces (index columns) that are not calendar time-columns
                    if (
                        calendar_time_keyspace is None
                        and ftks is not None
                        and ftks not in (fiscal_keyspace, vantage_keyspace)
                        and fiscal_keyspace not in q1.keyspaces
                        and vantage_keyspace not in q1.keyspaces
                    ):
                        # NOTE: calendar_time_keyspace will NOT be 'Fiscal' NOR 'Vantage'
                        calendar_time_keyspace = ftks

            if calendar_time_keyspace is not None:
                for q1key in qkeys_with_fiscal_keyspace:
                    new_qubles[q1key] = new_qubles[q1key].calendarize(
                        fiscal_keyspace=fiscal_keyspace,
                        calendar_keyspace=calendar_time_keyspace,
                        vantage_keyspace=vantage_keyspace,
                    )

        # ============================= END: SECTION #1 ===========================

        # ---------------------------------------------
        # unique_unmapped_keyspaces is a dictionary from ks to a list of quble numbers/indices indicating which qubles utilize this keyspace
        # ----------------------------------------------
        unique_unmapped_keyspaces = {}

        if reverse_iteration:
            new_qubles_key_iterator = reversed(list(new_qubles.keys()))
        else:
            new_qubles_key_iterator = list(new_qubles.keys())

        for qkey1 in new_qubles_key_iterator:
            quble1 = new_qubles[qkey1]
            if quble1 is None:
                continue

            for ks in quble1.keyspaces:
                if ks not in unique_unmapped_keyspaces:
                    unique_unmapped_keyspaces[ks] = [qkey1]
                    # continue
                else:
                    unique_unmapped_keyspaces[ks].append(qkey1)

        # Apply Pre-Test
        # [Remove keyspaces occuring across all qubles as linking does not apply here]
        for ks in list(unique_unmapped_keyspaces.keys()):
            # referencing <dict>.keys() in the iteration so that altering the unique_unmapped_keyspaces inside loop will not affect the iteration
            if len(unique_unmapped_keyspaces[ks]) == num_qubles:
                unique_unmapped_keyspaces.pop(ks)

        # Re-check length of unique_unmapped_keyspaces
        # Could have changed due to popping above and/or if multiple qubles (args) all exhibit the same single keyspace
        if len(unique_unmapped_keyspaces) <= 1:
            # There is only one keyspace left, so no more linking is required
            linking_still_needed = False
        else:
            # How many qubles are represented in the remaining unique_unmapped_keyspaces?
            qkeys_set = set()
            for ks in unique_unmapped_keyspaces:
                # unique_unmapped_keyspaces[ks] should be a list here
                qkeys_set.update(set(unique_unmapped_keyspaces[ks]))
                if len(qkeys_set) > 1:
                    break
            if len(qkeys_set) <= 1:
                # There is only one quble represented across the unmapped keyspaces so no more linking is required
                linking_still_needed = False

        # =====================================================
        # SECTION #2: Apply keyspace aliases (when provided)
        # =====================================================

        # ============================ START: SECTION #2 ==========================
        if linking_still_needed:
            if isinstance(keyspace_aliases, str):
                if keyspace_aliases == "<keyspace_aliases>":
                    keyspace_aliases = self.get_property(
                        "keyspace_aliases", appeal=True, try_json_loads=True, grace=True
                    )
                else:
                    raise TypeError(
                        f"Invalid keyspace_aliases:{keyspace_aliases}...expected list of lists or dict of lists or None or '<keyspace_aliases>'"
                    )

            # Convert keyspace_aliases from dict of lists/tuples -> list of lists/tuples
            if isinstance(keyspace_aliases, dict):
                aliases_as_list = []
                for alias_group_name, ks_aliases in keyspace_aliases.items():
                    if ks_aliases is None:
                        continue
                    elif not isinstance(ks_aliases, (list, tuple, set)):
                        raise TypeError(
                            f"Invalid dict entry keyspace_aliases:[{alias_group_name}]:{ks_aliases}...list/tuple/set dict values expected"
                        )
                    aliases_as_list.append(ks_aliases)
                keyspace_aliases = aliases_as_list

            # At this point, keyspace_aliases should be either:
            #    1) None
            # or 2) a list/tuple of lists/tuples
            # -----------------------------------------------
            if keyspace_aliases is None:
                pass
            elif not isinstance(keyspace_aliases, (list, tuple)):
                raise TypeError(
                    f"Invalid keyspace_aliases:{keyspace_aliases}...dict/list of lists/tuples/sets (or None) expected"
                )
            elif len(keyspace_aliases) == 0:
                pass
            elif len(unique_unmapped_keyspaces) <= 1:
                # Need atleast two keyspaces to invoke an alias
                pass
            else:
                # ========== START: len(unique_unmapped_keyspaces) >= 2 ==========
                # =================== START: keyspace_aliases LOOP ==================
                for ks_aliases1 in keyspace_aliases:
                    # Validate ks_aliases1
                    if ks_aliases1 is None:
                        continue
                    elif not isinstance(ks_aliases1, (list, tuple, set)):
                        raise TypeError(
                            f"Invalid keyspace_aliases:{keyspace_aliases}...dict/list of lists/tuples/sets (or None) expected"
                        )

                    # Remove None elements from current ks_aliases1 (to be safe)
                    # Then create a set for comparison (will eliminate duplicates)
                    ks_aliases1 = set([ks for ks in ks_aliases1 if ks is not None])

                    if len(ks_aliases1) <= 1:
                        # Need more than one alias to be relevant
                        continue

                    # Choosing the first unmapped keyspace that appears in the ks_aliases1 (if any)
                    # (current list of keyspace aliases)
                    tgt_keyspace = None
                    for ks in unique_unmapped_keyspaces:
                        if ks in ks_aliases1:
                            tgt_keyspace = ks
                            break
                    if tgt_keyspace is None:
                        # No unmapped keyspaces occur in the current ks_aliases1
                        # As such, simply continue to next set of keyspace aliases
                        continue
                    elif tgt_keyspace not in unique_unmapped_keyspaces:
                        # Should not happen
                        raise ValueError(
                            f"Internal inconsistency...tgt_keyspace:{tgt_keyspace} absent from unique_unmapped_keyspaces:{unique_unmapped_keyspaces}"
                        )

                    complement_unmapped_keyspaces = [
                        ks for ks in unique_unmapped_keyspaces if ks != tgt_keyspace
                    ]
                    if len(complement_unmapped_keyspaces) == 0:
                        break  # <-- Break out of keyspace_aliases loop if no complements exist

                    keyspaces_rename_map = {}
                    tgt_keyspaces_applied = []

                    # ======================= START: src_ks LOOP ====================
                    # Loop through the complementary unmapped keyspaces
                    for src_ks in complement_unmapped_keyspaces:
                        if src_ks in ks_aliases1:
                            # Does this src_ks appear in these ks_aliases1?
                            # AND do any of the absent_tgt_keyspaces ALSO appear in the same ks_aliases1?
                            keyspaces_rename_map[src_ks] = tgt_keyspace

                    # ======================== END: src_ks LOOP =====================
                    # Affect the renaming where/where applicable
                    for src_ks, tgt_ks in keyspaces_rename_map.items():
                        if src_ks in unique_unmapped_keyspaces:
                            # and unique_unmapped_keyspaces[src_ks] is not None \
                            # and len(unique_unmapped_keyspaces[src_ks]) > 0:
                            for qkey in unique_unmapped_keyspaces[src_ks]:
                                if tgt_ks not in tgt_keyspaces_applied:
                                    tgt_keyspaces_applied.append(tgt_ks)
                                new_qubles[qkey] = new_qubles[qkey].rename_spaces(
                                    space_map=keyspaces_rename_map
                                )
                            unique_unmapped_keyspaces.pop(src_ks)

                            if len(unique_unmapped_keyspaces) <= 1:
                                # Need atleast two keyspaces to invoke an alias
                                break

                    # Pop any tgt_ks that has been applied from the unique_unmapped_keyspaces
                    for tgt_ks in tgt_keyspaces_applied:
                        if tgt_ks in unique_unmapped_keyspaces:
                            unique_unmapped_keyspaces.pop(tgt_ks)

                    if len(unique_unmapped_keyspaces) <= 1:
                        # Need atleast two keyspaces to invoke an alias
                        break

                # Can happen due to popping above and/or if multiple qubles (args) all exhibit the same single keyspace
                if len(unique_unmapped_keyspaces) <= 1:
                    linking_still_needed = False
                else:
                    # How many qubles are represented in the remaining unique_unmapped_keyspaces?
                    qkeys_set = set()
                    for ks in unique_unmapped_keyspaces:
                        # unique_unmapped_keyspaces[ks] should be a list here
                        qkeys_set.update(set(unique_unmapped_keyspaces[ks]))
                        if len(qkeys_set) > 1:
                            break
                    if len(qkeys_set) <= 1:
                        # There is only one quble represented across the unmapped keyspaces so no more linking is required
                        linking_still_needed = False

                # ==================== END: keyspace_aliases LOOP ===================
                # =========== END: len(unique_unmapped_keyspaces) >= 2 ===========

        # ============================= END: SECTION #2 ===========================

        # =====================================================
        # SECTION #3: Apply reflibs remaps (when provided)
        # -----------------------------------------------------
        # Loop through the available reflibs (within/under the current library)
        # =====================================================

        # ============================ START: SECTION #3 ==========================
        if linking_still_needed:
            # ========================== START: reflib LOOP =========================
            # NOTE: get_reflibs() will take advantage of self._reflibs_cache
            for reflib in self.get_reflibs():
                if (reflib is not None) and not isinstance(
                    reflib, DataLib
                ):  # <-- Double check
                    if reflib in self.field_index:
                        reflib = self.get(reflib, suppress_recording=True, log=False)
                    else:
                        _logger.warning(
                            f"reflib:{reflib} not present in self.field_index"
                        )

                if reflib is None:
                    continue
                elif not isinstance(reflib, RefLib):
                    raise RefLibExpectedError

                # Check whether this reflib is to be exempted from auto_link
                # -------------------------------------------------------------
                auto_link_exemption = reflib.get_property(
                    "auto_link_exemption", grace=True
                )
                if auto_link_exemption:
                    continue

                curr_reflib_src_keyspaces = reflib.src_keyspaces

                # the unique keyspaces that occur within the current reflib.src_keyspaces list
                supported_unmapped_keyspaces = [
                    ks for ks in unique_unmapped_keyspaces if ks in reflib.src_keyspaces
                ]

                # continue if nothing can be resolved using the current reflib
                if len(supported_unmapped_keyspaces) == 0:
                    continue

                # Prioritize the sanctioned tgt_keyspace for this reflib if applicable
                tgt_keyspace = (
                    None  # <-- Added initilization at each pass through the loop
                )
                if impose_tgt_keyspace:
                    tgt_keyspace = reflib.get_tgt_keyspace(grace=True)

                # If needed, try to identify an appropriate tgt_keyspace for this reflib
                # by choosing the first unmapped keyspace that can be reached from this reflib
                # [It is only incrementally helpful to invoke reflibs to get to one of the unmapped keyspaces]
                if tgt_keyspace is None:
                    for ks in unique_unmapped_keyspaces:
                        if ks in reflib.tgt_keyspaces:
                            tgt_keyspace = ks
                            break

                # If no useful tgt_keyspace could be identified, then deem this reflib as not-helpful
                # to the keyspace resolution process and continue on to next reflib
                if tgt_keyspace is None:
                    continue
                else:
                    # Make sure tgt_keyspace is reachable using this reflib
                    if tgt_keyspace not in reflib.tgt_keyspaces:
                        raise ValueError(
                            f"Invalid/unsupported tgt_keyspace:{tgt_keyspace}"
                        )

                    qkeys_already_linked_by_curr_reflib = []
                    # ========================= START: ks LOOP ======================
                    for ks in deepcopy(unique_unmapped_keyspaces):
                        if ks not in curr_reflib_src_keyspaces:
                            continue

                        # See if a previous loop iteration might have already reconciled & removed this ks
                        # (as the target of a prior remap1d() call
                        if ks not in unique_unmapped_keyspaces:
                            continue

                        if ks != tgt_keyspace:
                            for qkey1 in unique_unmapped_keyspaces[ks]:
                                # Loop through the qubles containing this keyspace (ks)
                                if qkey1 not in qkeys_already_linked_by_curr_reflib:
                                    if ks in new_qubles[qkey1].keyspaces:
                                        new_qubles[qkey1] = reflib.remap1d(
                                            new_qubles[qkey1],
                                            src_keyspace=ks,
                                            tgt_keyspace=tgt_keyspace,
                                            link_check=link_check,
                                            link_dupe_grace=link_dupe_grace,
                                        )
                                        qkeys_already_linked_by_curr_reflib.append(
                                            qkey1
                                        )

                            unique_unmapped_keyspaces.pop(ks)

                        if (tgt_keyspace is not None) and (
                            tgt_keyspace in unique_unmapped_keyspaces
                        ):
                            unique_unmapped_keyspaces.pop(tgt_keyspace)

                        if len(unique_unmapped_keyspaces) == 0:
                            # No more unmapped keyspaces, All keyspaces have been mapped to tgt_keyspace
                            linking_still_needed = False
                            break

                    # ========================== END: ks LOOP =======================
            # =========================== END: reflib LOOP ==========================
        # ============================= END: SECTION #3 ===========================

        # -----------------------------------------
        # Return results from new_qubles
        # according to original qubles arg format
        # ----------------------------------------
        if isinstance(qubles, dict):
            # Was originally provided a dictionary (of qubles), return the modified new_qubles dict
            return new_qubles
        elif is_quble(qubles):
            # Was originally provided a single Quble, return first value of the new_qubles dict
            return list(new_qubles.values())[0]
        else:
            # Was originally provided a list/tuple, return only the values of the new_qubles dict
            return list(new_qubles.values())

    def remap(
        self,
        src_obj,
        src_keyspaces=None,
        tgt_keyspaces=None,
        link_check=False,
        link_dupe_grace=True,
        deep_copy=True,
        grace=True,
        exceptions_flag=False,
        exceptions_type=None,
        apply_vantage_aging=True,
        fiscal_keyspace="Fiscal",
        vantage_keyspace="Vantage",
    ):
        """
        Given a Quble, seeks to remap (possibly multiple) src keyspaces by consulting all embedded Reference Libraries
        Given a src_obj, (optional) src_keyspaces (scalar/list/tuple/None) and (optional) tgt_keyspaces (scalar/list/tuple/None),
        attempts to remap src_keyspace(s) to the candidate tgt_keyspace(s) using maps within any residence Reference Libraries
        -----------------------------------------------------------------------------------------------------------------------------------
        If tgt_keyspace = None, then RefLib.get_property("tgt_keyspace") is invoked for each Reference Libraries

        If deep_copy=True, if no ramapping is applied, then a copy of the source object is returned
        """
        from qubles.core.quble import Quble
        from qubles.io.base.rootlib import RootLib

        root_lib = RootLib()
        remaps_performed = False

        # Validate src_obj arg...
        # -------------------------
        if src_obj is None:
            return src_obj
        elif isinstance(src_obj, (Quble)):
            pass
        else:
            raise TypeError("Invalid src_obj: Quble expected")

        # Validate src_keyspaces arg...
        # ------------------------------
        if src_keyspaces is None:
            # Here, the keyspaces of src_obj serve as src_keyspaces
            # In this case, we need to make a deepcopy here as we may alter src_obj.keyspaces list below
            src_keyspaces = deepcopy(src_obj.keyspaces)
        elif isinstance(src_keyspaces, (list, tuple)):
            # We do not want to contaminate original src_keyspaces on exit
            src_keyspaces = deepcopy(src_keyspaces)
        else:
            src_keyspaces = [src_keyspaces]

        # Validate tgt_keyspaces arg...
        # ------------------------------
        if tgt_keyspaces is not None:
            if isinstance(tgt_keyspaces, (list, tuple)):
                # We do not want to contaminate original tgt_keyspaces on exit
                tgt_keyspaces = deepcopy(tgt_keyspaces)
            else:
                tgt_keyspaces = [tgt_keyspaces]

            for ks in tgt_keyspaces:
                if ks in src_obj.keyspaces:
                    # Don't want to target/transition
                    # to keyspaces that are already present
                    tgt_keyspaces.remove(ks)

        # Exit early if there are no src_keyspaces to remap or no tgt_keyspaces left
        if len(src_keyspaces) != 0 or len(tgt_keyspaces) != 0:
            # =================== START: src_keyspaces LOOP ===================
            for ks in src_keyspaces:
                # Validate ks against src_obj.keyspaces
                # ---------------------------------------
                if ks not in src_obj.keyspaces:
                    raise ValueError(
                        f"src_keyspace:{ks} not in src_obj.keyspaces: {src_obj.keyspaces}"
                    )

                # Handle special cases:
                # -----------------------
                if (ks == "SEDOL6") and (tgt_keyspaces == ["SEDOL"]):
                    if isinstance(src_obj, Quble):
                        table_name = generate_random_table_name()
                        execute(
                            f"""
                            create or replace table {dquote_dot(table_name)} as
                            select DISTINCT "{ks}", Sedol6To7_js("{ks}",False) as "{tgt_keyspaces[0]}"
                            from {dquote_dot(src_obj.table_name)}
                            """,
                            fetch="one",
                        )
                        ks_map = Quble.from_table(
                            f"{table_name}", valuespace=tgt_keyspaces[0]
                        )

                    else:
                        raise TypeError(
                            f"Expected instance of Quble, but received type:{type(src_obj)}"
                        )

                    ks_maps = {tgt_keyspaces[0]: ks_map}
                elif (ks == "SEDOL") and (tgt_keyspaces == ["SEDOL6"]):
                    if isinstance(src_obj, Quble):
                        table_name = generate_random_table_name()
                        execute(
                            f"""
                            create or replace table {dquote_dot(table_name)} as
                            select DISTINCT "{ks}", SUBSTRING("{ks}",1,6) as "{tgt_keyspaces[0]}"
                            from {dquote_dot(src_obj.table_name)}
                            """,
                            fetch="one",
                        )
                        ks_map = Quble.from_table(
                            f"{table_name}", valuespace=tgt_keyspaces[0]
                        )
                    else:
                        raise TypeError(
                            f"Expected instance of Quble, but received type:{type(src_obj)}"
                        )

                    ks_maps = {tgt_keyspaces[0]: ks_map}
                elif (ks == "CUSIP8") and (tgt_keyspaces == ["CUSIP"]):
                    if isinstance(src_obj, Quble):
                        table_name = generate_random_table_name()
                        execute(
                            f"""
                            create or replace table {dquote_dot(table_name)} as
                            select DISTINCT "{ks}", cusip8To9_js("{ks}",False) as "{tgt_keyspaces[0]}"
                            from {dquote_dot(src_obj.table_name)}
                            """,
                            fetch="one",
                        )
                        ks_map = Quble.from_table(
                            f"{table_name}", valuespace=tgt_keyspaces[0]
                        )
                    else:
                        raise TypeError(
                            f"Expected instance of Quble, but received type:{type(src_obj)}"
                        )

                    ks_maps = {tgt_keyspaces[0]: ks_map}
                elif (ks == "CUSIP") and (tgt_keyspaces == ["CUSIP8"]):
                    if isinstance(src_obj, Quble):
                        table_name = generate_random_table_name()
                        execute(
                            f"""
                            create or replace table {dquote_dot(table_name)} as
                            select DISTINCT "{ks}", SUBSTRING("{ks}",1,8) as "{tgt_keyspaces[0]}"
                            from {dquote_dot(src_obj.table_name)}
                            """,
                            fetch="one",
                        )
                        ks_map = Quble.from_table(
                            f"{table_name}", valuespace=tgt_keyspaces[0]
                        )
                    else:
                        raise TypeError(
                            f"Expected instance of Quble, but received type:{type(src_obj)}"
                        )

                    ks_maps = {tgt_keyspaces[0]: ks_map}
                # ------------------------------------------------------------------------------------------------------------------------
                # PIT Case: When converting from 'Fiscal' to another single (presumably time) tgt_keyspace (not 'Fiscal')
                #           but also in the presence of a secondary 'Vantage' keyspace within src_obj
                #
                # Apply PIT Fiscal --> Calendar Time Mapping Logic as follows:
                #
                #    1) Select the last non-missing value across the 'Fiscal' dimension
                #       for each (orthogonal) 'Vantage' key ==> Drops (squeezes) the 'Fiscal' keyspace away
                #
                #    2) Rename the 'Vantage' keyspace to the single tgt_keyspace provided
                #       (presumably to be later used/interpretted as a calendar time keyspace)
                # ------------------------------------------------------------------------------------------------------------------------
                elif (
                    (ks == fiscal_keyspace)
                    and isinstance(tgt_keyspaces, (list, tuple))
                    and (len(tgt_keyspaces) == 1)
                    and (fiscal_keyspace not in tgt_keyspaces)
                    and (vantage_keyspace in src_obj.keyspaces)
                ):
                    # ----------------------------------
                    # First, apply pre-cross fill here
                    # ----------------------------------
                    src_obj = src_obj.cross_fill1d(
                        keyspace=vantage_keyspace,
                        crosser_keyspace=fiscal_keyspace,
                        remove_dupes=True,
                    )

                    # -------------------------------------------------
                    # next, apply Vantage "Aging" (if/when applicable)
                    # -------------------------------------------------
                    # NOTE: Fiscal frequencies faster than monthly are likely not practical

                    fiscal_freq = src_obj.get_freq(space=fiscal_keyspace)
                    fiscal_ppy = Quble.ppy(fiscal_freq)

                    vantage_aging_clause = None

                    if not apply_vantage_aging:
                        # Do not apply vintage_aging_clause
                        pass
                    elif not fiscal_ppy:
                        # NOTE: This case ideally should not happen
                        pass
                    elif fiscal_ppy < 1.5:
                        # Annual Fiscal frequency (fiscal_ppy=1)
                        # Aging Condition: ("Vantage" - "Fiscal") <= 16 Months
                        # NOTE: DATEDIFF(<unit>, <earlier_datetime>, <latter_datetime>)
                        vantage_aging_clause = f'DATEDIFF(month, "{fiscal_keyspace}", "{vantage_keyspace}") <= 16'
                    elif fiscal_ppy < 2.5:
                        # Semi-Annual Fiscal frequency (fiscal_ppy=2)
                        # Aging Condition: ("Vantage" - "Fiscal") <= 10 Months
                        # NOTE: DATEDIFF(<unit>, <earlier_datetime>, <latter_datetime>)
                        vantage_aging_clause = f'DATEDIFF(month, "{fiscal_keyspace}", "{vantage_keyspace}") <= 10'
                    elif fiscal_ppy < 5:
                        # Quarterly Fiscal frequency (fiscal_ppy=4)
                        # Aging Condition: ("Vantage" - "Fiscal") <= 7 Months
                        # NOTE: DATEDIFF(<unit>, <earlier_datetime>, <latter_datetime>)
                        vantage_aging_clause = f'DATEDIFF(month, "{fiscal_keyspace}", "{vantage_keyspace}") <= 7'
                    elif fiscal_ppy < 14:
                        # Monthly Fiscal frequency (fiscal_ppy=12)
                        # Aging Condition: ("Vantage" - "Fiscal") <= 4 Months
                        # NOTE: DATEDIFF(<unit>, <earlier_datetime>, <latter_datetime>)
                        # NOTE: This case is likely not practical for "Fiscal" time-spaces
                        vantage_aging_clause = f'DATEDIFF(month, "{fiscal_keyspace}", "{vantage_keyspace}") <= 4'

                    if vantage_aging_clause:
                        src_obj = src_obj.select(where_clause=vantage_aging_clause)

                    # ------------------------------------------------------------------
                    # Next, apply last1d logic across "Fiscal" time-keyspace
                    # ------------------------------------------------------------------
                    # Assume the single tgt_keyspace provided to replace 'Fiscal'
                    # is intended to become the resultant valid (calendar) time-keyspace
                    # NOTE: cross-fill applied above, so use ==> pre_cross_fill=False
                    # ------------------------------------------------------------------
                    src_obj = src_obj.last1d(
                        keyspace=fiscal_keyspace,
                        ignore_missing=True,
                        auto_squeeze=True,
                        pre_cross_fill=False,
                    )

                    # -------------------------------------------------------------
                    # Rename keyspace (if needed): "Vantage" -> tgt_keyspaces[0]
                    # -------------------------------------------------------------
                    if tgt_keyspaces[0] != vantage_keyspace:
                        src_obj = src_obj.rename_space(
                            vantage_keyspace, tgt_keyspaces[0]
                        )

                    ks_maps = None
                    remaps_performed = True

                    # ------------------------------
                    # Assign time_basis, tfill_max
                    # ------------------------------
                    if tgt_keyspaces[0] in src_obj.keyspaces and src_obj.is_time_space(
                        tgt_keyspaces[0]
                    ):
                        src_calendar_freq = src_obj.get_freq(space=tgt_keyspaces[0])
                        tgt_calendar_freq = RootLib().get_control("freq")
                        if (
                            src_calendar_freq is not None
                            and tgt_calendar_freq is not None
                            and src_calendar_freq != tgt_calendar_freq
                        ):
                            # First, sort by keys ascending before convert_date_values
                            src_obj = src_obj.key_sort(key_ordering="asc")
                            src_obj = src_obj.convert_date_values(
                                space=tgt_keyspaces[0], freq=tgt_calendar_freq
                            )

                            # Remove any duplicate keys that may have been introduced via convert_date_values method
                            src_obj.check_remove_duplicate_keys()

                        if src_obj.is_variate:
                            # We have just remapped a time-space of a variate Quble
                            # Set tfill_max = 1.15 * ppy of freq of new time-keyspace

                            # We assume that the 'time_basis' was originally meant for use with the (now removed) Fiscal
                            # As such, we nullify 'time_basis' info_type for valuespaces
                            src_obj.set_space_info(
                                space="<valuespaces>",
                                info_type="time_basis",
                                info_value=None,
                            )

                            freq = src_obj.get_space_info(
                                info_type="freq", space=tgt_keyspaces[0], grace=True
                            )
                            ppy = datetools.ppy(freq, grace=True)
                            if ppy is not None:
                                src_obj.set_space_info(
                                    space=src_obj.valuespaces,
                                    info_type="tfill_max",
                                    info_value=int(1.15 * ppy),
                                )

                    continue  # <-- continue on to next ks in src_keyspaces
                # -------------------------------------------------
                # Non-PIT Case when either:
                #     1) tgt_keyspaces[0] is present as valuespace
                # OR  2) "REPDATE_ORG" is present as valuespace
                # -------------------------------------------------
                elif (
                    (ks == fiscal_keyspace)
                    and isinstance(tgt_keyspaces, (list, tuple))
                    and (len(tgt_keyspaces) == 1)
                    and (
                        tgt_keyspaces[0] in src_obj.valuespaces
                        or "REPDATE_ORG" in src_obj.valuespaces
                    )
                    and (fiscal_keyspace not in tgt_keyspaces)
                    and (vantage_keyspace not in src_obj.keyspaces)
                ):
                    tgt_keyspaces2 = [ks1 for ks1 in src_obj.keyspaces if ks1 != ks] + [
                        tgt_keyspaces[0]
                    ]
                    role_overrides = {}
                    for ks1 in tgt_keyspaces2:
                        role_overrides[ks1] = "keyspace"

                    tgt_calendar_freq = RootLib().get_control("freq")
                    date_values_converted = False

                    if "REPDATE_ORG" in src_obj.valuespaces:
                        src_calendar_freq = src_obj.get_freq(space="REPDATE_ORG")

                        # Convert date values (if needed)
                        if (
                            src_calendar_freq is not None
                            and tgt_calendar_freq is not None
                            and src_calendar_freq != tgt_calendar_freq
                        ):
                            # First, sort date values ascending before convert_date_values
                            src_obj = src_obj.value_sort(
                                valuespace="REPDATE_ORG", value_ordering="asc"
                            )
                            src_obj = src_obj.convert_date_values(
                                space="REPDATE_ORG", freq=tgt_calendar_freq
                            )
                            date_values_converted = True

                        tgt_valuespaces2 = [
                            vs1 for vs1 in src_obj.valuespaces if vs1 != "REPDATE_ORG"
                        ]
                        for vs1 in tgt_valuespaces2:
                            role_overrides[vs1] = "valuespace"

                        tgt_spaces2 = {}
                        for ks1 in tgt_keyspaces2:
                            if ks1 == tgt_keyspaces[0]:
                                tgt_spaces2["REPDATE_ORG"] = ks1
                            else:
                                tgt_spaces2[ks1] = ks1
                        for vs1 in tgt_valuespaces2:
                            tgt_spaces2[vs1] = vs1

                    else:
                        src_calendar_freq = src_obj.get_freq(space=tgt_keyspaces[0])

                        # Convert date values (if needed)
                        if (
                            src_calendar_freq is not None
                            and tgt_calendar_freq is not None
                            and src_calendar_freq != tgt_calendar_freq
                        ):
                            # First, sort date values ascending before convert_date_values
                            src_obj = src_obj.value_sort(
                                valuespace=tgt_keyspaces[0], value_ordering="asc"
                            )
                            src_obj = src_obj.convert_date_values(
                                space=tgt_keyspaces[0], freq=tgt_calendar_freq
                            )
                            date_values_converted = True

                        tgt_valuespaces2 = [
                            vs1
                            for vs1 in src_obj.valuespaces
                            if vs1 != tgt_keyspaces[0]
                        ]
                        for vs1 in tgt_valuespaces2:
                            role_overrides[vs1] = "valuespace"

                        tgt_spaces2 = tgt_keyspaces2 + tgt_valuespaces2

                    # Perform the selection
                    src_obj = src_obj.select(
                        column_names=tgt_spaces2,
                        custom_info_overrides={"role": role_overrides},
                    )

                    # Remove any duplicate keys that may have been introduced via convert_date_values method
                    if date_values_converted:
                        src_obj.check_remove_duplicate_keys()

                    ks_maps = None
                    remaps_performed = True

                    if (
                        tgt_keyspaces[0] in src_obj.keyspaces
                        and src_obj.is_time_space(tgt_keyspaces[0])
                        and src_obj.is_variate
                    ):
                        # We have just remapped a time-space of a variate Quble
                        # Set tfill_max = 1.15 * ppy of freq of new time-keyspace

                        # We assume that the 'time_basis' was originally meant for use with the (now removed) Fiscal
                        # As such, we nullify 'time_basis' info_type for valuespaces
                        src_obj.set_space_info(
                            space=src_obj.valuespaces,
                            info_type="time_basis",
                            info_value=None,
                        )

                        freq = src_obj.get_space_info(
                            info_type="freq", space=tgt_keyspaces[0], grace=True
                        )
                        ppy = datetools.ppy(freq, grace=True)
                        if ppy is not None:
                            src_obj.set_space_info(
                                space=src_obj.valuespaces,
                                info_type="tfill_max",
                                info_value=int(1.15 * ppy),
                            )
                    continue

                # ---------------------------------------
                # General remapping case, search RootLib
                # for applicable reference libs
                # ---------------------------------------
                else:
                    (ks_maps, reflib) = RootLib().ks_maps_with_reflib(ks, tgt_keyspaces)

                # Try to use available/sanctioned ks_maps to eliminate extra ks from keymap
                # -----------------------------------------
                # =================== START: if ks_maps is not None ===================
                if ks_maps is not None:
                    # Qualify key_maps as either None or a dictionary
                    if not isinstance(ks_maps, dict):
                        raise TypeError(
                            "RootLib().ks_maps_with_reflib() yeilded non-dictionary"
                        )

                    # =================== START: ks_maps LOOP ===================
                    src_ks = ks  # <-- Initialization
                    for tgt_ks, ks_map in ks_maps.items():
                        if ks_map is None:
                            continue
                        if (src_ks is not None) and (tgt_ks is not None):
                            if tgt_ks in src_obj.keyspaces:
                                continue
                            if src_ks not in src_obj.keyspaces:
                                # This should not happen if machinery is working properly (since we requested ks_map(ks,...) and ks has already been checked to fall within src_obj.keyspaces
                                _logger.debug(
                                    f"src_ks:{src_ks}, tgt_ks:{tgt_ks} src_obj.keyspaces:{src_obj.keyspaces}"
                                )
                            if src_ks not in src_obj.keyspaces:
                                continue

                            if isinstance(src_obj, Quble):
                                if link_check and not ks_map.fully_covers(
                                    src_obj,
                                    column_name=src_ks,
                                    other_column_name=src_ks,
                                    grace=False,
                                ):
                                    raise ValueError(
                                        f"Some index elements in src_obj[{src_ks}] are not covered by ks_map[{src_ks}]"
                                    )
                            else:
                                raise TypeError(
                                    f"Expected instance of Quble, but received type:{type(src_obj)}"
                                )

                            freq = None
                            if src_ks == fiscal_keyspace:
                                # Here, we use map_type='aggregate', map_basis='last:'+src_ks to handle the possibility of multiple
                                # Fiscal periods being reported on same (calendarized) report date
                                # ------------------------------------------------------------
                                map_type = "aggregate"
                                map_basis = f"last:{src_ks}"
                            else:
                                map_type = "<space>"
                                map_basis = "<space>"

                            src_obj = src_obj.remap1d(
                                ks_map,
                                keyspace=src_ks,
                                tgt_keyspace=tgt_ks,
                                map_type=map_type,
                                map_basis=map_basis,
                                tfill_method=None,
                                tfill_max=None,
                                tfill_end_mode="no_future",
                                auto_link=root_lib.get_control("auto_link"),
                                link_check=link_check,
                                link_dupe_grace=link_dupe_grace,
                                exceptions_flag=exceptions_flag,
                                exceptions_type=exceptions_type,
                            )

                            if (
                                tgt_ks in src_obj.keyspaces
                                and src_obj.is_time_space(tgt_ks)
                                and src_obj.is_variate
                            ):
                                # We have just remapped a time-space of a variate Quble
                                # Set tfill_max = 1.15 * ppy of freq of new time-keyspace

                                # We assume that the 'time_basis' was originally meant for use with the (now removed) Fiscal
                                # As such, we nullify 'time_basis' info_type for valuespaces
                                if src_obj.is_variate:
                                    src_obj.set_space_info(
                                        space="<valuespaces>",
                                        info_type="time_basis",
                                        info_value=None,
                                    )

                                ppy = datetools.ppy(
                                    src_obj.get_space_info(
                                        info_type="freq", space=tgt_ks, grace=True
                                    ),
                                    grace=True,
                                )
                                if ppy is not None:
                                    src_obj.set_space_info(
                                        space=src_obj.valuespaces,
                                        info_type="tfill_max",
                                        info_value=int(1.15 * ppy),
                                    )

                                remaps_performed = True
                                if src_ks in src_keyspaces:
                                    src_keyspaces.remove(src_ks)
                                if (
                                    tgt_keyspaces is not None
                                ) and tgt_ks in tgt_keyspaces:
                                    tgt_keyspaces.remove(tgt_ks)

                        src_ks = tgt_ks  # <-- Re-assign for second-stage maps

                    # ==================== END: ks_maps LOOP ====================
                # ==================== END: if ks_maps is not None ====================
            # ==================== END: src_keyspaces LOOP ====================

        # ----------------------
        # Clean-Up and Return
        # ----------------------
        if not remaps_performed:
            if not grace:
                raise Exception("No remaps were performed and no grace allowed")
            elif deep_copy:
                src_obj = src_obj.copy()

        return src_obj

    # ================================== I/O Methods ==================================

    def touch(self):
        """Updates DataLib's internal time_stamp with current time."""
        self.time_stamp = datetime.now()

    def to_lib(self, destruction_prep=False):
        if hasattr(self, "address") and self.address is not None:
            self.address.deliver(self, destruction_prep=destruction_prep)
        else:
            raise ValueError("self.address is None")

    def commit(self, recursive=False, destruction_prep=False):
        """Commits library information (as applicable)

        recursive=False: commits the library's own meta data only
        recursive=True: commits the library's fields-in-residence
        (as well recursively down through any sublibs)
        """
        if self.is_temporary:
            return

        # Loop through fields
        # -----------------------
        if recursive:
            for field in self.field_index:
                residence_key = (field, self.property_dep_hash_current(field))
                if not self._exists_in_residence(residence_key) and hasattr(
                    self.residence[residence_key], "commit"
                ):
                    self.residence[residence_key].commit(
                        recursive=recursive, destruction_prep=destruction_prep
                    )

        if hasattr(self, "address") and (self.address is not None):
            self.address.deliver(self, destruction_prep=destruction_prep)

    @classmethod
    def from_qdb(cls, qdb, qdb_key):
        from qubles.io.util.qdb import QDB

        qdbobj = None
        try:
            if isinstance(qdb, QDB) or isinstance(qdb, str) or qdb[-4:] == ".qdb":
                if isinstance(qdb, QDB):
                    qdbobj = qdb
                elif not os.path.exists(normalize(qdb)):
                    raise Exception(f"Bad file_name: {qdb}")
                else:
                    qdbobj = QDB(qdb, "r")

                lib = qdbobj.get(key=qdb_key)
            else:
                raise ValueError("Bad qdb argument provided")

            return lib
        finally:
            try:
                qdbobj.close()
            except AttributeError:
                pass

    @classmethod
    def from_lib(cls, address):
        """
        Loads element from one of the registered data libraries within RootLib

        :param fieldname: fieldname per the directory of the specified library
        :type fieldname: str
        """
        from qubles.io.base.rootlib import RootLib

        result = RootLib().get(address)
        if not isinstance(result, DataLib):
            raise ValueError(f"Non-DataLib result: DataLib.fromLib(address={address})")
        else:
            return result

    @classmethod
    def from_xl(
        cls,
        directory_book,
        directory_sheet,
        directory_property_map=None,
        defaults_book=None,
        defaults_sheet=None,
        defaults_header=None,
        ignore_empty_strings=True,
        verbose=False,
    ):
        """
        Loads a DataLib Schema from an Excel workbook

        :param directory_book: Excel workbook file containing the DataLib directory information
        :type filename: str

        :param directory_sheet: Excel sheetname containing the DataLib directory and associated (per field) property info
                                First row will be treated as header row
                                First column will treated as field name column (regardless of column header name)
        :type directory_sheet: str

        :param directory_property_map: (optional) Map (dict) from column header names to DataLib property names
                                        NOTE: the dictionary keys may be a sub-set of the columns in the sheet (if so, other columns will be ignored)
                                        if None provided, then ALL columns will be used as properties, and header names will be used directly as property names
        :type directory_file: dict

        :param defaults_sheet: (optional) Excel sheetname for defaults_file
                                First row will be treated as header row
                                First column will treated as property name column (regardless of column header name)
        :type defaults_sheet: string

        :param ignore_empty_strings: (default=True) flag to control whether to process/honor empty string values when parsing a library scchema file
                                     If true, then the record is not processes and any (empty string) property input will not be assigned (will be ignored)
        :type ignore_empty_strings: boolean

        :param verbose: (optional) verbose output  flag
        :type defaults_item: string

        :param verbose: (optional) verbose output  flag
        :type defaults_item: string

        NOTE: property values in Excel cells equaling 'None' will be assigned to python None values
        ----  aka blank cells will be SKIPPED / NOT PROCESSED (and will NOT result in a None property assignments
        """

        lib = cls()  # <-- Preserves potential parent class

        # Invoke common books (if appropriate)
        # ----------------------------------------
        if (
            (defaults_book is None)
            and (defaults_sheet is not None)
            and (directory_book is not None)
            and (directory_sheet is not None)
        ):
            defaults_book = directory_book
        elif (
            (directory_book is None)
            and (directory_sheet is not None)
            and (defaults_book is not None)
            and (defaults_sheet is not None)
        ):
            directory_book = defaults_book

        # -----------------------------------------------------
        # STEP 1: PROCESS DEFAULTS BOOK FIRST (IF APPLICABLE)
        # -----------------------------------------------------
        add_defaults = _process_defaults_sheet(
            defaults_book=defaults_book,
            defaults_sheet=defaults_sheet,
            defaults_header=defaults_header,
            ignore_empty_strings=ignore_empty_strings,
            verbose=verbose,
        )

        # Apply add_defaults
        for property_name, default_value in add_defaults.items():
            lib.properties.set(property_name, default_value)

        # ------------------------------------
        # STEP 2: PROCESS DIRECTORY BOOK
        # ------------------------------------
        if directory_book is not None:
            (fieldspace, add_fields, add_properties) = _process_directory_sheet(
                directory_book=directory_book,
                directory_sheet=directory_sheet,
                directory_property_map=directory_property_map,
                ignore_empty_strings=ignore_empty_strings,
            )
            # Add fieldspace
            if fieldspace is not None:
                lib.properties.set("fieldspace", fieldspace)

            # Apply add_fields
            if len(lib.field_index) > 0:
                raise ValueError(
                    "DataLib constructor unexpectedly encountered initial self.field_index."
                )
            else:
                lib.field_index = make_index(
                    add_fields
                )  # <-- NOTE: shallow copy but OK here

            # Apply add_properties
            if len(lib.properties.all_fldspec_properties) > 0:
                raise ValueError(
                    "DataLib constructor unexpectedly encountered initial fldspec_properties dictionary."
                )
            else:
                lib.properties.set_fldspec_properties(
                    add_properties
                )  # <-- NOTE: shallow copy but OK here

        if directory_book is not None:
            lib.time_stamp = datetime.fromtimestamp(os.path.getmtime(directory_book))
        elif defaults_book is not None:
            lib.time_stamp = datetime.fromtimestamp(os.path.getmtime(defaults_book))
        else:
            raise ValueError("Neither directory_book nor defaults_book provided")

        return lib

    def to_csv(
        self,
        csvfile,
        compress=False,
        header=True,
        rowkeys=True,
        date_format="%Y-%m-%d",
        missing_format="",
        tgt_keyspace=None,
        **fmtparams,
    ):
        """Convert DataLib to CSV format.

        Conform all fields in the DataLib and convert the resultant data to CSV.
        Modeled after the Python `csv <https://docs.python.org/2/library/csv.html>`_
        module and can be used in a similar manner::

            with open('eggs.csv', 'wb') as csvfile:
                lib.to_csv(csvfile, delimiter=' ', quotechar='|')

        The output is structured such that the beginning columns are the conformal
        keyspaces (from flattening the hyper index) and the remaining columns are
        the data for each field. For example, the following DataLib::

            Keyspaces: Dates, Fractiles, Tickers
            Fields: Company Name, Market Cap, Asset Quantity

        would generate the following CSV columns::

            Dates, Fractiles, Tickers, Company Name, Market Cap, Asset Quantity

        If *rowkeys* is ``False`` then the keyspaces will be left out, so the columns
        for the above example would be::

            Company Name, Market Cap, Asset Quantity

        :type csvfile: file-like object
        :param csvfile:
            Write the DataLib to a file-like object. Can be any object with a
            ``write()`` method. If *csvfile* is a file object, it must be opened with
            the 'b' flag on platforms where that makes a difference.

        :type compress: bool
        :param compress:
            Remove rows that contain only missing data.

        :type header: bool
        :param header:
            Include a header row in the CSV output.

        :type rowkeys: bool
        :param rowkeys:
            Write the associated keyspace key(s) for each row.

        :type date_format: str
        :param date_format:
            Format for serializing dates and datetimes. Uses formats accepted by
            Python's ``strftime`` method.

        :type missing_format: str
        :param missing_format:
            Format for serializing missing values. Writes the given string into the
            output when a missing value is encountered.

        :type tgt_keyspace: str
        :param tgt_keyspace:
            Remap identifiers into this format before constructing the CSV. Accepted
            values include any security format present in the environment's security
            master.

        :param \*\*fmtparams:
            Optional *fmtparams* keyword arguments can be given to override
            individual default formatting parameters. See the Python ``csv`` module
            for more details.

        """
        from qubles.io.base.rootlib import FrameContextManager, RootLib

        if self.is_empty:
            return

        try:
            strftime.strftime(datetime.now(), date_format)
        except Exception:
            raise DataLibCSVError(f"Invalid date format: {date_format}")

        try:
            missing_format = str(missing_format)
        except UnicodeEncodeError:
            missing_format = missing_format.encode("utf8")

        # Extract conformed Qubles for each field
        # ---------------------------------------
        if not self.is_conformal:
            lib = self.conform()
        else:
            lib = self

        conformal_fields = lib.conformal_fields().tolist()
        if len(conformal_fields) == 0:
            return

        field_qubles = lib.multi_eval_od()
        if tgt_keyspace:
            with FrameContextManager():
                RootLib().set_control("secmstr_tgt_keyspace", tgt_keyspace)
                for field, quble in field_qubles.items():
                    field_qubles[field] = RootLib().remap(quble)

        # Validations:
        # ------------------------------
        if field_qubles[conformal_fields[0]].is_undefined:
            return

        # Get the conformal hyper index
        # ------------------------------
        hyper_index = field_qubles[conformal_fields[0]].hyper_index
        if hyper_index.ndim == 0:
            return

        # Precomputation and memoization to improve formatting performance
        # -----------------------------------------------------------------
        ks_missing_vals = {}
        ks_is_time = {}
        ks_fmt_cache = {}
        for i, ks in enumerate(hyper_index.keyspaces):
            dtype = hyper_index[ks].get_dtype()
            if dtype not in (bool, np.bool_):
                ks_missing_vals[i] = missing_val_by_dtype(dtype)

            ks_is_time[i] = hyper_index[ks].is_time
            ks_fmt_cache[i] = {}

        field_missing_vals = {}
        field_is_time = {}

        for field, quble in field_qubles.items():
            dtype = quble.dtype
            if dtype not in (bool, np.bool_):
                field_missing_vals[field] = missing_val_by_dtype(dtype)

            field_is_time[field] = quble.is_datetime

        def fmt(val, keyspace=None, field=None):
            if keyspace is not None:
                if val in ks_fmt_cache[keyspace]:
                    return ks_fmt_cache[keyspace][val]

                if keyspace in ks_missing_vals and isnull(
                    val, missing_value=ks_missing_vals[keyspace]
                ):
                    fmt_val = missing_format
                elif ks_is_time[keyspace]:
                    fmt_val = strftime.strftime(val, date_format)
                else:
                    try:
                        fmt_val = str(val)
                    except UnicodeEncodeError:
                        fmt_val = val.encode("utf8")

                ks_fmt_cache[keyspace][val] = fmt_val

                return fmt_val
            else:
                if field in field_missing_vals and isnull(
                    val, missing_value=field_missing_vals[field]
                ):
                    return missing_format

                if field_is_time[field]:
                    return strftime.strftime(val, date_format)

                try:
                    return str(val)
                except UnicodeEncodeError:
                    return val.encode("utf8")

        # Reduce the conformal hyper index to a single dimension, folding if necessary
        # -----------------------------------------------------------------------------
        reduced_hyper_index = hyper_index
        is_folded = False
        if hyper_index.ndim > 1:
            for field, quble in field_qubles.items():
                field_qubles[field] = quble.fold(quble.keyspaces)

            reduced_hyper_index = field_qubles[conformal_fields[0]].hyper_index
            is_folded = True

        keyspace = reduced_hyper_index.keyspaces[0]
        writer = csv_writer(csvfile, **fmtparams)

        # Write the header, if applicable
        # --------------------------------
        if header:
            if not rowkeys:
                columns = conformal_fields
            elif is_folded:
                fold_delimiter = RootLib().get_control("fold_delimiter")
                columns = keyspace.split(fold_delimiter) + conformal_fields
            else:
                columns = [keyspace] + conformal_fields

            writer.writerow(columns)

        # Write the rows (there may be millions, so proceed with caution)
        # ----------------------------------------------------------------
        for i, key in enumerate(reduced_hyper_index[keyspace]):
            if not rowkeys:
                keys = []
            elif is_folded:
                keys = [fmt(k, keyspace=j) for j, k in enumerate(key)]
            else:
                keys = [fmt(key, keyspace=0)]

            vals = []
            all_missing = True
            for field in conformal_fields:
                val = field_qubles[field].values[i]
                if compress and all_missing:
                    is_missing = field in field_missing_vals and isnull(
                        val, missing_value=field_missing_vals[field]
                    )

                    if not is_missing:
                        all_missing = False

                vals.append(fmt(val, field=field))

            if compress and all_missing:
                continue

            writer.writerow(keys + vals)

    @abstractmethod
    def has_sql_query(
        self,
        field,
        sql_command_property_name="definition",
        sql_file_property_name="query_file",
    ):
        """
        Abstract base method
        (may be overridden by derived classes)
        """
        return NotImplementedError

    @abstractmethod
    def get_sql_query(
        self,
        field,
        sql_command_property_name="definition",
        sql_file_property_name="query_file",
    ):
        """
        Abstract base method
        (may be overridden by derived classes)
        """
        return NotImplementedError

    # ============================= Registration Methods ==============================

    @property
    def is_empty(self):
        if len(self.field_index) == 0:
            return True
        else:
            return False

    def register_field(self, fieldref, property_assignments=None, like=None):
        """
        Registers a field name (or multiple fieldnames) within of a DataLib...
        Can also accept an associated dictionary of field-specific property assignments (optional)

        :param fieldref: Field name to be registered
        :type fieldref: str

        :param property_assignments: (optional) field-specific property assignments
        :type property_assignments: dict (where keys are property names)
        """
        lib = self
        # If a multiple fields were provided...
        # --------------------------------------
        if (
            isinstance(fieldref, list)
            or isinstance(fieldref, tuple)
            or isinstance(fieldref, Index)
            or isinstance(fieldref, DateRange)
        ):
            if self.field_index.is_empty:
                if isinstance(fieldref, Index) or isinstance(fieldref, DateRange):
                    self.field_index = fieldref
                else:
                    self.field_index = make_index(fieldref)
            else:
                for f in fieldref:
                    self.register_field(f)
        # If a single field was provided...
        # ----------------------------------
        elif fieldref is not None:
            field = fieldref
            # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

            # Add fieldname if possible
            if field not in lib.field_index:
                lib.field_index = lib.field_index.add_key(field)
                lib.touch()

        # Implement properties (if applicable)
        # ---------------------------------------
        if property_assignments is None:
            if like is None:
                pass
            elif like not in self.field_index:
                raise ValueError(f"Unsupported like (field) arg: {like}")
            else:
                for property_name in self.properties.all_fldspec_properties:
                    if self.properties.is_fldspec_property(property_name, field=like):
                        like_value = self.properties.get_fldspec_property(
                            property_name, like
                        )
                        self.properties.set(property_name, like_value, field=field)

        elif not isinstance(property_assignments, dict):
            raise TypeError("Bad property_assignments arg: dict instance expected")
        else:
            # If assigning a builder, also set field_category/type
            if "builder" in property_assignments:
                lib.properties.set("field_category", "data", field=field)
                lib.properties.set("field_type", "Quble", field=field)
            registration_activity = False
            for property_name, property_value in property_assignments.items():
                lib.properties.set(property_name, property_value, field=field)
                registration_activity = True

            if registration_activity:
                lib.touch()

    def vacate(self, field="<all>", grace=False):
        """Clear the library's virtual residence.

        :type field: str
        :param field: Clear only this field from residence. If '<all>', the entire
                      residence will be cleared.

        :type grace: bool
        :param grace: Set to False to avoid raising an exception when vacating a
                      field that is not registered.

        """
        if field == self.ALL_FIELDS_KEYWORD:
            self.residence = {}
        elif field not in self.field_index:
            if grace:
                _logger.warning("Warning: Unregistered field: %s", field)
            else:
                raise UnregisteredFieldError(field)
        else:
            # Use .keys() since residence dictionary may change within loop below
            for residence_key in list(self.residence.keys()):
                if residence_key is None:
                    continue

                if not self._exists_in_residence(residence_key):
                    # Deletion of a previous residence key can trigger the Quble
                    # destructor which can cause other keys to be removed
                    continue

                residence_field = residence_key[0]

                if residence_field == field:
                    self.residence.pop(residence_key)
            try:
                # Cleanup all the cached res timestamps for the given lib id
                redis_keys = Redis.keys(self.REDIS_RESIDENCE_TIMESTAMP_KEY)
                for redis_key in redis_keys:
                    if str(self.id) in str(redis_key):
                        redis_item = str(redis_key, encoding="utf-8")
                        redis_list = redis_item.split(":")
                        Redis.pop(redis_list[0], redis_list[1])

                # Cleanup all the cached file timestamps for the given lib id
                redis_keys = Redis.keys(self.REDIS_FILE_TIMESTAMP_KEY)
                for redis_key in redis_keys:
                    if str(self.id) in str(redis_key):
                        redis_item = str(redis_key, encoding="utf-8")
                        redis_list = redis_item.split(":")
                        Redis.pop(redis_list[0], redis_list[1])
            except Exception as e:
                _logger.exception(e)

    def unregister_field(self, field, grace=False, retain_residence=False):
        """
        Unregisters the specified field from the DataLib
        NOTE: operates on/modifies self
        """
        if field not in self.field_index:
            if grace:
                _logger.warning("Warning: Unregistered field: %s", field)
            else:
                raise UnregisteredFieldError(field)
        else:
            # --------------------------------------------------------------------------------------
            #       VACATE VIRTUAL INSTANCE(S) FROM THE INTERNAL DICTIONARY (IF APPLICABLE)
            # --------------------------------------------------------------------------------------
            # When the <fieldname> element instances is 'popped' from the dictionary,
            # it will be eventually be destroyed (so long as not referenced from outside lib's dict),
            # at which time the assocoated destructor will be called
            # which will in-turn (hopefully) call <object>.toLib()
            # to persist (via file storage if applicable) any recent changes to instance before destruction
            # --------------------------------------------------------------------------------------
            if not retain_residence:
                self.vacate(field)

            if self.get_property(
                "file_type", field
            ) == "SNOWFLAKE" and not self.virtual_storage_only(field):
                self.field_to_id_map.pop(field)

            # Remove related field-specific properties
            for property_name in self.properties.all_fldspec_properties:
                if self.properties.is_fldspec_property(property_name, field=field):
                    self.properties.remove_fldspec_property(property_name, field=field)

            # Update the self.field_index
            self.field_index = self.field_index.pop_key(field)

            self.touch()

    def unregister_all_fields(self):
        """
        Unregisters all fields from the DataLib
        NOTE: operates on/modifies self
        """

        self.vacate()

        # Remove all field-specific properties
        self.properties.set_fldspec_properties({})

        self.field_index = Index(dtype=self.field_index.dtype)
        self.field_to_id_map = {}
        self.touch()

    def register_lib_from_xl(
        self,
        lib_field,
        directory_book,
        directory_sheet,
        directory_property_map=None,
        defaults_book=None,
        defaults_sheet=None,
        defaults_header=None,
        verbose=False,
    ):
        """
        Registers a data library from a single Excel workbook (master schema)
           Uses lib_fields (string, list or tuple) provided to identify for directory_sheet names as well as column headers in defaults_sheet
        Will return errors if:
             1) some directory_sheets not present for one of lib_fields provided
             2) defaults_sheet provided, but some headers absent in default sheet via-a-vis lib_fields provided

        If directory_property_map provided, it will be applied in common for all directory sheets

        :param lib_field: Library field underwhich to register the library
        :type lib_fields: string

        :param directory_book: Excel workbook file containing the DataLib directory information
        :type filename: str

        :param directory_sheet: Excel sheetname containing the DataLib directory and associated (per field) property info
                                First row will be treated as header row
                                First column will treated as field name column (regardless of column header name)
        :type directory_sheet: str

        :param directory_property_map: (optional) Map (dict) from column header names to DataLib property names
                                        NOTE: the dictionary keys may be a sub-set of the columns in the sheet (if so, other columns will be ignored)
                                        if None provided, then ALL columns will be used as properties, and header names will be used directly as property names
        :type directory_file: dict

        :param defaults_sheet: (optional) Excel sheetname for defaults_file
                                First row will be treated as header row
                                First column will treated as property name column (regardless of column header name)
        :type defaults_sheet: string

        :param defaults_header: (optional) header name used to locate the proper column in defaults sheet
        :type defaults_header: string

        :param verbose: (optional) verbose output  flag
        :type defaults_item: string
        """
        if lib_field is None or not isinstance(lib_field, str) or len(lib_field) == 0:
            lib_field = directory_sheet

        _logger.info(f"Registering Data Library: {lib_field}...")
        if verbose:
            if defaults_book is not None:
                _logger.debug(
                    "   ==> DEFAULTS: header: {0} in sheet: {1} within file: {2}".format(
                        defaults_header, defaults_sheet, defaults_book
                    )
                )
            _logger.debug(
                "   ==> PROPERTIES: sheet: {0} within file: {1}".format(
                    directory_sheet, directory_book
                )
            )

        new_lib = DataLib.from_xl(
            directory_book=directory_book,
            directory_sheet=directory_sheet,
            directory_property_map=directory_property_map,
            defaults_book=defaults_book,
            defaults_sheet=defaults_sheet,
            defaults_header=defaults_header,
            verbose=verbose,
        )

        self.register_field(
            lib_field
        )  # <-- Otherwise, self.get_property('auto_registration') will be checked
        self.set(lib_field, new_lib)
        self[lib_field].address = LibAddress(lib_field, self)

    def register_libs_from_xl(
        self,
        lib_fields,
        book,
        directory_property_map=None,
        defaults_sheet=None,
        verbose=False,
    ):
        """
        Registers multiple DataLib schemas from a single Excel workbook (master schema)
           Uses lib_fields (string, list or tuple) provided to identify for directory_sheet names as well as column headers in defaults_sheet
        Will return errors if:
             1) some directory_sheets not present for one of lib_fields provided
             2) defaults_sheet provided, but some headers absent in default sheet via-a-vis lib_fields provided

        If directory_property_map provided, it will be applied in common for all directory sheets

        :param lib_fields: Library fields underwhich to register the requisite libraries
                         NOTE: These fields will be used to find associated directory sheets & columns headers within default_sheet (if applicable)
                         failure to find necessary directory & default header information for each lib_field will result in an error
        :type lib_fields: list, tuple or scalar of string(s)

        :param book: Excel workbook file containing the master schema *directories & defaults) for the DataLibs
        :type book: str

        :param directory_sheet: Excel sheetname containing the DataLib directory and associated (per item) info
        :type directory_sheet: str

        :param directory_property_headers: (optional) Map (dict) from column header names to DataLib property names
                                        NOTE: the dictionary keys may be a sub-set of the columns in the sheet (if so, other columns will not be used)
                                        if None is provided, then all the columns of each directory sheet will be processed and entered as properties.
                                        NOTE: if not None, then directory_property_headers will be applied in COMMON to all directories
        :type directory_file: dict

        :param defaults_sheet: (optional) Excel sheetname for defaults_file
        :type defaults_item: string

        :param verbose: (optional) verbose output  flag
        :type defaults_item: string
        """
        # If this DataLib is currently empty,
        # assign (assume) default 'field_category'='lib'
        # --------------------------------------------------
        if self.is_empty:
            self.set_property("field_category", "lib")

        # If lib_fields arg is scalar, convert to list
        if isinstance(lib_fields, str):
            lib_fields = [lib_fields]

        # Set defaults_book based on presence of defaults_sheet (not None)
        if defaults_sheet is None:
            defaults_book = None
        else:
            defaults_book = book

        # Loop through the lib_fields and load & register each data library
        for lib_field in lib_fields:
            _logger.info(f"Registering Data Library: {lib_field}...")
            if verbose:
                if defaults_book is not None:
                    _logger.debug(
                        "   ==> DEFAULTS: header: {0} in sheet: {1} within file: {2}".format(
                            lib_field, defaults_sheet, defaults_book
                        )
                    )
                _logger.debug(
                    f"   ==> PROPERTIES: sheet: {lib_field} within file: {book}"
                )

            new_lib = DataLib.from_xl(
                directory_book=book,
                directory_sheet=lib_field,
                directory_property_map=directory_property_map,
                defaults_book=defaults_book,
                defaults_sheet=defaults_sheet,
                defaults_header=lib_field,
                verbose=verbose,
            )
            self.register_field(
                lib_field
            )  # <-- Otherwise, self.get_property('auto_registration') will be checked
            self.set(lib_field, new_lib)
            self[lib_field].address = LibAddress(lib_field, self)

    # =============================== Property Methods ================================

    def property_list(self):
        """List of property names.

        Provides a list of all recorded properties within the DataLib by crafting the
        UNION of default properties and field-specific properties.

        NOTE: Does NOT support/include native properties

        """
        return list(self.properties.all_properties)

    def has_property(
        self,
        property_name,
        fieldref=None,
        default_check=True,
        fldspec_check=True,
        native_check=True,
    ):
        """Return ``True`` if the property assignment exists.

        Checks whether specified property is currently a recorded property (across
        both default_properties, fldspec_properties & native property test)

        """
        if property_name is None:
            return False
        elif default_check and self.properties.is_default_property(property_name):
            return True
        elif fieldref is None:
            return False

        # Get lib & field...
        # -------------------
        lib = self
        field = fieldref

        # Apply appropriate logic to lib
        # -------------------------------
        if default_check and self.properties.is_default_property(property_name):
            return True
        elif (
            field is not None
            and fldspec_check
            and lib.properties.is_fldspec_property(property_name, field=fieldref)
        ):
            return True
        elif native_check and lib.is_native_property(property_name):
            return True
        else:
            return False

    def get_property_dict_element(
        self,
        property_name,
        dict_key,
        fieldref=None,
        grace=False,
        default_property_value=None,
        resolve_templates=False,
        template_brackets=("<", ">"),
        alt_keyspace=None,
        native_property_mode=NPModes.SECONDARY,
        appeal=False,
        suppress_recording=None,
        document_as_field_dep=False,
        **kwargs,
    ):
        """
        Returns the dictionary element (for key=dict_key)
        from a property assignment that is presumed to be a dictionary
        """
        # Validate dict_key
        if dict_key is not None:
            pass
        elif grace:
            return None
        else:
            raise ValueError("No dict_key")
        # Access the underlying property (as a dictionary if possible)
        property_value = self.get_property(
            property_name=property_name,
            fieldref=fieldref,
            grace=grace,
            default_property_value=default_property_value,
            resolve_templates=resolve_templates,
            template_brackets=template_brackets,
            alt_keyspace=alt_keyspace,
            native_property_mode=native_property_mode,
            appeal=appeal,
            suppress_recording=suppress_recording,
            document_as_field_dep=document_as_field_dep,
            try_json_loads=True,
            **kwargs,
        )

        if property_value is None:
            return None
        elif not isinstance(property_value, dict):
            raise TypeError(
                "For property_name:{0}: non-dictionary property_value:{1}".format(
                    property_name, property_value
                )
            )
        elif dict_key in property_value:
            return property_value[dict_key]
        elif grace:
            return None
        else:
            raise ValueError(
                "dict_key:{0} absent from (dictionary) property_value:{1}".format(
                    dict_key, property_value
                )
            )

    def get_active_connection_profile(self, fieldref: str = None):
        return self.get_property(
            "connection_profile",
            fieldref,
            grace=True,
            default_property_value=os.environ["DEFAULT_CONNECTION_PROFILE"],
        )

    def get_property(
        self,
        property_name,
        fieldref=None,
        grace=False,
        default_property_value=None,
        resolve_templates=False,
        template_brackets=("<", ">"),
        alt_keyspace=None,
        native_property_mode=NPModes.SECONDARY,
        appeal=False,
        suppress_recording=None,
        document_as_field_dep=False,
        try_json_loads=False,
        **kwargs,
    ):
        """Return the associated property value.

        :type property_name: str
        :param property_name:
            Retrieve the value for this property

        :type fieldref: str or list or tuple
        :param fieldref:
            Retrieve the property value for this field, if it is available. If
            *fieldref* is ``None`` or there is no field-specific property assignment,
            the default property is consulted.

        :type grace: bool
        :param grace:
            Ignore exceptions raised by accessing properties that don't exist.

        :type default_property_value: any
        :param default_property_value:
            If *grace* is ``True`` and the requested property does not exist, return
            this value.

        :type resolve_templates: bool
        :param resolve_templates:
            Return the property value with all templated tokens replaced by their
            associated property values.

        :type template_brackets: 2-tuple
        :param template_brackets:
            The start and end characters that denote a templated token.

        :type alt_keyspace: str
        :param alt_keyspace:
            If the property value is a RefLib name, return the alt_keyspace if it is
            found to be a source keyspace within the RefLib.

        :type native_property_mode: str
        :param native_property_mode:
            Control how native properties are invoked. ``None`` means native
            properties will not be consulted. "exclusive" means the native property
            method will *always* be called. "primary" means the native property method
            will be called if it exists, otherwise the stored property value will be
            returned. "secondary" means the system will only invoke the native
            property method if a stored property does not exist.

        :type appeal: bool
        :param appeal:
            If no property assignment is found, consult ancestor library chain if present.
            When no ancestors are present (i.e., lib does not have an address),
            appeal directly to the RootLib()

        :type suppress_recording: bool
        :param suppress_recording:
            Return the property value *without* recording the property as a property
            dependency.

        :type document_as_field_dep: bool
        :param document_as_field_dep:
            Indicate that this property value is referencing an addressable object
            and record it as a dynamic field dependency.

        :param \*\*kwargs:
            When a native property method is invoked, the keyword arguments will be
            passed as kwargs to that method.

        :type try_json_loads: bool
        :param try_json_loads: Flag to attempt json loads
                               for string/unicode-based property values
        """
        from qubles.core.quble import Quble
        from qubles.io.base.rootlib import RootLib

        if suppress_recording is None:
            if self.has_property_metadata("recording_flag", property_name):
                suppress_recording = not self.get_property_metadata(
                    "recording_flag",
                    property_name,
                )
            else:
                suppress_recording = False

        # Handle fieldref list/tuple args by recursively calling
        # get_property() with single elements of list/tuple
        # Otherwise assume a scalar fieldref is provided]
        # ------------------------------------------------------
        if isinstance(fieldref, (list, tuple)):
            properties = []
            for fref in fieldref:
                properties.append(
                    self.get_property(
                        property_name,
                        fieldref=fref,
                        grace=grace,
                        default_property_value=default_property_value,
                        resolve_templates=resolve_templates,
                        template_brackets=template_brackets,
                        alt_keyspace=alt_keyspace,
                        native_property_mode=native_property_mode,
                        appeal=appeal,
                        suppress_recording=suppress_recording,
                        document_as_field_dep=document_as_field_dep,
                        **kwargs,
                    )
                )

            return properties

        lib = self
        field = fieldref
        # -------------------------------------
        # Handle "linked" 'LIBADDRESS' Case
        # -------------------------------------
        if (
            (property_name in ("field_dependencies", "property_dependencies"))
            and (lib.get_property("access_mode", field, grace=True) is not None)
            and (lib.get_property("file_type", field, grace=True) == "LIBADDRESS")
        ):
            srcdomain, srcfield, _ = lib.lib_address_credentials(field)
            return srcdomain.get_property(
                property_name,
                fieldref=srcfield,
                grace=grace,
                default_property_value=default_property_value,
                resolve_templates=resolve_templates,
                template_brackets=template_brackets,
                alt_keyspace=alt_keyspace,
                native_property_mode=native_property_mode,
                appeal=appeal,
                suppress_recording=suppress_recording,
                document_as_field_dep=document_as_field_dep,
                **kwargs,
            )

        if (field is not None) and (field not in lib.field_index) and not grace:
            raise UnregisteredFieldError(f"{self} does not contain field: {field}")

        try:
            property_value = self.properties.get(
                property_name,
                field=field,
                native_property_mode=native_property_mode,
                **kwargs,
            )
        except MissingProperty:
            # Appeal the property request (if directed to do so)
            # ---------------------------------------------------
            if appeal and not lib.is_root:
                # NOTE: Intentionally not propagating a field-specific property
                # request here, only a default property request [Generally assuming
                # parental libs will not support same field names...trying to avoid
                # field/key errors]
                #
                # NOTE: Inclusion of kwargs was causing problems for this method
                # resubmission due to explicit setting of fieldref arg earlier in
                # signature
                # ------------------------------------------------------------------
                appellate_lib = (
                    lib.parent_lib if lib.parent_lib is not None else RootLib()
                )
                property_value = appellate_lib.get_property(
                    property_name=property_name,
                    fieldref=None,
                    grace=grace,
                    default_property_value=default_property_value,
                    resolve_templates=resolve_templates,
                    template_brackets=template_brackets,
                    alt_keyspace=alt_keyspace,
                    native_property_mode=native_property_mode,
                    appeal=appeal,
                    suppress_recording=True,
                    document_as_field_dep=document_as_field_dep,
                )
            elif grace:
                property_value = default_property_value
            else:
                raise

        # Invoke resolve_templates if applicable...
        # -------------------------------------------
        if (
            resolve_templates
            and isinstance(property_value, str)
            and search(template_brackets[0], property_value)
        ):
            # Assign local variables...
            # ---------------------------
            orig_property_value = property_value
            replace_ctr = None
            num_iterations = 0
            max_iterations = 10

            # Iteratively process property_value variable by trying to replacing
            # templated items with associated property values...
            #
            # NOTE: This technique allows the replacements of templated items to
            # themselves contain tempated items. Example:
            #
            #     lib.get_property('path','MY_FACTOR') = '<path1>'
            #         WHERE lib.get_property('path1','MY_FACTOR') = 'C:\<field>\'
            #
            # --------------------------------------------------------------------
            match_obj = search(r"\<(\*?\w+)\>", property_value)

            if match_obj is None:
                match_obj = search(r"\<(\$\w+)\>", property_value)

            while (match_obj is not None) and (num_iterations < max_iterations):
                replace_ctr = 0
                num_iterations += 1
                matched_pattern = match_obj.group(0)  # <-- (first) matched pattern
                if len(matched_pattern) < 2:
                    raise TemplateError(
                        "template match error...resultant match: {0} expected to be "
                        "atleast length 2".format(matched_pattern)
                    )
                elif (matched_pattern[0] != template_brackets[0]) or (
                    matched_pattern[-1] != template_brackets[1]
                ):
                    raise TemplateError(
                        "template match error...resultant match: {0} should "
                        "start/end with brackets: {1} & {2}".format(
                            matched_pattern, template_brackets[0], template_brackets[1]
                        )
                    )

                templated_property_name = matched_pattern[1:-1]

                if len(templated_property_name) == 0:
                    replacement_val = ""
                elif templated_property_name[0] == "$":
                    # Handle environment variable (May throw KeyError)
                    env_var = templated_property_name[1:]
                    replacement_val = os.environ[env_var]
                elif templated_property_name[0] == "*":
                    # Handle RootLib property/control consultation
                    # Here, we strip off the leading asterisk, then consult RootLib to evaluate associated property name
                    rootlib_property_name = templated_property_name[1:]
                    replacement_val = RootLib().get_property(
                        rootlib_property_name,
                        grace=False,
                        resolve_templates=True,
                        appeal=True,
                    )
                elif templated_property_name == "field":
                    replacement_val = field
                else:
                    replacement_val = lib.get_property(
                        templated_property_name,
                        field,
                        grace=False,
                        resolve_templates=True,
                        appeal=True,
                    )

                # Allow conversion of integer replacement_val to string
                if isinstance(replacement_val, int):
                    replacement_val = str(replacement_val)

                if property_name == "filename" and templated_property_name == "univ":
                    if replacement_val is None:
                        replacement_val = "NO_UNIV"

                    # Make universe address safe for filenames. This opens up a chance of filename collisions if part of a
                    # universe address already has the delimiter, but the chances are small enough we are ok taking this risk.
                    replacement_val = replacement_val.replace("|", "+")

                property_value, num_local_replaces = subn(
                    escape(matched_pattern),
                    lambda x: replacement_val,  # lambda fixes backslash escape issues
                    property_value,
                )

                replace_ctr += num_local_replaces

                # Next match iteration
                match_obj = search(r"\<(\w+)\>", property_value)

                if match_obj is None:
                    match_obj = search(r"\<(\$\w+)\>", property_value)

            # ------------------------------------------------------------------------
            # Here one of three following states should hold:
            #
            #     1) no templace_bracket characters (e.g., '<') remain in property
            #        value: GOOD Outcome
            #  OR 2) no replacements were made on most recent iteration: BAD Outcome
            #        (some templated items cannot be resolved)
            #  OR 3) max iterations exceeded: BAD Outcome (as a circular loop may be
            #        present)
            # ------------------------------------------------------------------------

            # Trap for bad outcomes
            # ----------------------
            if num_iterations >= max_iterations:
                raise TemplateError(
                    "Max template iterations {0} exceeded: Possibly circular loop? "
                    "for: get_property({1}, {2})".format(
                        max_iterations, property_name, field
                    )
                )
            elif replace_ctr == 0:
                raise TemplateError(
                    "Unsupported or improper template usage: original:{0}...final: "
                    "{1}".format(orig_property_value, property_value)
                )

        # For string/unicode property values, consider json.loads()
        if try_json_loads and isinstance(property_value, str):
            # Try json loads if directed
            try:
                property_value = loads(property_value)
            except:
                pass

        # Try to invoke reflibs if applicable
        # ------------------------------------
        if (property_value is not None) and (alt_keyspace is not None):
            if isinstance(alt_keyspace, Quble):  # <-- To be slightly more versatile
                alt_keyspace = alt_keyspace.keyspaces

            for reflib in RootLib().get_reflibs(recursive=False):
                reflib_field = reflib.address.min_field
                if property_value == reflib_field:
                    property_value = reflib.get_src_keyspace(src_keyspace=alt_keyspace)
                    break

        # Translate string field_fresh_time into datetime object
        # --------------------------------------------------------
        fresh_time_props = [
            "field_fresh_time",
            "field_fresh_time_override",
            "force_tgt_keys",
        ]

        if property_name in fresh_time_props and isinstance(property_value, str):
            # Build a dictionary look-up from (upper-cased) property_value to callable function
            property_value_upper = property_value.upper()

            if property_value_upper in FIELD_FRESH_TIME_OVERRIDE_FN_DICT:
                property_value = FIELD_FRESH_TIME_OVERRIDE_FN_DICT[
                    property_value_upper
                ]()  # <-- Call appopriate function with no args
            elif property_name == "force_tgt_keys":
                pass
            else:
                raise ValueError(
                    "property: field_fresh_time. string encountered:{0}".format(
                        property_value_upper
                    )
                )

        recorder = DataLib.dependency_recorder
        # Update property recorder
        # -------------------------
        if (
            not suppress_recording
            and property_name not in self.PROPERTY_RECORDER_PROHIBITION_LIST
            and recorder.current_address is not None
            and (
                recorder.current_address.base_domain == lib
                or (lib.is_root and field is None)
            )
        ):
            is_global = lib.is_root and not recorder.current_address.base_domain.is_root
            is_fldspec = (
                recorder.current_address.base_domain == lib and field is not None
            )

            if is_fldspec:
                recorder_field = recorder.current_address[-1]
                field_arg = (
                    ""
                    if ((recorder_field == field) and not recorder.fldspec_prop_tagging)
                    else field
                )
            else:
                field_arg = None

            annotated_name = self._annotate_property_dependency(
                property_name,
                is_global=is_global,
                appeal=appeal,
                resolve_templates=resolve_templates,
                is_field_dependency=document_as_field_dep,
                field_arg=field_arg,
            )

            recorder.track_access_for_property(annotated_name)

        return property_value

    def set_property(
        self,
        property_name,
        property_value,
        fieldref=None,
        redefine="auto_redefine",
        freeze=False,
        freeze_resolution="exception",
    ):
        """Assign a property of a DataLib.

        :type property_name: str
        :param property_name: The name of the property.

        :type property_value: any
        :param property_value: The value to assign.

        :type fieldref: str
        :param fieldref: Perform a field-specific assignment. Accepts the following:None
        The assignment will be applied to the default settings.
           '<all>'
                1) Assign default property (but do not create if originally absent)
                2) Assign all field-specific fields (but do not create if originally
                   absent)
           '<all_default>'
                1) Assign default property (and create if originally absent)
                2) Assign all field-specific fields (but do not create if originally
                   absent)
           '<all_fs>'
                1) Assign default property if present (but do not create if originally
                   absent)
                2) Assign all field-specific fields (and create if originally absent)
            str
                The name of the field used for the field-specific assignment.

        :type redefine: bool or str
        :param redefine: Touches the 'field_defn_time' property if 1) the property
                         name is in REDEFINE_PROPERTY_LIST and 2) the property value
                         changes. If redefine='auto_redefine', the library's
                         'auto_redefine' property will be consulted.

        :type freeze: bool
        :param freeze: Freeze the control on the given value for this frame and any
                       child frames. Any subsequent calls to ``set_property`` for the
                       given control will be ignored, until the current frame has been
                       popped off the stack.

        :type freeze_resolution: str
        :param freeze_resolution: How to behave when the property is frozen:
            'exception': Throw an exception
            'ignore': Do not set the property
            'override': Set the property even if it is frozen. USE WITH CAUTION!

        """
        from qubles.io.base.rootlib import RootLib

        field = fieldref

        if is_quble(property_value):
            raise ValueError(f"Cannot assign Qubles as properties! DataLib={self}")

        if property_name is None:
            return

        if property_name in self.NATIVE_PROPERTY_NONSTORE_LIST:
            raise ValueError(f"native only (no store) property: {property_name}")

        if self.is_root:
            DataLib.dependency_recorder.track_root_property_submission(property_name)

        # Handle non-root assignment for CONTROL_PROPERTY_LIST
        # -----------------------------------------------------
        if property_name in self.CONTROL_PROPERTY_LIST and not self.is_root:
            if field is not None:
                raise ValueError(
                    f"Restricted (Global-Only) property_name cannot be field-specific: {property_name}"
                )

            _logger.warning(
                f"Restricted (Global-Only) property_name: {property_name}...will be applied to "
            )

            RootLib().set_property(
                property_name=property_name,
                property_value=property_value,
                fieldref=field,
                redefine=redefine,
                freeze=freeze,
                freeze_resolution=freeze_resolution,
            )

            return

        # Resolve 'redefine' parameter
        # -----------------------------
        if redefine == "auto_redefine":
            redefine = self.get_property(
                "auto_redefine", grace=True, suppress_recording=True
            )

        redefine = redefine and property_name in self.REDEFINE_PROPERTY_LIST

        # Assign default property (if applicable)
        # ----------------------------------------
        if field in (None, self.ALL_FIELDS_KEYWORD, "<all_default>", "<all_fs>"):
            if property_name in self.DEFAULT_PROPERTY_PROHIBITION_LIST:
                raise ValueError(
                    f"Assigning default property prohibited: {property_name}"
                )

            if field in (self.ALL_FIELDS_KEYWORD, "<all_fs>"):
                allow = self.properties.is_default_property(property_name)
            else:
                allow = True

            if allow:
                self.set_default(
                    property_name,
                    property_value,
                    redefine=redefine,
                    freeze=freeze,
                    freeze_resolution=freeze_resolution,
                )

        # Assign field-specific property (if applicable)
        # -----------------------------------------------
        if field is None:
            return

        if property_name in self.FIELD_PROPERTY_PROHIBITION_LIST:
            raise ValueError(f"Assigning field property prohibited: {property_name}")

        if field == "<all_fs>":
            fields = self.fields()
        elif field in (self.ALL_FIELDS_KEYWORD, "<all_default>"):
            fields = self.properties.fields_with_property(property_name)
        else:
            fields = [field]

        for f in fields:
            if f not in self.field_index:
                if self.get_property("auto_register", grace=True):
                    self.register_field(f)
                else:
                    raise UnregisteredFieldError(f)

            if redefine:
                if self.properties.is_fldspec_property(property_name, f):
                    old_value = self.properties.get_fldspec_property(property_name, f)
                else:
                    old_value = None

            self.properties.set(
                property_name,
                property_value,
                field=f,
                freeze=freeze,
                freeze_resolution=freeze_resolution,
            )

            if redefine and old_value != property_value:
                self.set_property("field_defn_time", datetime.now(), f)

            self.touch()

    def remove_property(
        self, property_name, fieldref=None, redefine="auto_redefine", grace=False
    ):
        """Remove a property of a DataLib.

        if fieldref is None, removes default settings (but may leave field-specific
        property settings)

        if fieldref = '<all>', default settings and any existing field-specific
        property settings with be removed

        if fieldref is provided and fieldref != '<all>', the field-specific property
        will be removed (if present)

        """
        if property_name is None:
            return
        # Assign default property (if applicable)
        # -----------------------------------------
        elif (fieldref is None) or (fieldref == self.ALL_FIELDS_KEYWORD):
            self.remove_default(property_name, redefine=redefine)
            self.touch()
            if fieldref is None:
                return
            else:
                lib = self
                field = fieldref
        else:
            lib = self
            field = fieldref
            # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        # Assign field-specific property (if applicable)
        # -------------------------------------------------
        if (
            field == self.ALL_FIELDS_KEYWORD
        ):  # <-- Intentionally not an 'elif' condition
            if lib.properties.is_fldspec_property(property_name):
                affected_fields = []
                if property_name in self.REDEFINE_PROPERTY_LIST:
                    for field1 in lib.fields_with_property(property_name):
                        if redefine == "auto_redefine":
                            redefine = lib.get_property("auto_redefine", field)
                            if redefine is None:
                                redefine = False

                        if redefine:
                            affected_fields.append(field1)

                lib.properties.remove_fldspec_property(property_name)
                if property_name in self.REDEFINE_PROPERTY_LIST:
                    for field1 in affected_fields:
                        lib.set_property("field_defn_time", datetime.now(), field1)
                lib.touch()
        elif field not in lib.field_index:
            if grace:
                return
            else:
                raise UnregisteredFieldError(field)
        elif lib.properties.is_fldspec_property(property_name, field=field):
            # Record a (re)definition (if applicable)...
            # ----------------------------------------------
            if redefine == "auto_redefine":
                redefine = lib.get_property("auto_redefine", field)
                if redefine is None:
                    redefine = False

            if redefine and property_name in self.REDEFINE_PROPERTY_LIST:
                old_property = lib.properties.get_fldspec_property(property_name, field)

            # Change the field-specific property value
            lib.touch()
            lib.properties.remove_fldspec_property(property_name, field=field)

            if redefine and property_name in self.REDEFINE_PROPERTY_LIST:
                if old_property != lib.get_property(property_name, field):
                    lib.set_property("field_defn_time", datetime.now(), field)

    def validate_property(self, property_name, property_value):
        """Raise a ``PropertyValueError`` if the given property value is invalid."""
        self.properties.validate(property_name, property_value)

    def set_default(
        self,
        property_name,
        default_value,
        redefine="auto_redefine",
        freeze=False,
        freeze_resolution="exception",
    ):
        """Assign a default property of a DataLib.

        :type property_name: str
        :param property_name: The name of the property.

        :type property_value: any
        :param property_value: The value to assign.

        :type redefine: bool or str
        :param redefine: Touches the 'field_defn_time' property if 1) the property
                         name is in REDEFINE_PROPERTY_LIST and 2) the property value
                         changes. If redefine='auto_redefine', the library's
                         'auto_redefine' property will be consulted.

        :type freeze: bool
        :param freeze: Freeze the control on the given value for this frame and any
                       child frames. Any subsequent calls to ``set_default`` for the
                       given control will be ignored, until the current frame has been
                       popped off the stack.

        :type freeze_resolution: str
        :param freeze_resolution: How to behave when the property is frozen:
            'exception': Throw an exception
            'ignore': Do not set the property
            'override': Set the property even if it is frozen. USE WITH CAUTION!

        """
        if property_name is None:
            return

        if property_name in self.DEFAULT_PROPERTY_PROHIBITION_LIST:
            raise ValueError(f"Assigning default property prohibited: {property_name}")

        # Handle 'view' property as special case, since it is not
        # saved in the metadata and since it is usually a Quble
        # --------------------------------------------------------
        if property_name == "view":
            self.properties.set(
                property_name,
                default_value,
                freeze=freeze,
                freeze_resolution=freeze_resolution,
            )
            return

        # Check to see if new default matches existing default
        # -----------------------------------------------------
        if self.properties.is_default_property(property_name):
            existing_default = self.properties.get_default_property(property_name)

            if is_quble(existing_default) or is_quble(default_value):
                if existing_default is default_value:
                    return
            else:
                if existing_default == default_value:
                    return

        # Record a (re)definition (if applicable)
        # ----------------------------------------
        if (
            property_name not in self.FIELD_PROPERTY_PROHIBITION_LIST
            and property_name in self.REDEFINE_PROPERTY_LIST
        ):
            if redefine == "auto_redefine":
                redefine = self.get_property(
                    "auto_redefine", grace=True, suppress_recording=True
                )
                if redefine is None:
                    redefine = False

            # If making REDEFINE_PROPERTY_LIST changes, first record
            # old settings by calling and saving get_property()
            # -------------------------------------------------------
            if redefine:
                fields_inheriting_this_default = []
                for fieldname1 in self.field_index:
                    if self.properties.is_fldspec_property(property_name, fieldname1):
                        fields_inheriting_this_default.append(fieldname1)

                # If making REDEFINE_PROPERTY_LIST changes, check whether old properties (for each fieldname) were changed
                # If so, assign a new defn_time for the fieldname
                new_defn_time = datetime.now()
                for fieldname in fields_inheriting_this_default:
                    self.set_property("field_defn_time", new_defn_time, fieldname)

        # Now implement the actual default property update
        # -------------------------------------------------
        self.touch()

        self.properties.set(
            property_name,
            default_value,
            freeze=freeze,
            freeze_resolution=freeze_resolution,
        )

    def remove_default(self, property_name, redefine="auto_redefine"):
        """
        Removes a default property of a DataLib...
        """
        if self.properties.is_default_property(property_name):
            # Record a (re)definition (if applicable)...
            # ----------------------------------------------
            if redefine == "auto_redefine":
                redefine = self.get_property("auto_redefine")
                if redefine is None:
                    redefine = False

            self.touch()
            self.properties.remove_default_property(property_name)

    def native_storage(self, fieldref, case_sensitive=False, **kwargs):
        """
        Is this field to be stored in 'native' database format?
        (i.e., the Quble native database environment)
        """
        from qubles.core.quble import QUBLE_CONNECTION_TYPE

        file_type = self.get_property(
            property_name="file_type", fieldref=fieldref, **kwargs
        )
        if file_type is None:
            return False
        elif case_sensitive:
            return file_type == QUBLE_CONNECTION_TYPE
        else:
            return file_type.lower() == QUBLE_CONNECTION_TYPE.lower()

    def has_property_metadata(self, metadata_name, property_name):
        return self.properties.has_metadata(
            metadata_name,
            property_name,
        )

    def get_property_metadata(self, metadata_name, property_name, grace=True):
        return self.properties.get_metadata(
            metadata_name,
            property_name,
            grace=grace,
        )

    def register_control(self, property_name):
        """
        Registers a property name as a control property.
        """
        if property_name not in self.CONTROL_PROPERTY_LIST:
            self.CONTROL_PROPERTY_LIST.append(self.CONTROL_PROPERTY_LIST)

    def unregister_control(self, property_name):
        """
        Un-registers the specified property name the control list (if applicable)
        """
        if property_name in self.CONTROL_PROPERTY_LIST:
            self.CONTROL_PROPERTY_LIST.pop(self.CONTROL_PROPERTY_LIST)

    def get_control(self, property_name, fieldref=None, **kwargs):
        if property_name not in self.CONTROL_PROPERTY_LIST:
            raise ControlPropertyError(f"{property_name} is not a control property")

        if fieldref is not None:
            raise ControlPropertyError(
                "Control properties do not support field-specific assignments"
            )

        return self.get_property(property_name, fieldref=fieldref, **kwargs)

    def set_control(self, property_name, property_value, fieldref=None, **kwargs):
        """Assign a DataLib control property.

        Controls are reserved default properties. Calling ``set_control`` with an
        unregistered control with throw an error. The purpose of controls is to
        provide runtime error checking to prevent property name typos.

        NOTE: This wraps ``set_property`` and shares the same parameters.

        """
        if property_name not in self.CONTROL_PROPERTY_LIST:
            raise ControlPropertyError(f"{property_name} is not a control property")
        elif fieldref is not None:
            raise ControlPropertyError(
                "Control properties do not support field-specific assignments"
            )

        return self.set_property(
            property_name, property_value, fieldref=fieldref, **kwargs
        )

    def get_definition(self, fieldref):
        """Get the code definition for a field.

        Retrieves the definition using one of the following methods:
        #. Reading it from table using the stored definition_id
        #. Introspecting the source code of the builder function

        :type fieldref: str
        :param fieldref: The name of the field

        :rtype: tuple
        :returns: (code definition (or empty string), def_type (python, sql, etc., None))
        """
        def_type = None
        def_id = self.get_property(
            "definition", fieldref, grace=True, resolve_templates=True
        )

        builder = self.get_property(
            "builder", fieldref, grace=True, resolve_templates=True
        )

        if def_id:
            defn, def_type = definition_select(def_id)
        elif builder:
            builder_func = getattr(self, builder)
            defn = getsourcelines(builder_func)
            defn = "".join([sub(r"^ {4}", "", line) for line in defn[0]]).rstrip()
        else:
            defn = ""

        return defn, def_type

    def get_builder_definition(self, fieldref):
        """Get the builder code definition for a field.

        Retrieves the builder definition by introspecting the source code of the builder function

        :type fieldref: str
        :param fieldref: The name of the field

        :rtype: str
        :returns: code definition (or empty string)
        """

        builder = self.get_property(
            "builder", fieldref, grace=True, resolve_templates=True
        )

        if builder:
            builder_func = getattr(self, builder)
            defn = getsourcelines(builder_func)
            defn = "".join([sub(r"^ {4}", "", line) for line in defn[0]]).rstrip()
        else:
            defn = ""

        return defn

    def set_definition(
        self,
        def_name="<def_name>",
        code="<code>",
        fieldref="<fieldref>",
        def_type="<def_type>",
    ):
        """Sets the definition for a field.

        The definition will be stored in a Python file at the path specified by
        the ``definition`` property. If there is already a file at this path, a
        suffix will be appended to the existing file.

        :param def_name: Definition name
        :type def_name: str

        :param code: Python/SQL/etc. code
        :type code: str

        :param fieldref: The name of the field
        :type fieldref: str

        :param def_type: Type of definition, ex: python, sql, etc.
        :type def_type: str
        """
        # Marks the field as stale and in need of a rebuild
        self.set_property("field_defn_time", datetime.now(), fieldref)

        if def_name == "<def_name>":
            def_name = self.get_property(
                "builder", fieldref, grace=True, resolve_templates=True
            )

        if def_type == "<def_type>":
            def_type = self.get_property(
                "def_type", fieldref, grace=True, resolve_templates=True
            )
        if def_type is None:
            def_type = "python"

        return sql_set_definition(
            self.get_property("definition", fieldref, grace=True),
            def_name,
            code,
            def_type=def_type,
        )

    def delete_definition(self, fieldref):
        """Deletes the definition for a field from the definition table.

        :param fieldref: The name of the field
        :type fieldref: str
        """

        def_id = self.get_property(
            "definition", fieldref, grace=True, resolve_templates=True
        )

        Definition.objects.get(def_id=def_id).delete()

    @classmethod
    def _parse_complex_property(
        cls, complex_property_value, delimiter=",", sub_delimiter=":", int_flag=False
    ):
        """
        Class method to parse a 'complex' property value
        into a list or dictionary based on the presence of
        delimiter and sub_delimiter

        Examples:
           _parse_complex_property('A,B,C') -> ['A','B','C']
           _parse_complex_property('A:1,B:2,C:3') -> {'A':1,'B':2,'C':3}
        """
        # Declare result
        # -------------------
        if sub_delimiter is None:
            submodes = []
        else:
            submodes = {}

        if complex_property_value is None:
            pass
        elif not isinstance(complex_property_value, str):
            raise TypeError(
                f"Invalid property value:{complex_property_value}...string expected"
            )
        else:
            # Strip away leading/trailing spaces from each sub-string
            for i, submode_str in enumerate(complex_property_value.split(delimiter)):
                if sub_delimiter is None:
                    submodes.append(submode_str)
                else:
                    colon_locn = submode_str.find(
                        sub_delimiter
                    )  # <-- Split at first colon only (may be more after)
                    if colon_locn < 0:
                        submodes[submode_str.strip()] = None
                    elif len(submode_str) <= (colon_locn + 1):
                        submodes[submode_str[0:colon_locn].strip()] = None
                    elif int_flag:
                        submodes[submode_str[0:colon_locn].strip()] = int(
                            submode_str[(colon_locn + 1) :].strip()
                        )
                    else:
                        submodes[submode_str[0:colon_locn].strip()] = submode_str[
                            (colon_locn + 1) :
                        ].strip()

        return submodes

    def get_subparsed_property(
        self,
        property_name,
        fieldref=None,
        grace=False,
        default_property_value=None,
        resolve_templates=False,
        template_brackets=("<", ">"),
        alt_keyspace=None,
        native_property_mode="native_property_mode",
        appeal=False,
        delimiter=",",
        sub_delimiter=":",
        int_flag=False,
        **kwds_args,
    ):
        """
        Procures the specified property value and then parses it
        according to delimiter and (optional) sub_delimiter.

        If int_flag=True: Will convert each resultant parsed property to an integer

        If sub_delimiter is None:  Returns a list of the sub_properties

        Example: property_value:'IBM,AAPL' -> ['IBM','AAPL']

        If sub_delimiter is not None:  Returns an dictionary as follows:

           dictionary keys: sub-mode names
           dictionary values: submode assignments

        Example: property_value:'IBM:Intl Business Machines,AAPL:Apple Corp' -> {'IBM':Intl Business Machines, 'AAPL':Apple Corp'}

        """
        complex_property_value = self.get_property(
            property_name=property_name,
            fieldref=fieldref,
            grace=grace,
            default_property_value=default_property_value,
            resolve_templates=resolve_templates,
            template_brackets=template_brackets,
            alt_keyspace=alt_keyspace,
            native_property_mode=native_property_mode,
            appeal=appeal,
            **kwds_args,
        )

        return self._parse_complex_property(
            complex_property_value,
            delimiter=delimiter,
            sub_delimiter=sub_delimiter,
            int_flag=int_flag,
        )

    def is_field_specific_property(self, *args, **kwargs):
        return self.properties.is_fldspec_property(*args, **kwargs)

    # ============================ Native Property Methods ============================

    def is_native_property(self, property_name):
        """
        Designed to be called from set_property(property_name,...) & get_property(property_name,...)
        to evaluate whether the specified property_name has an associated build method
        NOTE: To function from set/get_property() methods, native_property footprint should be: self.<native_property>(self, field,...)
        """
        if property_name is None:
            return None
        return self.properties.is_native_property(property_name)

    def _instance_field_category(self, field_instance):
        if field_instance is None:
            return None
        elif isinstance(field_instance, DataLib):
            return "lib"
        else:
            return "data"

    def address_as_file_path(self, field=None):
        """Returns this library's address represented as a file system path.

        This feature is useful for libraries whose address structure mirrors the
        structure of the file system. For example, using the path property::
        TODO: REMOVED QUOTIENT_APP_DATA_PATH UPDATE DOCS
          <$QUOTIENT_APP_DATA_PATH>\\<address_as_file_path>

        will ensure that this library's field files will live in a directory
        that matches the address hierarchy.

        :param field: Unused argument. Allows this function to be used as a native
                      property for field specific property calls.

        :returns: The library's address represented as a file system path.
        :rtype: str
        """
        from qubles.io.base.rootlib import RootLib

        if isinstance(self, RootLib):
            raise ValueError(
                "Cannot convert RootLib address to file path. RootLib does "
                "not have an address."
            )
        if not self.address:
            raise ValueError(
                "Cannot convert address to file path. Library does not have "
                "an address."
            )
        if not all(isinstance(x, str) for x in self.address):
            raise ValueError(
                "Address-to-file-path is not supported on non-string "
                "addresses: {}".format(self.address)
            )

        return os.path.join(*self.address)

    # =============================== Freshness Methods ===============================

    def _valid_timestamp(self, timestamp, grace=True):
        if timestamp is None:
            return False
        elif not isinstance(timestamp, datetime):
            if grace:
                return False
            else:
                raise TypeError(f"Invalid timestamp recieved: {timestamp}")
        else:
            return timestamp != missing_date

    def _invalid_timestamp(self, timestamp, grace=True):
        return not self._valid_timestamp(timestamp, grace=grace)

    def _later_than(self, timestamp, other, grace=True):
        if self._invalid_timestamp(other, grace=grace):
            return True

        if self._invalid_timestamp(timestamp, grace=grace):
            return False

        return timestamp > other

    def _file_accessible(self, field):
        timestamp = None
        if field is not None and self.read_access(field):
            timestamp = self.file_timestamp(field)
            if timestamp:
                return timestamp

        return False

    def _prop_hash_in_filename(self, field):
        filename = self.get_property(
            "filename", field, grace=True, suppress_recording=True
        )

        return filename and "<property_dep_hash_current>" in filename

    def _annotate_property_dependency(
        self,
        property_name,
        is_global=False,
        appeal=False,
        resolve_templates=False,
        is_field_dependency=False,
        field_arg=None,
    ):
        """Inject property dependency qualifiers.

        Properties can be requested with a number of different settings, such as
        appeal, resolve_templates, document_as_field_dep, etc. When recording the
        dependency, we capture this state using special character qualifiers attached
        to the property name.

        """
        annotated_prop_name = property_name

        if is_global:
            annotated_prop_name += "*"
        if appeal:
            annotated_prop_name += "^"
        if resolve_templates:
            annotated_prop_name += "~"
        if is_field_dependency:
            annotated_prop_name += "`"
        if field_arg is not None:
            annotated_prop_name += f"@{field_arg}"

        return annotated_prop_name

    def parse_property_dependencies(self, field, **kwargs):
        """Return a consumable version of the 'property_dependencies' property.

        Strips out the special characters from the property names and converts them into a dictionary of metadata for each property dependency.

        NOTE: kwargs are passed through to the ``get_property`` method for the retrieval of the property dependencies.

        :type field: str
        :param field: Field name

        :rtype: list of dict
        :returns: Metadata for each property dependency:
            'property_name': The name of the property (without qualifiers)
            'appeal': True if the property was accessed with appeal=True

            'resolve_templates': True if the property was accessed with resolve_templates=True
            'is_field_dependency': True if the property was accessed with document_as_field_dep=True
            'domain': The library from which the property was accessed (currently either self or RootLib
            'field_arg': The name of the field, if it was a field-specific property call
        """
        from qubles.io.base.rootlib import RootLib

        prop_deps = self.get_property("property_dependencies", field, **kwargs)

        if not prop_deps:
            return prop_deps

        if not isinstance(prop_deps, (list, tuple)):
            raise ValueError(
                "property_dependencies({}) should be None/list/tuple... "
                "encountered: {}".format(field, prop_deps)
            )

        parsed_deps = []
        for prop_name in prop_deps:
            if prop_name is None or not isinstance(prop_name, str):
                raise ValueError(f"Invalid property_name:{prop_name}")

            if len(prop_name) == 0:
                _logger.warning("Encountered empty property_name (pre-cleansing)")
                continue

            cleaned_prop_name = prop_name
            appeal = False
            resolve_templates = False
            is_global = False
            is_field_dependency = False
            domain = self
            field_arg = None

            if "^" in prop_name:
                appeal = True
                cleaned_prop_name = cleaned_prop_name.replace("^", "")
            if "~" in prop_name:
                resolve_templates = True
                cleaned_prop_name = cleaned_prop_name.replace("~", "")
            if "*" in prop_name:
                is_global = True
                domain = RootLib()
                cleaned_prop_name = cleaned_prop_name.replace("*", "")
            if "`" in prop_name:
                is_field_dependency = True
                cleaned_prop_name = cleaned_prop_name.replace("`", "")
            if "@" in prop_name:
                parts = cleaned_prop_name.split("@", 1)
                if len(parts) > 1 and len(parts[1]) > 0:
                    field_arg = parts[1]
                else:
                    field_arg = field

                if is_global and self != RootLib():
                    field_arg = None

                cleaned_prop_name = parts[0]

            if len(prop_name) == 0:
                _logger.warning("Encountered empty property_name (post-cleansing)")
                continue

            parsed_deps.append(
                {
                    "property_name": cleaned_prop_name,
                    "appeal": appeal,
                    "resolve_templates": resolve_templates,
                    "is_field_dependency": is_field_dependency,
                    "domain": domain,
                    "field_arg": field_arg,
                }
            )

        return parsed_deps

    def field_dependencies(self, field):
        """Native property method for 'field_dependencies' property.

        Return the existing value for field_dependencies, if applicable, otherwise build the field to generate the dependencies.
        """
        if field is None:
            raise ValueError("Must specify field when asking for field_dependencies")

        if not self.is_derived_field(field):
            return None

        curr_deps = self.get_property(
            "field_dependencies",
            field,
            grace=True,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        if curr_deps is not None and isinstance(curr_deps, (list, tuple)):
            return curr_deps

        # If the current field dependencies are missing then the field will be considered stale.
        # Asking for the field will trigger a build which will generate the dependencies.
        self.get(field)

        new_deps = self.get_property(
            "field_dependencies",
            field,
            grace=True,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        return new_deps

    def property_dependencies(self, field):
        """Native property method for 'property_dependencies' property.

        Return the existing value for property_dependencies, if applicable, otherwise build the field to generate the dependencies.
        """
        if field is None:
            raise ValueError("Must specify field when asking for property_dependencies")

        if not self.is_derived_field(field):
            return None

        curr_deps = self.get_property(
            "property_dependencies",
            field,
            grace=True,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        if curr_deps is not None and isinstance(curr_deps, (list, tuple)):
            return curr_deps

        # If the current property dependencies are missing then the field will be
        # considered stale. Asking for the field will trigger a build which will generate the dependencies.
        self.get(field)

        new_deps = self.get_property(
            "property_dependencies",
            field,
            grace=True,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        return new_deps

    def local_property_dep_dict_current(self, field):
        """Generate a local dict from the current stored property dependencies.

        Generate a dict of the values for the current property dependencies of
        this library. Returns None if the property dependencies have not yet been generated.
            NOTE: Will not trigger a build of the given field
            NOTE: Property recording is suppressed

        :type field: str
        :param field: Generate a dict for this field

        :rtype: str
        :returns: The property dependency dict

        """
        prop_deps = self.parse_property_dependencies(
            field,
            grace=True,
            appeal=False,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        if prop_deps is None:
            _logger.debug(
                f"Unable to procure dict for {self}:{field}, prop_deps is None"
            )
            return None

        property_values_dict = {}
        for dep in prop_deps:
            skip_property = False
            # If resolve templates, first evaluate property value w/o template in order to isolate '<property_dep_hash_current>'
            if dep["resolve_templates"]:
                templated_value = dep["domain"].get_property(
                    dep["property_name"],
                    dep["field_arg"],
                    grace=True,
                    appeal=dep["appeal"],
                    resolve_templates=False,
                    suppress_recording=True,
                )

                # If a property depends on the property dependency hash, skip it to avoid a circular dependency
                skip_property = (
                    isinstance(templated_value, str)
                    and "<property_dep_hash_current>" in templated_value
                )

            if not skip_property:
                property_value = dep["domain"].get_property(
                    dep["property_name"],
                    dep["field_arg"],
                    grace=True,
                    appeal=dep["appeal"],
                    resolve_templates=dep["resolve_templates"],
                    suppress_recording=True,
                )

                property_values_dict[dep["property_name"]] = property_value

        # Make sure that dates are always in datetime format always so that we don't have prop dep hash comparison failures
        start_date = property_values_dict.get("start_date")
        end_date = property_values_dict.get("end_date")

        # If start_date and end_date are not None and they are not already datetime, attempt to cast them.
        if (start_date and end_date) and not (
            isinstance(start_date, datetime) and isinstance(end_date, datetime)
        ):
            try:
                property_values_dict["start_date"] = datetime.strptime(
                    start_date, "%Y-%m-%d %H:%M:%S"
                )
                property_values_dict["end_date"] = datetime.strptime(
                    end_date, "%Y-%m-%d %H:%M:%S"
                )
            except Exception as e:
                # Leaving this generic catch in the code so we can narrow down the inconsistencies in start_date
                # and end_date and where they come from.
                # Sometimes start date like 2 years/ 2 months, so typecast is not needed
                _logger.exception(e)

        return property_values_dict

    def property_dep_hash_current(self, field, validate=True):
        """Generate a recursive hash from the current stored property dependencies.

        Generate a hash of the values for the current property dependencies of this library combined with the hashes of all its field dependencies.
        Returns None if any library has not yet generated its property dependencies.
            NOTE: Will not trigger a build of the given field
            NOTE: Property recording is suppressed

        :type field: str
        :param field: Generate a hash for this field

        :rtype: str
        :returns: The property dependency hash

        """
        if self.get_property("field_category", field, grace=True) == "lib":
            return None
        return self.get(
            field,
            request_type="field_property_dep_hash",
            key_grace=True,
            suppress_recording=True,
            validate=validate,
        )

    def field_dependencies_current(self, field):
        """Return all field dependencies (explicit and implicit) for the given field.

        Field dependencies may be procured from the 'field_dependencies' property, or
        from items in the 'property_dependencies' property that indicate they are serving as field dependencies.

        If a field dependency resolves to a Screen, the dependency is considered to be on the last rule of the screen.
            NOTE: Returns None if the field dependencies have not yet been calculated.
            NOTE: Property recording will be suppressed.

        :type field: str
        :param field: Field name

        :rtype: list of tuple
        :returns: List of 2-tuple: (dependency domain, dependency field)

        """
        return self.get(
            field,
            request_type="field_dependencies",
            key_grace=True,
            suppress_recording=True,
        )

    def src_csv_time_stamp(
        self,
        field,
        path_property_name="path",
        filename_property_name="filename",
        file_meta_property_name="file_meta",
        suppress_recording=None,
    ):
        # Implemented in DataMapper child class
        pass

    def src_sql_table_time_stamp(
        self,
        field,
        suppress_recording=None,
    ):
        # Implemented in DataMapper child class
        pass

    def sql_query_timeout_check(
        self, field, file_meta_property_name="file_meta", suppress_recording=None
    ):
        # Implemented in DataMapper child class
        pass

    def src_time_stamp(self, field, type):
        # Implemented in DataMapper child class
        pass

    def field_fresher_than(
        self,
        field,
        timestamps,
        check_property_deps=True,
        check_field_deps=True,
        property_dep_hash=None,
        ad_hoc="ad_hoc",
    ):
        """Return whether the given field is fresher than the provided timestamps.

        Given some timestamps (presumably those of the file backing the field and the time_stamp of the field in residence), checks if the field or
        any of its dependencies have been updated since the timestamps. Returns a list of booleans corresponding to the list of timestamps, True if
        dependencies have been updated after the timestamp, since that implies the field will need to be recomputed. Checks the following:

            1. timestamp < field_fresh_time_override
            2. timestamp < field_defn_time
            3. property_dep_hash_last != property_dep_hash_current
            4. residence timestamp < file timestamp
            5. any dependency has been updated

        :type field: str
        :param field: Name of field

        :type timestamps: list
        :param timestamps: List of timestamps to check

        :rtype: list
        :returns: List of booleans, one for each provided timestamp, representing whether the timestamp is out of date.

        """
        from qubles.io.base.designlib import DesignLib

        timestamps_str = ", ".join(map(str, timestamps))  # Used for logging

        # Do this at the beginning to consistently record property dependencies
        is_derived_field = self.is_derived_field(field)

        # Begin by assuming that updates are needed for invalid timestamps
        # [invalid timestamps assumed to indicate that content is unavailable]
        # ---------------------------------------------------------------------
        invalid_time_stamps = [self._invalid_timestamp(ts) for ts in timestamps]
        updates_needed = invalid_time_stamps

        if all(updates_needed):
            _logger.debug(f"Stale: {field}: invalid timestamps: {timestamps}")
            return updates_needed

        # Check for changes in property dependencies
        # -------------------------------------------
        if not isinstance(check_property_deps, list):
            check_property_deps = [check_property_deps] * len(timestamps)

        if any(check_property_deps) and property_dep_hash is None:
            # Here, property_dep_hash (current) was not provided as an argument
            property_dep_hash = self.property_dep_hash_current(field)

        if any(check_property_deps):
            # NOTE: Logic below assumes that (property_dep_hash is None)
            # represents that property depenency hash current has not been generated
            # [as opposed to property depenency hash current is None because theer are no property dependencies]
            if is_derived_field and property_dep_hash is None:
                updates_needed = [
                    u or d for u, d in zip(updates_needed, check_property_deps)
                ]
                if all(updates_needed):
                    _logger.debug(f"Stale: {field}: property_dep_hash is None")
                    return updates_needed

        # All following checks are only relevant to derived fields
        # ---------------------------------------------------------
        if not is_derived_field:
            return updates_needed

        # Check if field_defn_time has been updated
        # ------------------------------------------
        field_defn_time = self.get_property(
            "field_defn_time",
            field,
            grace=True,
            suppress_recording=True,
        )

        if self._valid_timestamp(field_defn_time):
            after = [self._later_than(field_defn_time, ts) for ts in timestamps]
            updates_needed = [u or d for u, d in zip(updates_needed, after)]

            if all(updates_needed):
                _logger.debug(
                    "Stale: %s: field_defn_time (%s) > %s",
                    field,
                    field_defn_time,
                    timestamps,
                )
                return updates_needed

        # Check source "file" (may be table, etc.) via file_meta
        # -------------------------------------------------------
        # Is the file_meta_timestamp > each timestamp?
        # If so, the source file has been updated AFTER the cache file
        # And therefore, the cache_file is out of date and updates are needed
        if isinstance(self, DesignLib):
            file_meta = self.get_property(
                "file_meta",
                field,
                grace=True,
            )
            if file_meta is not None:
                file_meta = loads(file_meta)
                if "type" in file_meta:
                    file_meta_timestamp = self.src_time_stamp(field, file_meta["type"])
                    if file_meta_timestamp and isinstance(self, DesignLib):
                        after = [file_meta_timestamp > ts for ts in timestamps]
                        updates_needed = [u or d for u, d in zip(updates_needed, after)]
                        if all(updates_needed):
                            _logger.debug(
                                "Stale: %s: file_meta_timestamp (%s) > %s",
                                field,
                                file_meta_timestamp,
                                timestamps_str,
                            )
                            return updates_needed

        if check_field_deps:
            # Check if any dependencies have been updated
            # --------------------------------------------
            field_deps = self.field_dependencies_current(field)
            if field_deps is None:
                _logger.debug("Stale: %s: missing field dependencies", field)
                return [True] * len(timestamps)

            for args in field_deps:
                updates_needed = self._check_field_dep_updates(
                    args, updates_needed, timestamps, field, timestamps_str
                )
                # If all elements in updates_needed are True, then break out of the loop
                if all(updates_needed):
                    break

        return updates_needed

    def _check_field_dep_updates(
        self, args, updates_needed, timestamps, field, timestamps_str
    ):
        dep_lib, dep_field = args

        try:
            dep_time_stamp = dep_lib.get(
                dep_field,
                request_type="field_time_stamp",
                key_grace=True,
                suppress_recording=True,
            )
        except UnregisteredFieldError:
            _logger.warning(
                f"UNREGISTERED FIELD IN CHECK FIELD DEP UPDATES:: field -> {field}, dep_field -> {dep_field}"
            )
            dep_time_stamp = None

        after = [
            dep_time_stamp is None or self._later_than(dep_time_stamp, ts)
            for ts in timestamps
        ]
        updates_needed_result = [u or d for u, d in zip(updates_needed, after)]

        if all(updates_needed_result):
            if dep_time_stamp is None:
                _logger.debug(
                    "Stale: %s: dependency outdated: %s, time_stamp is None",
                    field,
                    dep_field,
                )
            else:
                _logger.debug(
                    "Stale: %s: dependency %s time_stamp (%s) > %s",
                    field,
                    dep_field,
                    dep_time_stamp,
                    timestamps_str,
                )

        return list(updates_needed_result)

    def field_time_stamp(self, field, **kwargs):
        return self.get(field, request_type="field_time_stamp", **kwargs)

    # ====================================== Get ======================================

    def __getitem__(self, fieldref):
        return self.get(fieldref)

    def get(
        self,
        fieldref,
        request_type="field_value",
        validate=True,
        auto_register="auto_register",
        auto_squeeze="auto_squeeze",
        suppress_recording=False,
        key_grace="key_grace",
        minimize_address=True,
        translate_address=True,
        force_build=False,
        log=None,
        property_dep_hash_override=None,
        write_to_tree=True,
        **kwargs,
    ):
        """Get an object from a Data Library.

        :param fieldref: A reference to the object to be retrieved.
        :type fieldref: Any value that is supported by the LibAddress constructor

        :type request_type: str
        :param request_type: Determines the return value. One of:
            'field_value': The object (DataLib, Quble, etc.)

            'field_time_stamp': The timestamp of the freshest version of the object,
                                whether it be from residence, file, or None if the object needs to be built
            'field_instance_sample': return an "empty" version of the object without actually building anything
            'field_dependencies': The current field dependencies (including field dependencies from properties)
            'field_property_dep_hash': The current property dependecy hash

        :type validate: bool
        :param validate: Set to False to request the field without a freshness check

        :type auto_register: bool
        :param auto_register: Register the requested field if it doesn't exist

        :type auto_squeeze: bool
        :param auto_squeeze: Auto squeeze behavior for hyper index, if requested

        :type suppress_recording: bool
        :param suppress_recording: Set to True to prevent the request from writing field and property dependencies

        :type key_grace: bool
        :param key_grace: If given a field reference that does not exist, return None

        :type minimize_address: bool
        :param minimize_address: If given a global address, minimize before resolving.

        :type translate_address: bool
        :param translate_address: If given a linked address, follow the link.

        :type force_build: bool
        :param force_build: Force the field to build, even if it is not out-of-date.

        :type log: bool
        :param log: Set to False to disable logging during this request, or True to
                    force logging.

        :param write_to_tree: Indicates if we should send a serialized tree over the WebSocket messenger.

        :rtype: See documentation for the <request_type> parameter.
        :returns: The object at the given field reference, or some piece of data related to it if <request_type> is specified.

        """
        from qubles.io.base.rootlib import FrameContextManager
        from qubles.io.base.rootlib import RootLib

        recorder = DataLib.dependency_recorder

        # Resolve lib and field
        #
        # E.g: If passed 'Pete|Backtests|Sample BT|ASSET_RETURNS', we want to
        # resolve lib to 'Pete|Backtests|Sample BT' and field to 'ASSET_RETURNS'
        # -----------------------------------------------------------------------
        address = LibAddress(fieldref, base_domain=self)

        if minimize_address:
            min_address = address.minimize() if len(address) != 1 else address
            lib = min_address.base_domain
            field = None if len(min_address) == 0 else min_address[-1]
            subkeys = None
        else:
            lib, field, subkeys = address.min_subparts

        field_id = None
        if lib is None:
            raise ValueError(f"Unable to minimize fieldref: {fieldref}")
        elif not is_rootlib(lib):
            field_id = self.field_to_id_map.get(field)

        # Check whether another thread has signalled for this thread to terminate
        #
        # This is a strategic location for interruption checking beacuse:
        #   1) Throwing an error here will not cause undesired state corruption
        #   2) The .get() method can take a long time, so interrupt before it runs
        # ------------------------------------------------------------------------
        if hasattr(current_thread(), "check_for_interruption_exception"):
            current_thread().check_for_interruption_exception()

        field_category = lib.get_property("field_category", field)

        # If the address minimized to a different library, forward the request
        # ---------------------------------------------------------------------
        if lib is not self:
            return lib.get(
                field,
                request_type=request_type,
                validate=validate,
                auto_register=auto_register,
                auto_squeeze=auto_squeeze,
                suppress_recording=suppress_recording,
                key_grace=key_grace,
                minimize_address=not minimize_address,
                translate_address=translate_address,
                force_build=force_build,
                log=log,
                property_dep_hash_override=property_dep_hash_override,
                write_to_tree=write_to_tree,
                **kwargs,
            )

        # Validate input parameters
        # --------------------------
        requests_types_with_context = [
            "field_value",
            "field_time_stamp",
            "field_property_dep_hash",
        ]
        requests_types_without_context = [
            "field_instance_sample",
            "field_dependencies",
        ]

        if (
            request_type
            not in requests_types_with_context + requests_types_without_context
        ):
            raise ValueError(f"Invalid request_type: {request_type}")

        # Initialize dependency recording for this request
        #
        # IMPORTANT: There should be no calls to get_property() before this point
        # ------------------------------------------------------------------------
        if log is None:
            logging_enabled = request_type == "field_value"
        else:
            logging_enabled = log

        recorder.increment_pointer(self, field, field_id, address, logging_enabled)

        value = None
        try:
            if self.is_derived_field(field) and request_type == "field_value":
                _logger.debug(
                    f"DataLib::get -> {self}:{field}",
                )

            # Initialize the build context and procure the requested value
            #
            # Introduce a new frame so the build context does not pollute the global
            # environment.
            #
            # IMPORTANT: There should be no calls to set_control() before this point
            # -----------------------------------------------------------------------
            procurement_duration_last = {}
            if isinstance(lib, DataLib) and field and lib.address:
                addr_string = "/".join(lib.address)
                addr_string = addr_string + "/" + str(field)
                procurement_duration_last[addr_string] = lib.get_property(
                    "procurement_duration_last", field, grace=True
                )

            with FrameContextManager():
                if (
                    request_type in requests_types_with_context
                    and field_category == "data"
                ):
                    with recorder.disable_property_access_recording():
                        with recorder.enable_root_property_submission_recording():
                            self.set_build_context(field)

                if recorder.has_cached_value(request_type):
                    value = recorder.get_cached_value(request_type)
                else:
                    value = self._get_field(
                        address=address,
                        field=field,
                        subkeys=subkeys,
                        request_type=request_type,
                        validate=validate,
                        auto_register=auto_register,
                        auto_squeeze=auto_squeeze,
                        suppress_recording=suppress_recording,
                        key_grace=key_grace,
                        minimize_address=minimize_address,
                        translate_address=translate_address,
                        force_build=force_build,
                        log=log,
                        recorder=recorder,
                        procurement_duration_last=procurement_duration_last,
                        property_dep_hash_override=property_dep_hash_override,
                        write_to_tree=write_to_tree,
                        **kwargs,
                    )
                    try:
                        # We attempt to log the procured field with all statistics like prop_dep_hash, build_tree_last, procurement_time, universe, etc
                        if field_id is not None:
                            field_val_fxs_last = self.get_property(
                                property_name="field_val_fxs_last",
                                fieldref=field,
                                grace=True,
                            )
                            field_col_types_last = self.get_property(
                                property_name="field_col_types_last",
                                fieldref=field,
                                grace=True,
                            )
                            field_valuespaces_last = self.get_property(
                                property_name="field_valuespaces_last",
                                fieldref=field,
                                grace=True,
                            )
                            build_tree_last = self.get_property(
                                property_name="build_tree_last",
                                fieldref=field,
                                grace=True,
                            )
                            procurement_duration_last = self.get_property(
                                property_name="procurement_duration_last",
                                fieldref=field,
                                grace=True,
                            )
                            property_dep_dict_last = self.get_property(
                                property_name="property_dep_dict_last",
                                fieldref=field,
                                grace=True,
                            )
                            property_dep_hash_last = self.get_property(
                                property_name="property_dep_hash_last",
                                fieldref=field,
                                grace=True,
                            )
                            properties_dict = {
                                "field_val_fxs_last": field_val_fxs_last,
                                "field_col_types_last": field_col_types_last,
                                "field_valuespaces_last": field_valuespaces_last,
                                "build_tree_last": build_tree_last,
                                "procurement_duration_last": procurement_duration_last,
                            }
                            properties_json = dumps(properties_dict)
                            universe = RootLib().get_control("univ", grace=True)
                            default_freq = RootLib().get_control("freq")
                            start_date = RootLib().get_control("start_date")
                            end_date = RootLib().get_control("end_date")
                            timestamp = str(datetime.now())
                            user_name = context_username.get()
                            file_type = self.get_property("file_type", field).upper()
                            tgt_table_name = None

                            if file_type == "SNOWFLAKE" and is_quble(value):
                                tgt_table_name = self.get_property(
                                    "filename",
                                    field,
                                    resolve_templates=True,
                                    grace=True,
                                )
                                tgt_table_name = (
                                    tgt_table_name.replace(".", "_")
                                    if "." in tgt_table_name
                                    else tgt_table_name
                                )
                                Activity.objects.create(
                                    quble_id=field_id,
                                    quble_address=address.maximize().to_string(),
                                    table_name=(
                                        tgt_table_name
                                        if tgt_table_name is not None
                                        else None
                                    ),
                                    properties_recording=properties_json,
                                    properties_dep_dict_last=property_dep_dict_last,
                                    properties_dep_hash_last=property_dep_hash_last,
                                    universe=universe,
                                    created_at=timestamp,
                                    user_name=user_name,
                                    default_freq=default_freq,
                                    start_date=start_date,
                                    end_date=end_date,
                                )
                    except Exception as e:
                        _logger.exception(
                            f"Error adding recent build statistics to activity table::{e}"
                        )

                    recorder.set_cached_value(request_type, value)
        finally:
            recorder.pop_pointer(
                increment_pointer=(request_type == "field_value" and is_quble(value)),
                suppress_recording=suppress_recording,
                write_to_tree=write_to_tree,
            )

        return value

    def _get_field(
        self,
        address,
        field,
        subkeys,
        request_type,
        validate,
        auto_register,
        auto_squeeze,
        suppress_recording,
        key_grace,
        minimize_address,
        translate_address,
        force_build,
        log,
        recorder,
        procurement_duration_last=None,
        property_dep_hash_override=None,
        write_to_tree=True,
        **kwargs,
    ):
        """Helper for ``DataLib.get()``.

        The ``DataLib.get()`` method does some initial parsing of input parameters and environment setup.
        The ``DataLib._get_field()`` method performs the actual retrieval of the field from the library.

        WARNING: This is a helper method for the ``get`` method and should not be called from any other context.

        See: :meth:`~qubles.io.base.datalib.DataLib.get`
        """
        from qubles.core.quble import Quble
        from qubles.io.base.rootlib import RootLib
        from qubles.io.util.multi_build_lib import MultiBuildLib

        if auto_squeeze == "auto_squeeze":
            auto_squeeze = RootLib().get_control("auto_squeeze")

        # Handle the unregistered field case
        # -----------------------------------
        if field not in self.field_index:
            try:
                self._handle_unregistered_field(
                    field,
                    auto_register,
                    key_grace,
                    request_type,
                )
            except GracefulUnregisteredFieldError:
                return None

        # Determine validity of cached objects
        # -------------------------------------
        field_category = self.get_property("field_category", field)
        if (
            validate
            and recorder.is_origin
            and request_type == "field_value"
            and field_category == "data"
            and not suppress_recording
        ):
            Messenger.write_message("Checking if data is up-to-date...")

        # Return the underlying object if the field is a link
        # ----------------------------------------------------
        if translate_address:
            underlying_address = address.translate(
                pre_minimized=minimize_address,
                minimize=minimize_address,
                filename_property_name="underlying_filename",
            )
            if minimize_address and len(underlying_address) == 1:
                underlying_lib = underlying_address.base_domain
                underlying_field = underlying_address[-1]
            else:
                underlying_lib, underlying_field, _ = underlying_address.min_subparts

            # Pass along the get request if the underlying object is different
            if underlying_lib is not self or underlying_field != field:
                return underlying_lib.get(
                    underlying_field,
                    request_type=request_type,
                    validate=validate,
                    auto_register=auto_register,
                    auto_squeeze=auto_squeeze,
                    suppress_recording=suppress_recording,
                    key_grace=key_grace,
                    minimize_address=not minimize_address,
                    translate_address=translate_address,
                    force_build=force_build,
                    log=log,
                    property_dep_hash_override=property_dep_hash_override,
                    **kwargs,
                )

        # Perform shortcuts for certain request types
        # --------------------------------------------
        if field not in self.field_index and request_type != "field_value":
            return None

        if request_type == "field_dependencies":
            try:
                return self._field_dependencies_current(field)
            except UnregisteredFieldError:
                return None

        if request_type == "field_property_dep_hash":
            try:
                return self._property_dep_hash_current(
                    field, recorder, validate=validate
                )
            except UnregisteredFieldError:
                return None

        builder = self.build_method(field, build_access_check=True)
        if builder is not None and request_type == "field_instance_sample":
            return Quble.undefined_instance()

        with recorder.disable_property_access_recording():
            (
                residence_outdated,
                file_outdated,
                residence_key,
                residence_timestamp,
                file_timestamp,
            ) = self._check_freshness(
                field,
                (
                    property_dep_hash_override
                    if (
                        property_dep_hash_override is not None
                        and request_type != "field_property_dep_hash"
                    )
                    else self.property_dep_hash_current(field, validate=validate)
                ),
                force_build,
                validate,
                recorder._build_tree,
                suppress_recording,
            )

        # Procure the value
        # ------------------
        sourced_from_virtual = False
        sourced_from_file = False
        sourced_from_build = False

        # -----------------------------------------------------------------------------
        # Step 1: Try to access virtual in-memory storage (if applicable)
        # -----------------------------------------------------------------------------
        # NOTE: For request_type == 'field_instance_sample', it does not matter if the residence is outdated.
        # -----------------------------------------------------------------------------
        if not residence_outdated or (
            request_type == "field_instance_sample"
            and self._exists_in_residence(residence_key)
        ):
            if request_type == "field_time_stamp":
                return residence_timestamp

            value = self.residence[residence_key]

            if request_type == "field_instance_sample":
                return value

            if recorder.is_origin and builder and not suppress_recording:
                Messenger.write_message("Sourcing from memory...")
            sourced_from_virtual = True
            if address and isinstance(value, Quble) and field != "Exceptions":
                recorder.add_current_pointer_to_build_tree(
                    cache_flag="Memory",
                    procurement_duration_last=procurement_duration_last,
                    node_table_name=value.table_name,
                    write_to_tree=write_to_tree,
                )

        # -----------------------------------------------------------------------------
        # Step 2: Try to read from file (if applicable)
        # -----------------------------------------------------------------------------
        # NOTE: For request_type == 'field_instance_sample', it does not matter if the file is outdated.
        # -----------------------------------------------------------------------------
        elif not file_outdated or (request_type == "field_instance_sample"):
            if request_type == "field_time_stamp":
                return file_timestamp
            elif request_type == "field_instance_sample":
                return self._get_instance_sample(field)

            if recorder.is_origin and not suppress_recording:
                Messenger.write_message("Collecting data from source...")

            sourced_from_file = True
            value = self.read_field(field, time_stamp=file_timestamp)

            if address and isinstance(value, Quble) and field != "Exceptions":
                recorder.add_current_pointer_to_build_tree(
                    cache_flag="Source",
                    procurement_duration_last=procurement_duration_last,
                    node_table_name=value.table_name,
                    write_to_tree=write_to_tree,
                )

        # -----------------------------------------------------------------------------
        # Step 3: Throw an appropriate error if item is not derived
        # -----------------------------------------------------------------------------
        elif builder is None:
            if force_build:
                raise BuilderError(
                    f"force build=True (erroneously) for non-derived field: {field}"
                )
            elif self.read_grace(field) and file_outdated:
                if request_type != "field_value":
                    return None

                sourced_from_file = True
                value = None
            elif not self.read_access(field):
                raise BuilderError(
                    f"Non-derived (or no build access) field: {field}: read access denied...virtual storage and/or access_grace(read) should apply"
                )
            elif field_category == "lib":
                raise BuilderError(
                    f"Unable to procure library field: {field}. Could not be loaded from cached or persisted state."
                )
            else:
                raise BuilderError(
                    f"Non-derived (or no build access) field: {field}: (stale or absent) and (access_grace(read)=False)"
                )
        # -----------------------------------------------------------------------------
        # Step 4: Make sure build is authorized
        # -----------------------------------------------------------------------------
        elif not self.build_access(field):
            raise BuilderError(
                f"Derived field: {field}: stale or absent or dependencies unvailable or read prohibited & build not authorized"
            )
        # -----------------------------------------------------------------------------
        # Step 5: Call the requisite build method (if no aggregation is applicable)
        # -----------------------------------------------------------------------------
        else:
            if request_type == "field_time_stamp":
                return None

            if request_type == "field_instance_sample":
                return Quble.undefined_instance()

            sourced_from_build = True
            _logger.debug(f"Starting build process for: {self}:{field}")
            if recorder.is_origin and not suppress_recording:
                Messenger.write_message("Initiating build process...")

            value = self._call_builder(
                field,
                builder,
                recorder,
                **kwargs,
            )
            # Hard assign current timestamp, avoids inheriting old timestamp from previous build
            value.time_stamp = datetime.now()

        if value:
            value.address = address.maximize()
        # -----------------------------------------------------------------------------
        # Validate state before continuing...
        # -----------------------------------------------------------------------------
        if request_type != "field_value":
            raise BuilderError(
                f"request_type: {request_type}...Should not have reached this point!"
            )

        # -----------------------------------------------------------------------------
        # Clear DependencyRecorder cache, if necessary
        # -----------------------------------------------------------------------------
        if (
            sourced_from_build
            or (builder is None and sourced_from_file)
            and field_category == "data"
        ):
            recorder.clear_cached_values()

        # -----------------------------------------------------------------------------
        # Procure newly recorded field & property dependencies from DependencyRecorder
        # -----------------------------------------------------------------------------
        new_field_deps = None
        new_prop_deps = None
        new_duration = None
        record_fxs = False
        record_valuespaces = False
        record_col_types = False
        record_fx = False
        record_type = False
        record_duration = False
        record_prop_deps = False
        record_field_deps = False

        if sourced_from_build:
            record_field_deps = (
                "field_dependencies" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
                and not self.virtual_access_only(field)
            )
            record_prop_deps = (
                "property_dependencies" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
                and not self.virtual_access_only(field)
            )
            record_duration = (
                "procurement_duration_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
            )
            record_type = "value_type_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
            record_fx = "value_fx_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
            record_col_types = (
                "field_col_types_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
            )
            record_valuespaces = (
                "field_valuespaces_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST
            )
            record_fxs = "field_val_fxs_last" in self.NATIVE_PROPERTY_AUTOSTORE_LIST

            if record_field_deps:
                new_field_deps = []

                for dep_address in recorder.get_field_recording():
                    if dep_address.base_domain != self:
                        dep_address = dep_address.maximize()

                    new_field_deps.append(dep_address)
                    # Set up dependent_fields
                    dep_fields = dep_address.min_domain.get_property(
                        "dependent_fields",
                        dep_address[-1],
                        grace=True,
                        default_property_value=[],
                        suppress_recording=True,
                    )
                    dep_fields = [] if dep_fields is None else dep_fields
                    max_address = list(address.maximize())
                    if max_address not in dep_fields:
                        if max_address[0] != "__TEMP__":
                            dep_fields.append(max_address)
                            dep_address.min_domain.set_property(
                                "dependent_fields", dep_fields, dep_address[-1]
                            )
                            dep_address.min_domain.commit()

            if record_prop_deps:
                new_prop_deps = recorder.get_property_access_recording()

            if record_duration:
                new_duration = recorder.get_cumulative_duration()

        # -----------------------------------------------------------------------------
        # Build Multi-Value Dictionary
        # -----------------------------------------------------------------------------
        # TYPICAL CASE:The above value access logic yields a single value
        # (Quble, DataLib or None) representing the field being requested.
        #
        # SPECIAL CASE: When a builder function returns DataLib where one of the
        # returned library's fields matches the original field being requested, assume
        # the builder is returning multiple, inter-related derived fields. In this
        # case, we want to 1) validate, 2) write and 3) reside each of the field/value
        # pairs within the new library from the builder.
        # -----------------------------------------------------------------------------
        multi_values = {}

        is_multi_build = (
            value is not None
            and sourced_from_build
            and (
                isinstance(value, MultiBuildLib)
                or (isinstance(value, DataLib) and field in value.field_index)
            )
        )

        if is_multi_build:
            if isinstance(value, MultiBuildLib) and field not in value.field_index:
                raise MultiBuildError(
                    f"field (sourced_from_build): {field} absent from MultiBuildLib result: {value.field_index}"
                )
            value.parent_id = self.id
            for f in value.field_index:
                multi_values[f] = value.get(
                    f,
                    suppress_recording=True,
                    write_to_tree=False,
                )

            # If we received a MultiBuildLib from a builder method, make sure to return just the field that was requested
            return_value = multi_values[field]
        else:
            multi_values[field] = value
            return_value = value

        if (
            return_value is not None
            and self.get_property("file_type", field) == "SNOWFLAKE"
            and not self.virtual_storage_only(field)
        ):
            return_value.id = self.field_to_id_map[return_value.name]

        # -----------------------------------------------------------------------------
        # Persist post-processing information for each multi-value
        # -----------------------------------------------------------------------------
        # 2a. write "field_col_types_last" property
        # 2b. write "field_valuespaces_last" property
        # 2c. write "field_val_fxs_last" property
        # 2d. write "value_type_last" property
        # 2e. write "value_fx_last" property
        # 3. write "field_dependencies" property
        # 4. write "property_dependencies" property
        # 5. write "procurement_duration_last" property
        # 6. write to residence
        # 7. write to file
        # NOTE: No messenger writes should occur within the multi-build loop
        # This prevents flickering of the build tree.
        # -----------------------------------------------------------------------------

        lib_commit_flag = False
        multi_build_is_written = False
        for multi_field, multi_value in multi_values.items():
            if multi_field not in self.field_index:
                continue

            # --------------------------
            # Update field_col_types
            # For univariate Quble: saved as str
            # For multivariate Quble: saved as dict keyed by each valuespace
            # For undefined and non-variate Quble: not saved (and removed from properties)
            # --------------------------
            if record_col_types and isinstance(multi_value, Quble):
                if multi_value.is_undefined:
                    col_types = None
                else:
                    col_types = multi_value.get_column_type(multi_value.spaces)

                lib_commit_flag = True
                # Deciding to save an empty dict here to document that information has been requested/saved
                if col_types in [None]:
                    self.remove_property(
                        "field_col_types_last", multi_field, grace=True
                    )
                elif isinstance(col_types, str):
                    self.set_property("field_col_types_last", col_types, multi_field)
                else:
                    self.set_property(
                        "field_col_types_last", dumps(col_types), multi_field
                    )

            # --------------------------
            # Update field_valuespaces
            # For nonvariate, univariate or multivariate Quble: saved as list
            # For undefined Quble: not saved (and removed from properties)
            # --------------------------
            if record_valuespaces and isinstance(multi_value, Quble):
                valuespaces = (
                    multi_value.valuespaces
                )  # <-- May yield None for undefined Quble
                lib_commit_flag = True
                # Decided to formally set_property for case: valuespaces=[] or valuespaces=()
                if valuespaces in [None]:
                    self.remove_property(
                        "field_valuespaces_last", multi_field, grace=True
                    )
                else:
                    self.set_property(
                        "field_valuespaces_last",
                        dumps(valuespaces) if valuespaces is not None else valuespaces,
                        multi_field,
                    )

            # --------------------------
            # Update field fxs
            # For variate Quble: saved as dict keyed by each valuespace
            # For undefined and non-variate Quble: not saved (and removed from properties)
            # --------------------------
            if record_fxs and isinstance(multi_value, Quble):
                if multi_value.is_undefined or multi_value.is_nonvariate:
                    value_fxs = None
                else:
                    # Here, multi_value is a variate Quble, so we expect value_fxs to be a dict keyed on each valuespace
                    # NOTE: omit_unassigned=False below would only include valuespaces in dict that store non-None fx assignments
                    value_fxs = multi_value.get_space_info(
                        info_type="fx",
                        space=multi_value.valuespaces,
                        grace=True,
                    )
                    if value_fxs is None:
                        pass
                    elif isinstance(value_fxs, (dict, list, tuple)):
                        if len(value_fxs) > 0:
                            value_fxs = dumps(value_fxs)
                        else:
                            value_fxs = None
                    elif not isinstance(value_fxs, str):
                        value_fxs = str(value_fxs)

                lib_commit_flag = True
                # Deciding to save an empty dict here to document that information has been requested/saved
                if value_fxs in [None]:
                    self.remove_property("field_val_fxs_last", multi_field, grace=True)
                else:
                    self.set_property("field_val_fxs_last", value_fxs, multi_field)

            # --------------------------
            # Update field value type
            # For univariate Quble: saved as str
            # For multivariate Quble: saved as dict keyed by each valuespace
            # For undefined and non-variate Quble: not saved (and removed from properties)
            # --------------------------
            if record_type and isinstance(multi_value, Quble):
                if multi_value.is_undefined:
                    value_type = None
                elif multi_value.is_univariate:
                    value_type = multi_value.get_column_type(multi_value.valuespace)
                elif multi_value.is_multivariate:
                    value_type = multi_value.get_column_type(multi_value.valuespaces)
                    if value_type is None:
                        pass
                    elif isinstance(value_type, dict):
                        value_type = dumps(value_type)
                    elif not isinstance(value_type, str):
                        value_type = str(value_type)
                else:
                    value_type = None

                lib_commit_flag = True
                if value_type in [None, {}]:
                    self.remove_property("value_type_last", multi_field, grace=True)
                else:
                    self.set_property("value_type_last", value_type, multi_field)

            # --------------------------
            # Update field fx
            # For univariate Quble: saved as str
            # For multivariate Quble: saved as dict keyed by each valuespace
            # For undefined and non-variate Quble: not saved (and removed from properties)
            # --------------------------
            if record_fx and isinstance(multi_value, Quble):
                if multi_value.is_undefined:
                    value_fx = None
                elif multi_value.is_univariate:
                    value_fx = multi_value.get_space_info(
                        info_type="fx",
                        space=multi_value.valuespace,
                        grace=True,
                    )
                elif multi_value.is_multivariate:
                    value_fx = multi_value.get_space_info(
                        info_type="fx",
                        space=multi_value.valuespaces,
                        grace=True,
                    )
                    if value_fx is None:
                        pass
                    elif isinstance(value_fx, dict):
                        value_fx = dumps(value_fx)
                    elif not isinstance(value_fx, str):
                        value_fx = str(value_fx)
                else:
                    value_fx = None

                lib_commit_flag = True
                if value_fx in [None, {}]:
                    self.remove_property("value_fx_last", multi_field, grace=True)
                else:
                    self.set_property("value_fx_last", value_fx, multi_field)

            # --------------------------
            # Update field dependencies
            # --------------------------
            if record_field_deps:
                lib_commit_flag = True
                if new_field_deps is None:
                    self.remove_property("field_dependencies", multi_field)
                else:
                    self.set_property("field_dependencies", new_field_deps, multi_field)

            # --------------------------------------------------------
            # Update property dependencies & property dependency hash
            # --------------------------------------------------------
            if record_prop_deps:
                lib_commit_flag = True
                self.set_property("property_dependencies", new_prop_deps, multi_field)

                prop_hash_current = self._property_dep_hash_current(
                    multi_field, recorder, validate=False
                )
                if prop_hash_current is None:
                    self.remove_property("property_dep_dict_last", multi_field)
                    self.remove_property("property_dep_hash_last", multi_field)
                else:
                    self.set_property(
                        "property_dep_dict_last",
                        self.local_property_dep_dict_current(multi_field),
                        multi_field,
                    )
                    self.set_property(
                        "property_dep_hash_last", prop_hash_current, multi_field
                    )

            # --------------------------------------------------------------------
            # Write to residence
            #
            # IMPORTANT: Must happen AFTER persisting the "property_dependencies"
            # property so the residence key is generated correctly.
            # --------------------------------------------------------------------
            residence_key = None
            if not sourced_from_virtual and self.virtual_storage(multi_field):
                result_is_library = (
                    self.get_property("field_category", multi_field) == "lib"
                )
                prop_hash_current = (
                    None
                    if result_is_library
                    else self._property_dep_hash_current(
                        multi_field, recorder, validate=False
                    )
                )
                if (
                    prop_hash_current is not None
                    or result_is_library
                    or self.virtual_access_only(multi_field)
                ):
                    residence_key = (multi_field, prop_hash_current)
                    if builder:
                        _logger.debug("Writing to residence: %s", residence_key)
                    if not is_quble(multi_value) or not multi_value.is_undefined:
                        self._assign_to_residence(residence_key, multi_value)
                        Redis.set(
                            self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                            str(self.id)
                            + str(residence_key)
                            + str(multi_value.time_stamp),
                            datetime.now(),
                        )
            # Resolve field dependencies
            if sourced_from_file and isinstance(multi_value, DataLib):
                multi_value.resolve_field_dependencies()

            # ---------------------------------------------------------------
            # Write to file
            #
            # IMPORTANT: Must happen AFTER writing the field to residence so
            # the residence timestamp is available when writing the file.
            # ---------------------------------------------------------------
            file_storage = self.file_storage(multi_field) or self.touch_storage(
                multi_field
            )
            filename = self.get_property(
                "filename",
                multi_field,
                grace=True,
                resolve_templates=True,
                suppress_recording=True,
            )
            if sourced_from_build and multi_value is not None and file_storage:
                if isinstance(return_value, Quble) and filename and self.address:
                    addr_string = "/".join(self.address)
                    addr_string = addr_string + "/" + str(field)
                    procurement_duration_last[addr_string] = new_duration

                # Check if a multi_build table has already been written.
                # all_columns_xyz only needs to be written once.
                # If the table name is all_columns, then set the multi_build_is_written flag
                # and write the table. The next iteration of this loop, the write operation will not happen due to the
                # multi_build_is_written_flag.
                if not multi_build_is_written:
                    if multi_value.table_name.startswith("all_columns_"):
                        multi_build_is_written = True
                    if builder:
                        _logger.debug("Writing to file: %s", filename)
                    self.write_field(
                        multi_field, multi_value, residence_key=residence_key
                    )

                # quble table name has changed, update cache
                if not is_quble(multi_value) or not multi_value.is_undefined:
                    if self._exists_in_residence(residence_key):
                        tmp_val = self.residence[residence_key]
                        tmp_val.table_name = multi_value.table_name
                        self._assign_to_residence(residence_key, tmp_val)

            # ----------------------------
            # Update procurement duration
            # ----------------------------
            if record_duration and new_duration is not None:
                self.set_property(
                    "procurement_duration_last", new_duration, multi_field
                )
        address = address.maximize()
        if isinstance(return_value, Quble) and sourced_from_build:
            # Final build tree write. In the case of a recursion call, is_completed may be false,
            # If this is indeed the final write, then is_completed will be True
            if (filename) and (field in filename):
                node_table_name = filename
            else:
                node_table_name = return_value.table_name
            recorder.add_current_pointer_to_build_tree(
                procurement_duration_last=procurement_duration_last,
                node_table_name=node_table_name,
                completed_node=address,
                is_completed=recorder.is_origin,
                write_to_tree=write_to_tree,
            )
            # If this is the final write, then send the serialized build tree over the Messenger WebSocket
            # And set the build_tree_last property.
            if recorder.is_origin and not suppress_recording:
                self.set_property(
                    "build_tree_last",
                    dumps(recorder.serialize_build_tree()),
                    field,
                )
                Messenger.write_message("Persisting result...")

        # ---------------
        # Commit the lib
        # ---------------
        if lib_commit_flag and not isinstance(self, MultiBuildLib):
            self.commit()

        # -----------------------------------------------------------------------------
        # Return the procured value
        # -----------------------------------------------------------------------------
        if return_value is None:
            return None

        return_value.address = address
        fieldspace = self.get_property("fieldspace", grace=True)

        if isinstance(return_value, DataLib):
            if subkeys is None or subkeys.is_empty:
                return return_value
            elif fieldspace is None or fieldspace not in subkeys.keyspaces:
                # Content in subkeys is deemed to be 'orthogonal' to library and therefore does not change resolution
                return return_value
            else:
                return None
        else:
            if subkeys is None or subkeys.is_empty:
                return return_value
            else:
                raise TypeError(f"Invalid subkeys type encountered: {type(subkeys)}")

    def field_keyspaces_last(self, field, **kwargs):
        """
        Native property methods for generating field_keyspaces_last from properties: 'field_col_types_last' and 'field_valuespaces_last'

        May return a list, tuple or None
        """
        col_types = self.get_property(
            property_name="field_col_types_last", fieldref=field, **kwargs
        )
        if col_types is None:
            return None
        elif not isinstance(col_types, dict):
            raise TypeError

        spaces = list(col_types.keys())
        if len(spaces) == 0:
            return []

        valuespaces = self.get_property(
            property_name="field_valuespaces_last", fieldref=field, **kwargs
        )
        if isinstance(valuespaces, str):
            return [space1 for space1 in spaces if space1 != valuespaces]
        elif isinstance(valuespaces, (list, tuple)):
            return [space1 for space1 in spaces if space1 not in valuespaces]
        else:
            # No valid valuespaces recording, this field does not have a procurement history to leverage
            return None

    def field_val_types_last(self, field, auto_squeeze=False, **kwargs):
        """
        Native property method for: get_property('field_val_types_last',...) from properties: field_col_types_last and valuespaces_last

        May return a string (univariate) or a dictionary (keyed by valuespace)
        """

        return self._subspace_types(
            field=field,
            subspaces_property_name="field_valuespaces_last",
            col_types_property_name="field_col_types_last",
            auto_squeeze=auto_squeeze,
            **kwargs,
        )

    def field_key_types_last(self, field, auto_squeeze=False, **kwargs):
        """
        Native property method for get_property('field_key_types_last',...) from properties: field_col_types_last and valuespaces_last

        May return a string (univariate) or a dictionary (keyed by valuespace)
        """
        return self._subspace_types(
            field=field,
            subspaces_property_name="field_keyspaces_last",
            col_types_property_name="field_col_types_last",
            auto_squeeze=auto_squeeze,
            **kwargs,
        )

    def _subspace_types(
        self,
        field,
        subspaces_property_name,
        col_types_property_name,
        auto_squeeze=False,
        **kwargs,
    ):
        col_types = self.get_property(
            property_name=col_types_property_name,
            fieldref=field,
            try_json_loads=True,
            **kwargs,
        )
        if col_types is None:
            return None

        subspaces = self.get_property(
            property_name=subspaces_property_name,
            fieldref=field,
            try_json_loads=True,
            **kwargs,
        )

        # Build value_type_last (None, str or dict) from respective col_types and subspaces
        if subspaces is None:
            return None
        elif isinstance(col_types, str):
            if isinstance(subspaces, str):
                # if auto_squeeze, return as string, otherwise single-entry dict
                return col_types if auto_squeeze else {subspaces: col_types}
            elif isinstance(subspaces, (list, tuple)) and len(subspaces) == 1:
                # if auto_squeeze, return as string, otherwise single-entry dict
                return col_types if auto_squeeze else {subspaces[0]: col_types}
            else:
                return None

        elif isinstance(col_types, dict):
            if isinstance(subspaces, str):
                if subspaces in col_types:
                    # if auto_squeeze, return as string, otherwise single-entry dict
                    return (
                        col_types[subspaces]
                        if auto_squeeze
                        else {subspaces: col_types[subspaces]}
                    )
                else:
                    # if auto_squeeze, return as None, otherwise empty dict
                    return None if auto_squeeze else {}
            elif isinstance(subspaces, (list, tuple)):
                # Trap for auto_squeeze case
                if (
                    (len(subspaces) == 1)
                    and (subspaces[0] in col_types)
                    and auto_squeeze
                ):
                    return col_types[subspaces[0]]

                subcol_types = {}
                for space1 in subspaces:
                    if space1 in col_types:
                        subcol_types[space1] = col_types[space1]
                    else:
                        subcol_types[space1] = None
                return subcol_types
            else:
                return None

        else:
            return None

    def set_build_context(self, field):
        """Set global controls before resolving the field.

        Impose DataLib properties as RootLib 'environmental' controls (e.g. 'FX') during construction of fields.

        NOTE: Does nothing on the base class. Overridden on a case-by-case basis by subclasses.
        """
        pass

    def _check_freshness(
        self,
        field,
        property_dep_hash,
        force_build,
        validate,
        build_tree,
        suppress_recording,
    ):
        """Check the freshness of the residence and file caches.

        NOTE: DO NOT CALL THIS METHOD. This is only to be used inside ``get()``.

        :type field: str
        :param field: The field name

        :type property_dep_hash: str
        :param property_dep_hash: The current property dependency hash

        :type force_build: bool
        :param force_build: Do not check freshness. Consider all caches to be stale.

        :type validate: bool
        :param validate: When False, residence and/or file are automatically considered fresh, if they exist.

        :rtype: tuple
        :returns: (residence outdated, file outdated, residence key)
        """
        from qubles.io.base.rootlib import RootLib
        from qubles.io.util.multi_build_lib import MultiBuildLib

        failed_freshness_msg = "Failed field/prop dependency freshness test."
        if force_build:
            try:
                table_to_drop = None
                residence_key = (field, property_dep_hash)
                is_multi_build = isinstance(self, MultiBuildLib)
                if residence_key in self.residence:
                    table_to_drop = self.residence[residence_key].table_name
                    if is_multi_build:
                        Redis.pop(
                            self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                            str(self.parent_id)
                            + str(residence_key)
                            + str(self.residence[residence_key].time_stamp),
                        )
                    else:
                        Redis.pop(
                            self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                            str(self.id)
                            + str(residence_key)
                            + str(self.residence[residence_key].time_stamp),
                        )
                    self.residence.pop(residence_key, None)
                else:
                    prop_hash_last = self.get_property(
                        "property_dep_hash_last", field, grace=True
                    )
                    if prop_hash_last:
                        table_to_drop = f"{field}_{prop_hash_last}_qdb"

                if table_to_drop:
                    drop_table(table_to_drop)
            except Exception as e:
                _logger.error(
                    f"something went wrong purging the frontend cache as part of force rebuild: {e}"
                )

            return (True, True, None, None, None)

        residence_outdated = True
        file_outdated = True
        is_derived = self.is_derived_field(field)
        resync_time = RootLib().get_control("resync_time")

        # Get residence timestamp
        # ------------------------
        residence_key = (field, property_dep_hash)
        residence_timestamp = None
        residence_msg = None
        if not self._exists_in_residence(residence_key):
            residence_msg = f"missing key: {residence_key}"
        elif self.residence[residence_key] is None:
            residence_msg = f"value is null: {residence_key}"
        else:
            try:
                residence_timestamp = self.residence[residence_key].time_stamp
            except AttributeError:
                residence_msg = 'no "timestamp" attribute: {}'.format(
                    type(self.residence[residence_key])
                )

            if residence_timestamp and self._invalid_timestamp(residence_timestamp):
                residence_timestamp = None
                residence_msg = f"invalid timestamp: {residence_timestamp}"
            elif residence_timestamp is not None:
                # Validate the residence timestamp (skipping field/prop deps for performance)
                residence_outdated = self.field_fresher_than(
                    field,
                    [residence_timestamp],
                    check_property_deps=False,
                    check_field_deps=False,
                )[0]
                if residence_outdated:
                    residence_timestamp = None
                    residence_msg = failed_freshness_msg
                else:
                    # Shortcut! If inside the recync time window, just bail
                    is_multi_build = isinstance(self, MultiBuildLib)
                    if is_multi_build:
                        last_evaluated = Redis.get(
                            self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                            str(self.parent_id)
                            + str(residence_key)
                            + str(residence_timestamp),
                        )
                    else:
                        last_evaluated = Redis.get(
                            self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                            str(self.id)
                            + str(residence_key)
                            + str(residence_timestamp),
                        )

                    if last_evaluated and int(
                        (datetime.now() - last_evaluated).total_seconds()
                    ) <= int(resync_time * 60):
                        return (False, False, residence_key, residence_timestamp, None)
            else:
                residence_msg = "Value in residence had time_stamp of 'None'"

        if (
            self.read_field_category(field) == "data"
            and len(build_tree._nodes) == 0
            and not suppress_recording
        ):
            Messenger.write_message(
                f"Checking if data is up-to-date... Validating {field}"
            )

        # Get file timestamp
        # -------------------
        file_timestamp = self._file_accessible(field)
        file_msg = None
        if file_timestamp == False:
            file_msg = "not accessible"
        elif file_timestamp and self._invalid_timestamp(file_timestamp):
            file_timestamp = None
            file_msg = f"invalid timestamp: {file_timestamp}"
        elif file_timestamp is not None:
            # Validate the residence timestamp (skipping field/prop deps for performance)
            file_outdated = self.field_fresher_than(
                field,
                [file_timestamp],
                check_property_deps=False,
                check_field_deps=False,
            )[0]
            if file_outdated:
                file_timestamp = None
                file_msg = failed_freshness_msg
            elif residence_outdated:
                # Shortcut! If inside the recync time window, just bail
                is_multi_build = isinstance(self, MultiBuildLib)
                if is_multi_build:
                    last_evaluated = Redis.get(
                        self.REDIS_FILE_TIMESTAMP_KEY,
                        str(self.parent_id) + str(residence_key) + str(file_timestamp),
                    )
                else:
                    last_evaluated = Redis.get(
                        self.REDIS_FILE_TIMESTAMP_KEY,
                        str(self.id) + str(residence_key) + str(file_timestamp),
                    )
                if last_evaluated and int(
                    (datetime.now() - last_evaluated).total_seconds()
                ) <= int(resync_time * 60):
                    return (
                        True,
                        False,
                        residence_key,
                        residence_timestamp,
                        file_timestamp,
                    )

        # If validate=False, any non-null timestamp is considered valid
        # --------------------------------------------------------------
        if not validate and (residence_timestamp or file_timestamp):
            residence_outdated = residence_timestamp is None
            file_outdated = file_timestamp is None

            return (
                residence_outdated,
                file_outdated,
                residence_key,
                residence_timestamp,
                file_timestamp,
            )

        if residence_outdated and file_outdated:
            # If both are invalid, bail early
            return (
                residence_outdated,
                file_outdated,
                residence_key,
                residence_timestamp,
                file_timestamp,
            )

        # Only need to validate the fresher timestamp of the two
        # -------------------------------------------------------
        if residence_timestamp and file_timestamp:
            if file_timestamp > residence_timestamp:
                residence_timestamp = None
                residence_outdated = True
                if is_derived:
                    file_timestamp = None
                    file_msg = "updated outside builder"
            else:
                file_outdated = True
                file_timestamp = None
                file_msg = "residence was fresher"

        # ------------------------------------------
        # Consult ad_hoc setting
        # ------------------------------------------
        ad_hoc = self.get_property(
            "ad_hoc",
            field,
            grace=True,
            suppress_recording=True,
            default_property_value=False,
        )

        # At this point, we know we either residence or file isn't outdated, so no need to build
        if ad_hoc:
            return (
                residence_outdated,
                file_outdated,
                residence_key,
                residence_timestamp,
                file_timestamp,
            )

        # Consult field_fresh_time_override
        # ------------------------------------
        fresh_time_override = self.get_property(
            "field_fresh_time_override",
            field,
            grace=True,
            appeal=True,
            suppress_recording=True,
        )

        valid_fresh_time_override = fresh_time_override is None or isinstance(
            fresh_time_override, datetime
        )
        if not valid_fresh_time_override:
            raise TypeError(
                "Invalid field_fresh_time_override: {}... expected None, or "
                "datetime".format(fresh_time_override)
            )

        if self._valid_timestamp(fresh_time_override):
            for timestamp in [residence_timestamp, file_timestamp]:
                if timestamp is not None and self._later_than(
                    fresh_time_override, timestamp
                ):
                    _logger.debug(
                        "Stale: %s: fresh time override (%s) > %s",
                        field,
                        fresh_time_override,
                        timestamp,
                    )
                    return (True, True, None, None, None)

        # Check if the field is fresher than the residence/file
        # ------------------------------------------------------
        if residence_timestamp:
            residence_outdated = self.field_fresher_than(
                field,
                [residence_timestamp],
                check_property_deps=[is_derived],
                property_dep_hash=property_dep_hash,
            )[0]
            residence_msg = failed_freshness_msg
        elif file_timestamp:
            # 1. Read-only fields are not dependent on the property hash
            # 2. Fields that already include the property hash in the filename do not need to check it again
            file_outdated = self.field_fresher_than(
                field,
                [file_timestamp],
                check_property_deps=[is_derived],
                property_dep_hash=property_dep_hash,
            )[0]
            file_msg = failed_freshness_msg

        if is_derived:
            if residence_outdated:
                _logger.debug(
                    f"Residence Outdated for field {field} with timestamp {residence_timestamp} and property hash {property_dep_hash}"
                )
                _logger.debug(f"Reason: {residence_msg}")
            else:
                _logger.debug(
                    f"Residence Fresh for field {field} with timestamp {residence_timestamp} and property hash {property_dep_hash}"
                )

            if residence_outdated and residence_msg is not None and file_outdated:
                _logger.debug(
                    f"File Outdated for field {field} with timestamp {file_timestamp} and property hash {property_dep_hash}"
                )
                _logger.debug(f"Reason: {file_msg}")
            elif not file_outdated:
                _logger.debug(
                    f"File Fresh for field {field} with timestamp {file_timestamp} and property hash {property_dep_hash}"
                )

        # set current timestamp as last evaluated
        if not residence_outdated:
            is_multi_build = isinstance(self, MultiBuildLib)
            if is_multi_build:
                Redis.set(
                    self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                    str(self.parent_id) + str(residence_key) + str(residence_timestamp),
                    datetime.now(),
                )
            else:
                Redis.set(
                    self.REDIS_RESIDENCE_TIMESTAMP_KEY,
                    str(self.id) + str(residence_key) + str(residence_timestamp),
                    datetime.now(),
                )

        if not file_outdated:
            is_multi_build = isinstance(self, MultiBuildLib)
            if is_multi_build:
                Redis.set(
                    self.REDIS_FILE_TIMESTAMP_KEY,
                    str(self.parent_id) + str(residence_key) + str(file_timestamp),
                    datetime.now(),
                )
            else:
                Redis.set(
                    self.REDIS_FILE_TIMESTAMP_KEY,
                    str(self.id) + str(residence_key) + str(file_timestamp),
                    datetime.now(),
                )

        return (
            residence_outdated,
            file_outdated,
            residence_key,
            residence_timestamp,
            file_timestamp,
        )

    def _property_dep_hash_current(
        self, field, recorder, include_field_deps=True, validate=True
    ):
        """Generate a hash from the current stored property dependencies.

        NOTE: DO NOT CALL THIS METHOD. This is only to be used inside ``get()`` when request_type=field_property_dep_hash.
        To retrieve the current property hash, call ``lib.get(field, request_type='field_property_dep_hash')``.
        """
        if recorder.current_property_dep_hash_current is not None:
            return recorder.current_property_dep_hash_current

        property_dep_dict = self.local_property_dep_dict_current(field)
        if property_dep_dict is not None:
            property_dep_hash = make_hash(property_dep_dict)
        else:
            # This field has never been procured before
            _logger.debug(
                f"_property_dep_hash_current: {field} property_dep_dict is None"
            )
            return None

        hashes = [property_dep_hash]
        if include_field_deps:
            field_deps = self.field_dependencies_current(field)
            if field_deps is None:
                # The field dependencies have never been calculated, so the aggregate hash is invalid
                _logger.debug(f"_property_dep_hash_current: {field} field_deps is None")
                return None

            for dep_lib, dep_field in field_deps:
                try:
                    dep_property_dep_hash = dep_lib.get(
                        dep_field,
                        request_type="field_property_dep_hash",
                        key_grace=True,
                        suppress_recording=True,
                        validate=validate,
                    )
                except UnregisteredFieldError:
                    _logger.warning(
                        f"_property_dep_hash_current: {field}'s dependency {dep_field} was unregisterd at domain {dep_lib}"
                    )
                    return None

                if dep_property_dep_hash is None:
                    # The aggregated property dependency hash is only valid if every dependency has a valid hash
                    _logger.debug(
                        f"_property_dep_hash_current: {field}'s dependency {dep_field} property hash is None"
                    )
                    return None

                hashes.append(dep_property_dep_hash)

        pdhc = make_hash(hashes)
        recorder.current_property_dep_hash_current = (
            pdhc  # <-- Assign for cached benefit
        )

        if pdhc != self.get_property(
            "property_dep_hash_last", field, grace=True, suppress_recording=True
        ):
            _logger.debug(
                f"PROPERTY HASHES DON'T MATCH for FIELD:{field} ---> {pdhc} != {self.get_property('property_dep_hash_last', field, grace=True, suppress_recording=True)}"
            )
            _logger.debug(property_dep_dict)
            _logger.debug(
                f"{self.get_property('property_dep_dict_last', field, grace=True, suppress_recording=True)}"
            )
        return pdhc

    def _field_dependencies_current(self, field):
        """Return all field dependencies (explicit and implicit) for the given field.

        NOTE: DO NOT CALL THIS METHOD. This is only to be used inside ``get()`` when request_type=field_dependencies.
        To retrieve the current field dependencies, call ``lib.get(field, request_type='field_dependencies')``.

        """

        field_deps = self.get_property(
            "field_dependencies",
            field,
            grace=True,
            appeal=False,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        prop_deps = self.parse_property_dependencies(
            field,
            grace=True,
            appeal=False,
            native_property_mode=NPModes.NONE,
            suppress_recording=True,
        )

        if field_deps is None or prop_deps is None:
            # The field has never been accessed, so the dependencies have not been recorded
            return None

        all_deps = []
        for dep_address in field_deps:
            all_deps.append((LibAddress(dep_address), False, None))

        for prop_dep in prop_deps:
            if not prop_dep["is_field_dependency"]:
                continue

            dep_address = prop_dep["domain"].get_property(
                prop_dep["property_name"],
                prop_dep["field_arg"],
                appeal=prop_dep["appeal"],
                resolve_templates=prop_dep["resolve_templates"],
                grace=True,
                suppress_recording=True,
            )

            if not dep_address:
                continue

            if not isinstance(dep_address, LibAddress):
                dep_address = LibAddress(dep_address, base_domain=prop_dep["domain"])

            all_deps.append((dep_address, True, prop_dep))

        parsed_deps = []
        for dep_address, sourced_from_prop, prop_dep in all_deps:
            if dep_address is None or dep_address.is_empty:
                continue

            try:
                dep_lib, dep_field, _ = dep_address.min_subparts
            except UnregisteredFieldError as e:
                _logger.critical(
                    f"UnregisteredField!!!! {prop_dep} {field} {self}  {dep_address}"
                )
                raise e

            # Field dependencies stored in properties might resolve to screens.
            # For these cases, we must resolve the field in order to know the type.
            # ----------------------------------------------------------------------
            field_value = None

            if sourced_from_prop and (
                dep_lib.get_property(
                    "field_category",
                    dep_field,
                    grace=True,
                    suppress_recording=True,
                )
                != "data"
            ):
                field_value = dep_lib.get(
                    dep_field,
                    request_type="field_value",
                    key_grace=True,
                    suppress_recording=True,
                    log=False,
                )

            if is_screen(field_value):
                if field_value.is_empty:
                    continue
                dep_lib = field_value
                dep_field = field_value.active_rule_fields[-1]

            parsed_deps.append((dep_lib, dep_field))

        return parsed_deps

    def _handle_unregistered_field(self, field, auto_register, key_grace, request_type):
        if auto_register == "auto_register":
            auto_register = self.get_property("auto_register", grace=True)

        if auto_register:
            self.register_field(field)
            self.touch()
        elif ((key_grace == "key_grace") and self.key_grace()) or (key_grace == True):
            if request_type != "field_value":
                raise GracefulUnregisteredFieldError(
                    "Gracefully exiting, should be caught!"
                )
        else:
            raise UnregisteredFieldError(
                f"'{field}' not contained in {self}::field_index -> {self.fields()}."
            )

    def _call_builder(self, field, builder, recorder, **kwargs):
        """
        Calls the given builder function, passing in kwargs. Returns the result.
        Add field arg (when expected).
        """
        if "field" in getfullargspec(builder).args:
            kwargs["field"] = field
        if "fieldref" in getfullargspec(builder).args:
            kwargs["fieldref"] = field

        if self.address:
            recorder.add_current_pointer_to_build_tree()

        return builder(**kwargs)

    def _get_instance_sample(self, fieldref):
        from qubles.core.quble import Quble

        # Handle list/tuple case...
        # ----------------------------
        if isinstance(fieldref, list) or isinstance(fieldref, tuple):
            instance_samples = []
            for fref in fieldref:
                instance_samples.append(self.instance_sample(fieldref=fref))
            return instance_samples

        # Procure lib that contains fieldref (may be an embedded lib)
        # --------------------------------------------------------------
        lib = self
        field = fieldref

        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        # Validate field...
        # ---------------------------
        if field not in lib.field_index:
            raise ValueError(f"Unsupported field: {field}")

        # Proceed with load of field from lib
        # [Here lib is not necessarily self, may be imdedded library containing requested field]
        # ------------------------------------------------------------

        # Step 1: Try to access virtual in-memory storage (if applicable)...
        # ---------------------------------------------------------------------
        property_dep_hash = self.property_dep_hash_current(field, validate=False)
        residence_key = (field, property_dep_hash)

        if self._exists_in_residence(residence_key):
            return self.residence[residence_key]
        elif not lib.is_derived_field(field):
            # If not derived, try to read the type
            if lib.read_grace(field) and (
                not lib.read_access(field)
                or not lib.inspect_file(
                    field, validate=False, check_property_deps=False
                )
            ):  # <-- Only honor access_grace(read)=True logic if: 1) NOT read_access OR 2) file does NOT exist
                return None
            elif not lib.read_access(field):
                raise BuilderError(
                    "Non-derived field: %s: read access denied...virtual assignment/storage should apply (but did not)"
                    % (field)
                )
            elif not lib.inspect_file(field, check_property_deps=False):
                raise BuilderError(f"Non-derived field: {field}: stale or absent")

            field_type = lib.read_field_type(field)
        elif lib.read_access(field) and lib.inspect_file(
            field, validate=True, check_property_deps=True
        ):
            # If derived & file available & valid: try to read the type
            _logger.debug(
                f"Procurring instance sample for derived field: {self}:{field}"
            )
            field_type = lib.read_field_type(field)
        else:
            # Assume that all non-present derived items are Qubles
            return Quble.undefined_instance()

        # Instantiate accordingly
        # --------------------------
        if field_type is None:
            raise ValueError(f"field_type could not be determined for field: {field}")

        # Import and instantiate the proper class
        # ------------------------------------------
        cls = libtypes.get(field_type, grace=True)

        if cls:
            return cls()

        return None

    def key_grace(self, fieldref=None, grace=True):
        """Return a boolean interpretation of access_grace for invalid key usage.

        Will return True if self.get_property('access_grace') is either True or a string containing sub-string 'key'.
        """
        lib = self
        field = fieldref
        access_grace = lib.get_property("access_grace", field, grace=grace)
        if access_grace is None:
            return False
        elif not isinstance(access_grace, str) and not isinstance(access_grace, bool):
            raise TypeError(
                "Invalid property value for access_grace:{0}...None,string or bool expected".format(
                    access_grace
                )
            )
        elif isinstance(access_grace, str) and search("key", access_grace):
            return True
        elif isinstance(access_grace, bool) and access_grace:
            return True
        else:
            return False

    def read_grace(self, fieldref=None, grace=True):
        """Return a boolean interpretation of access_grace for (file) read failure.

        Will return True if self.get_property('access_grace') is either True or a string containing sub-string 'read'.
        """
        lib = self
        field = fieldref
        access_grace = lib.get_property("access_grace", field, grace=grace)
        if access_grace is None:
            return False
        elif not isinstance(access_grace, str) and not isinstance(access_grace, bool):
            raise BuilderError(
                "Invalid property value for access_grace:{0}...None,string or bool expected".format(
                    access_grace
                )
            )
        elif isinstance(access_grace, str) and search("read", access_grace):
            return True
        elif isinstance(access_grace, bool) and access_grace:
            return True
        else:
            return False

    def read_field_time_stamp(
        self,
        fieldref,
        file_type_property_name="file_type",
        path_property_name="path",
        filename_property_name="filename",
        grace=False,
    ):
        """Read field's timestamp from library."""
        lib = self
        field = fieldref

        file_type = lib.get_property(
            file_type_property_name, field, grace=grace
        ).upper()

        if file_type == "SNOWFLAKE":
            residence_key = (field, lib.property_dep_hash_current(field))
            if lib._exists_in_residence(residence_key):
                return lib.residence[residence_key].time_stamp

            field_category = lib.read_field_category(field)
            if field_category == "lib":
                return read_last_updated(lib.id, lib.field_to_id_map[field])
            elif field_category == "data":
                return snowflake_get_quble_timestamp(
                    lib.get_property("filename", field, resolve_templates=True)
                )
        elif file_type == "LIBADDRESS":
            (domain, local_field, _) = lib.lib_address_credentials(
                field,
                path_property_name=path_property_name,
            )
            if domain is None:
                raise UnresolvedDomainError(field=field)
            else:
                return domain.read_field_time_stamp(local_field, grace=grace)
        elif file_type == "QDB":
            (filename, objname) = lib.qdb_credentials(
                field,
                path_property_name=path_property_name,
                filename_property_name=filename_property_name,
            )
            filename = normalize(filename)
            from qubles.io.util.qdb import QDB

            if not os.path.exists(filename):
                if grace:
                    return missing_date
                else:
                    raise IOError(f"Bad file_name: {filename}")

            with closing(QDB(filename, "r")) as qdb:
                time_stamp = qdb.time_stamp(objname)

            return time_stamp
        elif grace:
            return missing_date
        else:
            raise ValueError(f"Bad file_type: {file_type}")

    def read_field_category(
        self,
        fieldref,
        file_type_property_name="file_type",
        path_property_name="path",
        filename_property_name="filename",
        grace=True,
    ):
        """
        Reads fieldref's category from library
        """

        lib = self
        field = fieldref

        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        file_type = lib.get_property(
            file_type_property_name, field, grace=False
        ).upper()

        # Check if we are a linked field, and forward the request if so
        # ------------------------
        if file_type == "LIBADDRESS":
            (domain, local_field, _) = lib.lib_address_credentials(
                field,
                path_property_name=path_property_name,
            )
            if domain is None:
                raise UnresolvedDomainError(field=field)
            else:
                return domain.read_field_category(local_field)
        else:
            # Otherwise, attempt to fetch the assignment from props directly
            field_category = lib.get_property("field_category", field, grace=grace)
            if field_category is not None:
                return field_category
            elif lib.is_derived_field(field):
                lib.set_property("field_category", "data", field)
                return "data"
            # If no field_category assignement is found, confirm file_type is SNOWFLAKE and we aren't virtual_storage_only,
            # and attempt to collect that information from the db or throw an appropriate exception otherwise
            elif file_type == "SNOWFLAKE":
                if self.virtual_storage_only(field):
                    raise BuilderError(
                        f"Field category not assigned for virtual access only field: {field}"
                    )
                return properties_field_select(self.id, self.field_to_id_map[field])[5]
            elif file_type == "QDB":
                from qubles.io.util.qdb import QDB

                (filename, objname) = lib.qdb_credentials(
                    field,
                    path_property_name=path_property_name,
                    filename_property_name=filename_property_name,
                )
                filename = normalize(filename)

                if not os.path.exists(filename):
                    raise IOError(f"Bad file_name: {filename}")

                with closing(QDB(filename, "r")) as qdb:
                    type_str = qdb.kind(objname)

                return "data" if type_str in ("Qube", "Quble") else "lib"
            else:
                raise ValueError(
                    f"Unsupported file_type: {file_type} for field: {field}"
                )

    def read_field_type(
        self,
        fieldref,
        file_type_property_name="file_type",
        path_property_name="path",
        filename_property_name="filename",
        grace=True,
    ):
        """
        Reads fieldref's type from library
        """

        lib = self
        field = fieldref

        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        file_type = lib.get_property(
            file_type_property_name, field, grace=False
        ).upper()

        # Check if we are a linked field, and forward the request if so
        # ------------------------
        if file_type == "LIBADDRESS":
            (domain, local_field, _) = lib.lib_address_credentials(
                field,
                path_property_name=path_property_name,
            )
            if domain is None:
                raise UnresolvedDomainError(field=field)
            else:
                return domain.read_field_type(local_field)
        else:
            # Otherwise, attempt to fetch the assignment from props directly
            field_type = lib.get_property("field_type", field, grace=grace)
            if field_type is not None:
                return field_type
            # If no field_type assignement is found, confirm file_type is SNOWFLAKE and we aren't virtual_storage_only,
            # and attempt to collect that information from the db or throw an appropriate exception otherwise
            elif file_type == "SNOWFLAKE":
                if self.virtual_storage_only(field):
                    raise BuilderError(
                        f"Field type not assigned for virtual access only field: {field}"
                    )
                return properties_field_select(self.id, self.field_to_id_map[field])[4]
            elif file_type == "QDB":
                from qubles.io.util.qdb import QDB

                (filename, objname) = lib.qdb_credentials(
                    field,
                    path_property_name=path_property_name,
                    filename_property_name=filename_property_name,
                )
                filename = normalize(filename)

                if not os.path.exists(filename):
                    raise IOError(f"Bad file_name: {filename}")

                with closing(QDB(filename, "r")) as qdb:
                    type_str = qdb.kind(objname)

                return type_str
            else:
                raise ValueError(
                    f"Unsupported file_type: {file_type} for field: {field}"
                )

    def read_field(
        self,
        fieldref,
        file_type_property_name="file_type",
        path_property_name="path",
        time_stamp=None,
    ):
        """Reads <field> from the data library"""
        from qubles.core.quble import Quble

        address = LibAddress(fieldref, self)
        (lib, field, subkeys) = address.min_subparts
        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        file_type = lib.get_property(file_type_property_name, field).upper()

        if file_type == "SNOWFLAKE":
            field_category = lib.read_field_category(field)
            if field_category == "lib":
                result = get_lib(self.field_to_id_map[field], self.name, is_root=False)
            elif field_category == "data":
                result = Quble.from_table(
                    lib.get_property(
                        "filename",
                        field,
                        grace=False,
                        resolve_templates=True,
                        suppress_recording=True,
                    ),
                    time_stamp=time_stamp,
                )
            else:
                raise ValueError(
                    f"ERROR: SNOWFLAKE Libs require a valid field_category prop, but received: {field_category}, for lib|field: {lib}|{field}."
                )
        elif file_type == "LIBADDRESS":
            (domain, local_field, _) = lib.lib_address_credentials(
                field,
                path_property_name=path_property_name,
            )
            if domain is None:
                raise UnresolvedDomainError(field=field)
            else:
                return domain.get(local_field)
        else:
            raise ValueError(f"Bad file_type: {file_type}")

        if subkeys is not None:
            result = result.get(subkeys)
        result.address = address

        if is_quble(result):
            pass
        elif is_screen(result):  # Screen fields are sorted by default
            result.field_index = make_index(sorted(list(result.field_index)))
        elif result.__class__.__name__ not in (
            "Portfolio",
            "Backtest",
            "Forecast",
            "Screen",
            "RootLib",
        ) and is_user_data(
            result.address.globalize(), get_usernames()
        ):  # Certain libs have field order that are specific
            result = sort_field_index(result)

        return result

    def read_underlying(
        self,
        fieldref,
        file_type_property_name="underlying_file_type",
        path_property_name="underlying_path",
    ):
        """
        Reads an underlying field data according to the field-specific properties:
            'underlying_file_type'
            'underlying_path'
            'underlying_filename'

        NOTE: This method designed for use in derived DataLibs.
        """

        return self.read_field(
            fieldref,
            file_type_property_name=file_type_property_name,
            path_property_name=path_property_name,
        )

    # ====================================== Set ======================================

    def __setitem__(self, key, value):
        return self.set(key, value)

    def set(
        self,
        fieldref,
        value,
        destruction_prep=False,
        validate=True,
        redefine="auto_redefine",
        auto_register="auto_register",
        grace=False,
    ):
        """Set the field to the given value.

        :type destruction_prep: bool
        :param destruction_prep: Indicates that this DataLib.set() call was instigated by a Quble/DataLib destructor call..
                                    special internal logic will apply

        :type validate: bool
        :param validate: Flag to indicate to check time_stamp before writing the result to field. If validate=True (and write storage is
                            applicable), then new field will be written only if: new value timestamp > file storage timestamp

        :type redefine: bool
        :param redefine: Flag to indicate whether to record a new field "definition" with the assignment.
                            If redefine_flag='auto_redefine', the self.get_property(key, 'auto_redefine') will be consulted.

        """
        from qubles.io.ref.reflib import RefLib

        if value and getattr(value, "is_temporary", False):
            return

        address = LibAddress(fieldref, self)
        if value.address is None:
            value.address = address
        lib, field, subkeys = address.min_subparts

        # ----------------------------------------------------------------
        # Register field within field_index (if required & allowed)...
        # ----------------------------------------------------------------
        #   1) field is currently absent from lib
        #   2) library's default "store_mode" is not None (not no write)
        #   3) library supports "auto registration" of (new) fields
        # ----------------------------------------------------------------
        if field not in lib.field_index:
            if auto_register == "auto_register":
                auto_register = lib.get_property("auto_register", grace=True)

            if lib.no_storage(field, grace=True):
                raise ValueError(f"Storage prohibited for field: {field}")
            elif auto_register:
                lib.register_field(field)
                lib.touch()
            else:
                raise ValueError("auto_registration prohibited for field: {0}", field)

        # ------------------
        # Check store_mode
        # ------------------
        if lib.no_storage(field):
            if destruction_prep:
                lib.vacate(field)
                return
            elif grace:
                return
            else:
                raise ValueError(f"Storage prohibited for field: {field}")

        # Set field_category
        lib.set_property("field_category", "data" if is_quble(value) else "lib", field)

        # -------------------------------
        # Handle non-trivial hyper case
        # -------------------------------
        if subkeys is not None:
            # Deep copy is not needed here. Note that we were actually given a
            # subset value
            sub_value = value

            # May be linked to virtual object in DataLib Dictionary
            value = lib.get(field)

            # Should be calling value (Quble) set() method - not DataLib set method
            value.set(subkeys, sub_value)

        virtual_storage_flag = lib.virtual_storage(field)

        if virtual_storage_flag:
            property_dep_hash = self.property_dep_hash_current(field)
            residence_key = (field, property_dep_hash)
        else:
            residence_key = None

        # ------------------------------------------------------------------
        # Virtual Dictionary Storage / Update when NOT Destruction Prep...
        # ------------------------------------------------------------------
        if not destruction_prep and virtual_storage_flag:
            self._assign_to_residence(residence_key, value)

        # ----------------------------------------------------------------------------
        # Record a field redefinition (if applicable) when not destruction prep mode
        # ----------------------------------------------------------------------------
        if not destruction_prep:
            if redefine == "auto_redefine":
                redefine = lib.get_property("auto_redefine", field)

                if redefine is None:
                    redefine = False

            if redefine is True:  # <-- Trying to be robust to other values of redefine
                lib.set_property("field_defn_time", datetime.now(), field)

            # When assigning a RefLib object, update self._reflibs_cache if/when applicable
            if (
                self._reflibs_cache is not None
                and isinstance(value, RefLib)
                and value not in self._reflibs_cache
            ):
                self._reflibs_cache.append(value)

        # -------------------------------------
        # Library external file operations...
        # -------------------------------------
        if lib.file_storage(field) or lib.touch_storage(field):
            lib.write_field(
                field, value, validate=validate, residence_key=residence_key
            )

        # ---------------------------------------------------
        # Virtual Dictionary Purge when Destruction Prep...
        # ---------------------------------------------------
        if (
            virtual_storage_flag
            and residence_key in lib.residence
            and lib.residence[residence_key] is value
            and destruction_prep
        ):
            lib.residence.pop(residence_key)

        # Finally, if we did any storing, ensure field_type/category are populated
        if virtual_storage_flag or lib.file_storage(field) or lib.touch_storage(field):
            # If we don't have field_type/category assigned, do it now
            if self.get_property("field_type", field, grace=True) is None:
                lib.set_property("field_type", value.__class__.__name__, field)
            self.set_property(
                "field_category", "data" if is_quble(value) else "lib", field
            )  # Overwrite doesn't matter here

    def write_field(self, field, result, validate=True, residence_key=None):
        """Writes the given value to the file associated with the given field.

        :param field: The name of the field.
        :type field: str

        :param result: The value to write.
        :type result: DataLib or Quble

        :param validate: When True, write is forced. Otherwise, the write only happens if the result's timestamp > the file time stamp.
        :type validate: bool

        :param residence_key: If provided, will touch the associated residence item
                                (if present) to keep the residence item fresh relative to subsequent file comparisons.
        :type residence_key: 2-tuple of (field, property dep hash)

        """
        file_type = self.get_property("file_type", field).upper()

        if file_type == "SNOWFLAKE":
            if isinstance(result, DataLib):
                set_lib(self, result)
            elif is_quble(result):
                if result.is_undefined:
                    return
                tgt_table_name = self.get_property(
                    "filename", field, resolve_templates=True, suppress_recording=True
                )
                if tgt_table_name is None:
                    raise ValueError(
                        f"Error writing field: {field} of library: {self}. Missing filename property."
                    )

                tgt_table_name = (
                    tgt_table_name.replace(".", "_")
                    if "." in tgt_table_name
                    else tgt_table_name
                )

                # THIS APPROACH NEEDS TO SET A FLAG ON THE RESULT, TO PREVENT INTERACTION WITH THE TABLE BEFORE IT IS TRANSITIONED
                # THAT CAN BE DONE POST REDIST UPDATE, SO COMMENTING THE BELOW AND REVERTING TO THE OLDER SLOWER SAVE TEMPORARILY
                # self._write_quble.delay(
                #     result.table_name,
                #     tgt_table_name,
                #     self.get_property("field_type", field, grace=True),
                #     result.column_info,
                #     result.time_stamp,
                # )

                snowflake_persist_quble(
                    result.table_name, tgt_table_name, result.time_stamp
                )

                # Finally, assign the new table_name to the result, update the comments & timestamp
                result.table_name = tgt_table_name
                multi_col_info_custom_writer(result.table_name, result._column_info)
            else:
                raise TypeError(
                    f"ERROR: Recieved bad result, can only write DataLib/Quble fields but recieved: {type(result)}"
                )
        elif file_type == "LIBADDRESS":
            domain, local_field, _ = self.lib_address_credentials(field)

            if domain is None:
                raise UnresolvedDomainError(field=field)

            return domain.write_field(
                local_field, result, validate=validate, residence_key=residence_key
            )
        else:
            raise ValueError(f"Bad file_type: {file_type}")

    # ============================ Storage Mode Short-Cuts ============================

    def virtual_storage(self, fieldref=None, grace=False, suppress_recording=False):
        store_mode = self.get_property(
            "store_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (store_mode is None) or not search("virtual", store_mode):
            return False
        else:
            return True

    def file_storage(self, fieldref=None, grace=False, suppress_recording=False):
        store_mode = self.get_property(
            "store_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (store_mode is None) or not search("write", store_mode):
            return False
        else:
            return True

    def touch_storage(self, fieldref=None, grace=False, suppress_recording=False):
        store_mode = self.get_property(
            "store_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (store_mode is None) or not search("touch", store_mode):
            return False
        else:
            return True

    def no_storage(self, fieldref=None, grace=False, suppress_recording=False):
        store_mode = self.get_property(
            "store_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if store_mode is None:
            return True
        else:
            return False

    def read_only(self, fieldref=None, grace=False, suppress_recording=False):
        return not self.file_storage(
            fieldref, grace=grace, suppress_recording=suppress_recording
        ) and not self.touch_storage(
            fieldref, grace=grace, suppress_recording=suppress_recording
        )

    def virtual_storage_only(
        self, fieldref=None, grace=False, suppress_recording=False
    ):
        return (
            self.virtual_storage(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
            and not self.file_storage(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
            and not self.touch_storage(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
        )

    # ============================ Access Mode Short-Cuts =============================

    def accessible(self, fieldref=None, grace=False, suppress_recording=False):
        access_mode = self.get_property(
            "access_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if access_mode is None:
            return False
        else:
            return True

    def virtual_access(self, fieldref=None, grace=False, suppress_recording=False):
        access_mode = self.get_property(
            "access_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (access_mode is None) or not search("virtual", access_mode):
            return False
        else:
            return True

    def virtual_access_only(self, fieldref=None, grace=False, suppress_recording=False):
        return (
            self.virtual_access(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
            and not self.read_access(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
            and not self.build_access(
                fieldref, grace=grace, suppress_recording=suppress_recording
            )
        )

    def read_access(self, fieldref=None, grace=False, suppress_recording=False):
        access_mode = self.get_property(
            "access_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (access_mode is None) or not search("read", access_mode):
            return False
        else:
            return True

    def build_access(self, fieldref=None, grace=False, suppress_recording=False):
        access_mode = self.get_property(
            "access_mode", fieldref, grace=grace, suppress_recording=suppress_recording
        )
        if (access_mode is None) or not search("build", access_mode):
            return False
        else:
            return True

    @property
    def is_root(self):
        from qubles.io.base.rootlib import RootLib

        if self == RootLib():
            return True
        else:
            return False

    # ========================== File Credentials Short-Cuts ==========================

    def aggr_builder(self, field):
        """
        Special builder method that aggregates across a sub-set of other fields in a library
        """
        from qubles.core.quble import Quble
        from qubles.io.base.rootlib import RootLib

        root_lib = RootLib()

        lib = self
        if (field is None) or (field not in self.field_index):
            raise ValueError(f"Unsupported field: {field}")

        # -------------------------------------------------
        # PARSE aggrop INTO PARTS (delimiter = '|'):
        #    [0]: inclusion property name
        #    [1]: aggregation operator
        #    [2] = post aggrrgation operation
        # -------------------------------------------------
        aggrop = lib.get_property("aggrop", field, grace=True)

        if aggrop is None:
            raise ValueError("No aggrop property assigned")
        elif not isinstance(aggrop, str):
            raise TypeError("Invalid aggrop property value: string expected")

        aggrop_parts = aggrop.split("|")
        if len(aggrop_parts) < 2:
            raise ValueError(
                "Invalid aggrop: atleast two parameters required (delimiter=|)"
            )

        # Procure inclusion_name
        # ----------------------------
        if not isinstance(aggrop_parts[0], str) or (len(aggrop_parts[0]) == 0):
            raise ValueError("Invalid/empty inclusion property name (aggrop_parts[0])")

        inclusion_name = aggrop_parts[0]

        if not isinstance(aggrop_parts[1], str) or (len(aggrop_parts[1]) == 0):
            raise ValueError(
                "Invalid/empty aggrgation function abbrev (aggrop_parts[1])"
            )

        # Procure aggrop_fn_name & weight_property_name
        # ------------------------------------------------
        weight_property_name = None
        # Case 1: Operator Format = 'WTD_'<OP>
        if aggrop_parts[1][0:4].upper() == "WTD_":
            temp_parts = aggrop_parts[1][4:].split(":")
            if len(temp_parts) > 1:
                weight_property_name = temp_parts[1]
            aggrop_fn_name = f"WTD_{temp_parts[0].upper()}"
        # Case 2: Operator Format = 'WTD'<OP>
        elif aggrop_parts[1][0:3].upper() == "WTD":
            temp_parts = aggrop_parts[1][3:].split(":")
            if len(temp_parts) > 1:
                weight_property_name = temp_parts[1]
            aggrop_fn_name = f"WTD_{temp_parts[0].upper()}"
        # Case 3: Operator Format = 'W'<OP>
        elif aggrop_parts[1][0:1].upper() == "W":
            temp_parts = aggrop_parts[1][1:].split(":")
            if len(temp_parts) > 1:
                weight_property_name = temp_parts[1]
            aggrop_fn_name = f"WTD_{temp_parts[0].upper()}"
        # Case 4: Operator Format = <OP>
        else:
            aggrop_fn_name = aggrop_parts[1].upper()

        # Validate resultant aggrop_fn_str
        if aggrop_fn_name not in (
            "AVE",
            "MEAN",
            "MEDIAN",
            "SUM",
            "WTD_MEAN",
            "WTD_AVE",
            "WTD_MEDIAN",
            "WTD_SUM",
        ):
            raise ValueError(f"Invalid/unsupported aggrop_fn_str:{aggrop_fn_name}")

        # Procure postop settings (if applicable)
        postop_fn_name = None
        postop_keyspace = None
        if len(aggrop_parts) > 2:
            temp_parts = aggrop_parts[2].split(":")
            if len(temp_parts) > 1:
                postop_keyspace = temp_parts[1]
            postop_fn_name = temp_parts[0]

        # -------------------------------------
        # PROCURE AGGREGATION FIELD LIST...
        # -------------------------------------
        sub_fields = make_index(lib.fields(inclusions={inclusion_name: field}))
        if sub_fields.is_empty:
            return Quble()

        # -----------------------------------------------------
        # CONSTUCT COLLECTIVE MULTI-FACTOR DATA (DIM+1)...
        # -----------------------------------------------------
        multi_data = None
        multi_wts = None
        id_keyspace = lib.get_property("id_keyspace", grace=False)
        aggr_keyspace = "<AGGR>"  # <-- This is an arbitrary keyspace choice as it is merely a internal place-holder (will not be present in result)
        for sub_field in sub_fields:
            local_data = self.get(sub_field)
            if local_data is None:
                continue

            if (weight_property_name is not None) and aggrop_fn_name[0:4] == "WTD_":
                local_weight = lib.get_property(
                    weight_property_name, sub_field, grace=True
                )
            else:
                local_weight = None

            # Procure (local) preop settings (if applicable)
            preop_fn_name = None
            preop_keyspace = None
            preop = lib.get_property("preop", sub_field, grace=True)
            if (preop is not None) and len(preop) > 0:
                temp_parts = preop.split(":")
                if len(temp_parts) > 1:
                    preop_keyspace = temp_parts[1]
                preop_fn_name = temp_parts[0]

            # No Pre-Op
            # --------------
            if (
                (preop_fn_name is None)
                or not isinstance(preop_fn_name, str)
                or len(preop_fn_name) == 0
            ):
                pass
            # Pre-Op: Percentile Rank ('PR')
            # -----------------------------------
            elif preop_fn_name == "PR":
                if local_weight is None:
                    local_data = local_data.rank1d(
                        keyspace=id_keyspace,
                        ascending=True,
                        ignore_missing=True,
                        pct_rank_flag=True,
                    )
                elif np.isscalar(local_weight):
                    if isnull(local_weight):
                        local_data = local_data.clear()
                    else:
                        local_data = (local_data * np.sign(local_weight)).rank1d(
                            keyspace=id_keyspace,
                            ascending=True,
                            ignore_missing=True,
                            pct_rank_flag=True,
                        )
                        local_weight = np.absolute(
                            local_weight
                        )  # Take absolute value of weight
                else:  # <-- Nominally, local_weight is Quble here
                    local_data = (local_data * local_weight.sign()).rank1d(
                        keyspace=id_keyspace,
                        ascending=True,
                        ignore_missing=True,
                        pct_rank_flag=True,
                    )
                    local_weight = (
                        local_weight.absolute()
                    )  # Take absolute value of weight
            # Pre-Op: Sub=Percentile Rank ('SUB_PR')
            # -------------------------------------------
            elif (preop_fn_name == "SPR") or (preop_fn_name == "SUB_PR"):
                if preop_keyspace is None:
                    raise ValueError(
                        f"No preop_keyspace parameter for sub_field:{sub_field}"
                    )
                elif preop_keyspace == id_keyspace:
                    raise ValueError(
                        "preop_keyspace:{0} should NOT match id_keyspace:{1} for sub_field:{2}".format(
                            preop_keyspace, id_keyspace, sub_field
                        )
                    )

                sub_keymap = root_lib.ks_map(id_keyspace, preop_keyspace)
                if sub_keymap is None:
                    raise ValueError(
                        "Unable to locate sub_keymap:{0}->{1} for sub_field:{2}".format(
                            id_keyspace, postop_keyspace, sub_field
                        )
                    )

                if local_weight is None:
                    local_data = local_data.sub_pct_rank1d(
                        keymap=sub_keymap,
                        keyspace=id_keyspace,
                        ascending=True,
                        ignore_missing=True,
                        pct_required=0.0,
                    )
                elif np.isscalar(local_weight):
                    if isnull(local_weight):
                        local_data = local_data.clear()
                    else:
                        local_data = (
                            local_data * np.sign(local_weight)
                        ).sub_pct_rank1d(
                            keymap=sub_keymap,
                            keyspace=id_keyspace,
                            ascending=True,
                            ignore_missing=True,
                        )
                        local_weight = np.absolute(
                            local_weight
                        )  # Take absolute value of weight
                else:  # <-- Nominally, local_weight is Quble here
                    local_data = (local_data * local_weight.sign()).sub_pct_rank1d(
                        keymap=sub_keymap,
                        keyspace=id_keyspace,
                        ascending=True,
                        ignore_missing=True,
                    )
                    local_weight = (
                        local_weight.absolute()
                    )  # Take absolute value of weight

            else:
                raise ValueError(f"Unsupported preop:{preop_fn_name}")

            raise NotImplementedError("Quble logic not yet implemented")

        # ------------------------------------
        # PERFORM AGGREGATION OPERATION...
        # ------------------------------------
        _logger.debug(
            f"Performing Aggregation: {aggrop_fn_name} ({len(sub_fields)})..."
        )
        if multi_data is None:
            return Quble()
        elif (aggrop_fn_name == "AVE") or (aggrop_fn_name == "MEAN"):
            result = multi_data.mean1d(
                aggr_keyspace, ignore_missing=True, auto_squeeze=True, pct_required=0.0
            )
        elif aggrop_fn_name == "MEDIAN":
            result = multi_data.median1d(
                aggr_keyspace, ignore_missing=True, auto_squeeze=True, pct_required=0.0
            )
        elif aggrop_fn_name == "SUM":
            result = multi_data.sum1d(
                aggr_keyspace, ignore_missing=True, auto_squeeze=True, pct_required=0.0
            )
        elif (aggrop_fn_name == "WTD_AVE") or (aggrop_fn_name == "WTD_MEAN"):
            result = multi_data.wtd_mean1d(
                aggr_keyspace,
                ignore_missing=True,
                auto_squeeze=True,
                pct_required=0.0,
                view=multi_wts,
            )
        elif aggrop_fn_name == "WTD_MEDIAN":
            result = multi_data.wtd_median1d(
                aggr_keyspace,
                ignore_missing=True,
                auto_squeeze=True,
                pct_required=0.0,
                view=multi_wts,
            )
        elif aggrop_fn_name == "WTD_SUM":
            result = multi_data.wtd_sum1d(
                aggr_keyspace,
                ignore_missing=True,
                auto_squeeze=True,
                pct_required=0.0,
                view=multi_wts,
            )
        else:
            raise ValueError(f"Invalid/unsupported aggrop_fn_str:{aggrop_fn_name}")

        # ----------------------------------
        # PERFORM POSTOP (IF APPLICABLE)
        # ----------------------------------
        if postop_fn_name is not None:
            _logger.debug(f"Applying Post-Op: {postop_fn_name}")

        if postop_fn_name is None:
            pass
        elif postop_fn_name == "PR":
            result = result.rank1d(
                keyspace=id_keyspace,
                ascending=True,
                ignore_missing=True,
                pct_rank_flag=True,
                pct_required=0.0,
            )
        elif (postop_fn_name == "SPR") or (postop_fn_name == "SUB_PR"):
            if postop_keyspace is None:
                raise ValueError("postop_keyspace parameter not provided within aggrop")
            elif postop_keyspace == id_keyspace:
                raise ValueError(
                    "postop_keyspace:{0} should NOT match id_keyspace:{1}".format(
                        postop_keyspace, id_keyspace
                    )
                )
            sub_keymap = root_lib.ks_map(id_keyspace, postop_keyspace)
            if sub_keymap is None:
                raise ValueError(
                    f"Unable to locate sub_keymap:{id_keyspace}->{postop_keyspace}"
                )
            result = result.sub_rank1d(
                keymap=sub_keymap,
                keyspace=id_keyspace,
                ascending=True,
                ignore_missing=True,
                pct_rank_flag=True,
                pct_required=0.0,
            )
        else:
            raise ValueError(f"Invalid/unsupported postop_fn_name:{postop_fn_name}")

        return result

    # ============================ File Interface Routines ============================

    def file_timestamp(self, fieldref, verbose=False, suppress_recording=None):
        """
        Returns the timestamp of the file backing the passed field.
        Returns None if there is no file or if the user lacks read permissions.
        """
        lib = self
        field = fieldref
        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently
        # ================================================================
        #  PART 1: Verify File Exists & Procure File's Data Time-Stamp
        # ================================================================
        file_type = lib.get_property(
            "file_type",
            field,
            suppress_recording=suppress_recording,
            grace=True,
            default_property_value="SNOWFLAKE",  # DEFAULTING TO SNOWFLAKE
        ).upper()

        # If the filename depends on the property hash but the property dependencies have not yet been generated, then we know there is no file accessible.
        # ---------------------------------------------------------------------------
        prop_hash_in_filename = self._prop_hash_in_filename(field)
        is_derived = self.is_derived_field(field)
        if prop_hash_in_filename:
            prop_deps = self.parse_property_dependencies(
                field,
                grace=True,
                appeal=False,
                native_property_mode=NPModes.NONE,
                suppress_recording=True,
            )

            if prop_deps is None and is_derived:
                _logger.debug(
                    f"PROP DEPS IS NONE ---> RETURNING FILE_TIMESTAMP OF NONE for {self}:{field}"
                )
                return None

        if file_type == "SNOWFLAKE":
            field_category = lib.read_field_category(field)
            if field_category == "lib":
                if field in lib.field_to_id_map.keys():
                    return lib_select(lib.field_to_id_map[field])["last_updated"]
                else:
                    return None
            elif field_category == "data":
                return snowflake_get_quble_timestamp(
                    lib.get_property(
                        "filename",
                        field,
                        grace=False,
                        resolve_templates=True,
                        suppress_recording=True,
                    )
                )
            else:
                raise ValueError(
                    f"ERROR: SNOWFLAKE Libs require a valid field_category prop, but received: {field_category}, for lib|field: {lib}|{field}."
                )
        elif file_type == "LIBADDRESS":
            (srcdomain, srcfield, _) = lib.lib_address_credentials(
                field, suppress_recording=suppress_recording
            )

            if srcfield not in srcdomain.fields():
                if srcdomain.get_property("auto_register", grace=True):
                    srcdomain.register_field(srcfield)
                else:
                    return None

            return srcdomain.get(srcfield).time_stamp
        elif file_type == "QDB":
            from qubles.io.util.qdb import QDB

            filename, objname = lib.qdb_credentials(
                field, suppress_recording=suppress_recording
            )
            filename = normalize(filename)

            if filename is None:
                raise ValueError(f"(QDB) filename not provided for field: {field}")
            elif not os.path.exists(filename):
                if verbose:
                    _logger.debug(f"Inspect Warning: File Path Not Found: {filename}")
                return None
            elif objname is None:
                raise ValueError(f"(QDB) objname not provided for field: {field}")
            else:
                with closing(QDB(filename, "r")) as qdb:
                    return qdb.time_stamp(objname)
        else:
            raise ValueError(f"Unsupported file_type: {file_type} for field: {field}")

    def inspect_file(
        self,
        fieldref,
        validate=True,
        check_property_deps=False,
    ):
        """
        Search for <fieldref> in the library...

        Returns True if: Associated file exists AND:
                        a) (field is NOT DERIVED)
                    OR  b) (validate==False)
                    OR  c) (file has a non-stale time-stamp)

        Otherwise, return False
        """
        lib = self
        field = fieldref

        file_time_stamp = lib.file_timestamp(field)
        file_exists = file_time_stamp is not None

        # -----------------------------------------------
        # Handle Case where: file/table DOES NOT EXIST
        # -----------------------------------------------
        if not file_exists:
            return False

        # If the file exists, and validate is False, we are done
        if not validate:
            return True

        # If file_time_stamp is not available, assume this feature is NOT supported by underlying file storage, and return True
        # -----------------------------------------------------------------
        if self._invalid_timestamp(file_time_stamp):
            return True

        return not lib.field_fresher_than(
            field, [file_time_stamp], check_property_deps=check_property_deps
        )[0]

    # THIS IS A LEGACY METHOD, IT SHOULD NOT BE REFERENCED IN THE CODEBASE
    def qdb_credentials(
        self,
        fieldref,
        path_property_name="path",
        filename_property_name="filename",
        objname_property_name="objname",
        suppress_recording=None,
    ):
        """
        Returns the tuple of (qdb_filename,qdb_objname)
        affiliated with the specified fieldname
        """
        lib = self
        field = fieldref
        # NOTE: Make sure 'self' is NOT referenced again within this method...should use 'lib' subsequently

        if field not in lib.field_index:
            raise ValueError(f"Unsupported field: {field}")

        file_path = lib.get_property(
            path_property_name,
            field,
            grace=False,
            resolve_templates=True,
            suppress_recording=suppress_recording,
        )
        local_filename = lib.get_property(
            filename_property_name,
            field,
            grace=False,
            resolve_templates=True,
            suppress_recording=suppress_recording,
        )
        objname = lib.get_property(
            objname_property_name,
            field,
            grace=False,
            resolve_templates=True,
            suppress_recording=suppress_recording,
        )

        filename = os.path.join(normalize(file_path), local_filename)

        return (filename, objname)

    def lib_address_credentials(
        self,
        fieldref,
        domain=None,
        local_keys=None,
        path_property_name="path",
        post_minimize=True,
        suppress_recording=None,
    ):
        """
        Given a fieldref, requests the specified path_property_name and uses these to construct a (domain, field, sub_keys)
        When post_minimize=True, the resulting domain will be the minimum possible domain
        """
        from qubles.io.base.rootlib import RootLib

        if domain is None:
            domain = self.get_property(
                path_property_name,
                fieldref,
                grace=True,
                suppress_recording=suppress_recording,
            )

        if local_keys is None:
            local_keys = self.get_property(
                "underlying_filename",
                fieldref,
                grace=True,
                suppress_recording=suppress_recording,
            )
            if local_keys is None:
                raise ValueError(
                    "LibAddress credentials: underlying_filename property (absent or None)"
                )
            elif isinstance(local_keys, LibAddress):
                if domain is not None:
                    raise TypeError(
                        f"domain and/or path (name:{path_property_name}) should be None when providing a LibAddress as a filename_(name:underlying_filename)"
                    )
                else:
                    domain = local_keys.base_domain
                    local_keys = list(local_keys)  # .toString()
            elif isinstance(local_keys, str):
                local_keys = local_keys.split(RootLib().get_control("index_delimiter"))

        if isinstance(domain, DataLib):
            pass
        elif domain is None:
            domain = RootLib()
        elif not isinstance(domain, str):
            raise TypeError("Invalid domain: DataLib or string (proxy) expected")
        elif domain[0:4].lower() == "root":
            domain = RootLib()
        elif (domain[0:5].lower() == "local") or (domain[0:4].lower() == "self"):
            domain = self
        else:
            raise ValueError(
                "LibAddress credentials: Invalid {0} (path / domain) property:{1}".format(
                    path_property_name, domain
                )
            )

        # Minimize (if applicable)
        # --------------------------
        if post_minimize:
            linked_address = LibAddress(local_keys, domain)
            domain, local_field, sub_keys = linked_address.min_subparts
        elif not isinstance(local_keys, list):
            raise TypeError("local_keys is expected to be a list")
        elif len(local_keys) == 0:
            local_field = None
            sub_keys = None
        elif len(local_keys) == 1:
            local_field = local_keys[0]
            sub_keys = None
        else:
            local_field = local_keys[0]
            sub_keys = local_keys[1:]

        return (domain, local_field, sub_keys)

    # =================================================================================
    #                                 DataLib States:
    # ---------------------------------------------------------------------------------
    #    1) Storage: Info on where/how is the library stored
    #    2) Structure: Info on the architecture/structure of the library elements
    #    3) Typical: Info on the library's elemental type make-up
    # =================================================================================

    # ---------------------------------------------------------------------------------
    #                             DataLib: Storage States...
    # ---------------------------------------------------------------------------------
    @property
    def is_virtual(self):
        """
        Indicates whether DataLib has stored all items virtually (in resident dictionary)
        """
        for field in self.field_index:
            property_dep_hash = self.property_dep_hash_current(field, validate=False)
            residence_key = (field, property_dep_hash)
            if not self._exists_in_residence(residence_key):
                return False
        return True

    def ensure_virtual(self):
        self.virtualize(inplace=True)

    def virtualize(self, inplace=False):
        """
        Stores all items in the Library virtually (providing allowable store_mode)
        NOTE: implicitly 'inplace' (i.e., Operates on self. Does NOT return a copy)

        If inplace = True, self will be modified accordingly
        If inplace = False, self will be unchanged, and a fully-virtual copy will be returned
        """
        if inplace:
            for field in self.field_index:
                property_dep_hash = self.property_dep_hash_current(field)
                residence_key = (field, property_dep_hash)
                if not self._exists_in_residence(residence_key):
                    if not self.virtual_storage(field):
                        raise BuilderError(
                            "Store mode inconsistent with virtual storage"
                        )
                    else:
                        self.get(
                            field
                        )  # <-- the self.get() method should result in a population of the virtual dictionary
            return self
        else:
            return self.copy(virtual_only=True)

    @property
    def has_lib_fields(self):
        for field in self.field_index:
            if self.get_property("field_category", field) == "lib":
                return True
        return False

    @property
    def has_shared_data(self):
        for field in self.field_index:
            if self.get_property("category1", field) == "Shared Data":
                return True
        return False

    @property
    def has_data_fields(self):
        for field in self.field_index:
            if self.get_property("field_category", field) == "data":
                return True
        return False

    @property
    def data_only(self):
        if not self.has_data_fields:
            return False  # 'None_only' or 'lib_only'
        elif self.has_lib_fields:
            return False  # 'mixed'
        else:
            return True

    @property
    def lib_only(self):
        if not self.has_lib_fields:
            return False  # 'None_only' or 'data_only'
        elif self.has_data_fields:
            return False  # 'mixed'
        else:
            return True

    # ---------------------------------------------------------------------------------
    #                            DataLib Typical States...
    # ---------------------------------------------------------------------------------

    @property
    def dtype(self):
        """Returns the common dtype (if applicable) or None if dtypes are inconsistent"""
        common_dtype = None

        # Ensure self is a non-empty, fully virtual DataLib...
        # -----------------------------------------------------
        if self.is_empty:
            return common_dtype

        # Loop thorugh the fields...
        # --------------------------------
        for field in self.field_index:
            # Procure local_dtype...
            # --------------------------
            local_dtype = None
            local_value = self.get(field)
            if local_value is None:
                pass
            elif hasattr(
                local_value, "dtype"
            ):  # <-- should be true for np.scalar, Quble
                local_dtype = local_value.dtype
            else:
                raise AttributeError(f"self[{field}] has no attribute: dtype")

            # Evaluate primal_dtype...
            # -------------------------
            if common_dtype is None:
                common_dtype = local_dtype
            elif common_dtype != local_dtype:
                return None  # <-- Mutiple/inconsistent dtypes exist

        return common_dtype

    # ======================== DataLib Properties & Transforms ========================

    def conformal_fields(self):
        return self.fields()

    @property
    def is_conformal(self):
        if self.is_empty:
            return False

        # ------------------------------------------------
        # Loop through library objects and validate that:
        #  1) They are all data objects (Qubles)
        #  2) All data fields have a common hyper_index
        # ------------------------------------------------
        lhi = None  # <-- lower hyper index (lhi)
        for fld1 in self.conformal_fields():
            val1 = self.get(fld1)

            if val1 is None:
                return True
            elif is_quble(val1):
                # Conformal means coindexed (same table)
                if lhi is None:
                    lhi = val1  # <-- we could take val1.index here, but more overhead for no functional benefit
                elif not val1.are_coindexed(lhi, grace=True):
                    return False
            else:
                return False

        return True

    def conform(self):
        return self.to_report()

    def to_report(self):
        """
        Conforms a DataLib by generating a Report instance
        """
        from qubles.io.base.report import Report

        if isinstance(self, Report):
            return self.copy()

        fields = [fld for fld in self.fields()]
        default_properties = {
            "underlying_file_type": "LIBADDRESS",
            "underlying_path": self,
        }
        fldspec_properties = {"underlying_filename": {}}
        fieldspace = self.get_property("fieldspace", grace=False)
        for fld in fields:
            fldspec_properties["underlying_filename"][fld] = fld

        lib = Report(
            fields=fields,
            default_properties=default_properties,
            fldspec_properties=fldspec_properties,
            fieldspace=fieldspace,
            access_mode="build",
            store_mode="virtual",
        )

        return lib

    def to_quble(
        self,
        unpivot=False,
        keys_join_op="union",
        compress=True,
        treat_false_as_null=False,
        variate_mode="uni",
        allow_recursion=False,
        join_chunck_size=10,
        unpivot_keyspace="<fieldspace>",
    ):
        """
        Converts a DataLib to a (multi-variate or uni-variate) Quble

        If this library is a Report instance, simply access the first display field

        If not a report, joins the primary valuespaces of the Quble fields using the specified keys_join_op then
        compress the multi/uni-variate Quble as directed

        :param unpivot: Flag to create a uni-variate Quble result by unpivoting the multiple valuespaces resulting from the fields of the library.
                        The unpivot operation will introduce a unpivot_keyspace. If unpivot_keyspace=="<fieldspace>", the new keyspace will be
                        assigned according to library's fieldspace. The resultant new_valuespace=DEFAULT_VALUESPACE.
        :type unpivot: boolean (False*/True)

        :param keys_join_op: keys join operator [See: :meth:`~qubles.core.quble.Quble.join`]
        :type keys_join_op: str

        :param compress: Controls compression
        :type compress: boolean or str (True*, False, 'all', 'any' or None)
            ==> compress=False: no compression
            ==> compress=True and unpivot=True: compress the uni-variate Quble
            ==> compress=True and unpivot=False: compress multi-variate Quble w/summarize='any' to keep rows with 'any' non-nulls
            ==> compress='any' and unpivot=False: compress multi-variate Quble w/summarize='any' to keep rows with 'any' non-nulls
            ==> compress='all' and unpivot=False: compress multi-variate Quble w/summarize='all' to only keep rows with 'all' non-nulls

        :param treat_false_as_null: For boolean values, controls how to treat compression/removal of False values (only applies to unpivot=False case)
        :type treat_false_as_null: boolean (False*/True)

        :param variate_mode: controls valuespace pairings -> None or 'uni': only integrate primary valuespace from each field
        'multi' or 'mixed': integrate all valuespaces from each field
                ==> For multi-variate Qubles, add valuespace suffix to fieldname
                ==> For example: field='Pricing' src valespaces: ['Close', 'Volume']
                ==> tgt valespaces: ['Pricing:Close', 'Pricing:Volume']
        :type variate_mode: str

        :param allow_recursion: Flag to allow recursive calls for sub-library fields
        :type allow_recursion: boolean (False*/True)

        :param join_chunck_size: Maximum table joins per join chunck
        :type join_chunck_size: int (>0) or None

        :param unpivot_keyspace: new keyspace created via the unpivot operation (if unpivot==True)
        :type unpivot_keyspace: str
        """
        from qubles.io.base.report import Report
        from qubles.core.quble import Quble

        # Validate variate_mode
        if variate_mode not in (None, "uni", "mixed", "multi"):
            raise ValueError(
                f"Invalid variate_mode:{variate_mode}...valid options: 'uni','mixed','multi'"
            )
        elif variate_mode is None:
            variate_mode = "uni"

        # Validate the join_chunck_size
        if join_chunck_size is None:
            # No join chuncking
            pass
        elif not isinstance(join_chunck_size, int) or join_chunck_size <= 1:
            raise ValueError(
                f"Invalid join_chunck_size:{join_chunck_size}...integer >= 2 expected"
            )

        # First, convert DataLib to a Report if needed
        if isinstance(self, Report):
            tgt_quble = self.get_first_display_field()
        elif len(self.field_index) == 0:
            tgt_quble = Quble.undefined_instance()
        else:
            valuespaces_join_op = {}
            qubles_to_be_joined = []
            base_tgt_valuespaces = []
            field_ctr = 0  # <-- NOTE: Some fields may not participate in join
            intrachunk_quble_ctr = 0
            join_chuncks = []
            for fieldname in self.field_index:
                # NOTE: All underlying_vals should be variate here
                val = self[fieldname]
                if val is None:
                    # Could throw an Exception here
                    continue
                elif isinstance(val, DataLib) and allow_recursion:
                    val = val.to_quble(
                        unpivot=unpivot,
                        keys_join_op=keys_join_op,
                        compress=compress,
                        treat_false_as_null=treat_false_as_null,
                        variate_mode=variate_mode,
                        allow_recursion=allow_recursion,
                        join_chunck_size=join_chunck_size,
                    )
                elif not isinstance(val, Quble):
                    raise TypeError(
                        f"Non-Quble for field:{fieldname} w/allow_recursion:{allow_recursion}"
                    )
                elif val.is_undefined:
                    continue
                elif val.is_nonvariate:
                    val = val.index_to_bool(valuespace=fieldname)

                # Build valuespaces_join_op[tgt_vs] = (field_ctr, src_valuespace)
                if val.is_multivariate and variate_mode in ("multi", "mixed"):
                    for src_vs in val.valuespaces:
                        tgt_vs = fieldname + ":" + src_vs
                        valuespaces_join_op[tgt_vs] = (intrachunk_quble_ctr, src_vs)
                else:
                    valuespaces_join_op[fieldname] = (
                        intrachunk_quble_ctr,
                        val.valuespace,
                    )

                qubles_to_be_joined.append(val)
                field_ctr += 1
                intrachunk_quble_ctr += 1
                if (join_chunck_size is not None) and (
                    (field_ctr % join_chunck_size) == 0
                ):
                    # Log qubles_to_be_joined and valuespaces_join_op
                    join_chuncks.append([qubles_to_be_joined, valuespaces_join_op])

                    # Then start a new join chunck and seed with the prior valuespaces
                    for tgt_vs in valuespaces_join_op:
                        base_tgt_valuespaces.append(tgt_vs)

                    # For the next chunck to be compiled, the previous stage will serve as the "base" element[0]
                    # For this base, the source valuespaces and tgt valuspaces are same
                    valuespaces_join_op = {
                        0,
                        [vs for vs in base_tgt_valuespaces],
                    }  # <-- making a deep copy here
                    qubles_to_be_joined = [
                        "base"
                    ]  # <-- placeholder/proxy for running joined table
                    intrachunk_quble_ctr = (
                        1  # <-- The previous stage will serve as the "base" element[0]
                    )

            # Log the residual chunck (if present)
            if intrachunk_quble_ctr == 0:
                # Nothing in this join chunck to process
                pass
            elif (
                intrachunk_quble_ctr == 1
                and isinstance(qubles_to_be_joined[0], str)
                and (qubles_to_be_joined[0] == "base")
            ):
                # Only the base here in this join chunck, so nothing extra to be added on top of the base
                pass
            elif (
                len(join_chuncks) == 0
                and isinstance(qubles_to_be_joined[0], str)
                and (qubles_to_be_joined[0] == "base")
            ):
                raise ValueError(
                    f"Internal inconsistency...len(join_chuncks):{len(join_chuncks)} yet qubles_to_be_joined[0]:{qubles_to_be_joined[0]}"
                )
            else:
                join_chuncks.append([qubles_to_be_joined, valuespaces_join_op])
                # For Posterity, reset the intr-chunck valuespaces_join_op and qubles_to_be_joined
                valuespaces_join_op = {}
                qubles_to_be_joined = []

            # Initialize the tgt_quble as an undefined Quble
            tgt_quble = Quble.undefined_instance()

            # Apply the joins in apportioned chuncks
            for chunck_no in range(len(join_chuncks)):
                qubles_to_be_joined = join_chuncks[chunck_no][0]
                valuespaces_join_op = join_chuncks[chunck_no][1]
                if len(qubles_to_be_joined) == 0:
                    # No qubles to be joined from this chunck
                    pass
                elif not tgt_quble.is_undefined:
                    if isinstance(qubles_to_be_joined[0], str) and (
                        qubles_to_be_joined[0] == "base"
                    ):
                        # Some (one or more) qubles to be joined from this chunck to an existing, defined tgt_quble
                        tgt_quble = tgt_quble.join(
                            other=qubles_to_be_joined[1:],
                            keys_join_op=keys_join_op,
                            keyspaces_join_op="union",
                            valuespaces_join_op=valuespaces_join_op,
                        )
                    else:
                        raise ValueError(
                            f"Internal inconsistency...tgt_quble is not undefined yet qubles_to_be_joined[0]:{qubles_to_be_joined[0]}"
                        )
                elif len(qubles_to_be_joined) == 1:
                    if isinstance(qubles_to_be_joined[0], str) and (
                        qubles_to_be_joined[0] == "base"
                    ):
                        raise ValueError(
                            f"Internal inconsistency...tgt_quble is undefined yet qubles_to_be_joined[0]:{qubles_to_be_joined[0]}"
                        )
                    # A single quble to be joined from this chunck to a non-existing, undefined tgt_quble
                    # Here, we merely assign tgt_quble to the sole quble to be joined
                    tgt_quble = qubles_to_be_joined[0]
                else:
                    if (
                        isinstance(qubles_to_be_joined[0], str)
                        and qubles_to_be_joined[0] == "base"
                    ):
                        raise ValueError(
                            f"Internal inconsistency...tgt_quble is undefined yet qubles_to_be_joined[0]:{qubles_to_be_joined[0]}"
                        )
                    # Mutiple qubles to be joined from this chunck to a non-existing, undefined tgt_quble
                    tgt_quble = qubles_to_be_joined[0].join(
                        other=qubles_to_be_joined[1:],
                        keys_join_op=keys_join_op,
                        keyspaces_join_op="union",
                        valuespaces_join_op=valuespaces_join_op,
                    )

        # Pivot and/or compress the resultant Quble
        if tgt_quble is None:
            pass
        elif not isinstance(tgt_quble, Quble):
            raise TypeError(
                f"Invalid type(tgt_quble):{type(tgt_quble)}...Quble expected"
            )
        elif tgt_quble.is_undefined:
            # Cannot compress or unpivot an undefined Quble
            pass
        elif tgt_quble.is_nonvariate:
            # Cannot compress or unpivot a non-variate Quble
            pass
        elif unpivot:
            # In this case, compress should ideally be a boolean
            if unpivot_keyspace == "<fieldspace>":
                unpivot_keyspace = self.get_property("fieldspace", grace=False)

            if not isinstance(unpivot_keyspace, str):
                raise TypeError(
                    f"Invalid unpivot_keyspace:{unpivot_keyspace} string required for unpivot={unpivot}"
                )
            tgt_quble = tgt_quble.unpivot(
                new_keyspace=unpivot_keyspace, compress=compress
            )
        elif compress is None:
            pass
        elif isinstance(compress, str):
            tgt_quble.compress(
                treat_false_as_null=treat_false_as_null,
                summarize=compress,
                inplace=True,
            )
        elif compress not in (True, False):
            raise TypeError(f"Invalid compress:{compress}...str or bool required")
        elif compress:
            tgt_quble.compress(
                treat_false_as_null=treat_false_as_null, summarize="any", inplace=True
            )

        return tgt_quble

    def __repr__(self):
        return f"""<{self.__class__.__name__} object{"" if not self.address else f" with address '{'|'.join(self.address)}'"} at {hex(id(self))}>"""

    def __str__(self):
        return f"<{self.__class__.__name__} @address: '{'|'.join(self.address if self.address else [])}'>"

    def multi_eval_od(self, fields=None, request_type="field_value"):
        """
        Returns a dictionary of field values: Motivated by need for efficient multi-field evalution for Report class
        """
        result = {}
        if self.is_empty:
            return result
        elif fields is None:
            fields = self.field_index

        for fld1 in fields:
            result[fld1] = self.get(fld1, request_type=request_type)

        return result

    def reconcile(
        self,
        fields=None,
        keyspaces_to_resolve=None,
        key_grace=True,
        auto_link=True,
        suppress_recording=False,
        keyspaces_join_op="union",
        keys_join_op="intersection",
        valuespaces_join_op="all",
    ):
        """
        Builds a dictionary of "reconciled" (sub)field content

        Initializes dictionary with associated self[field]
        Next, resolves (remaps) the specified keyspaces (if preent in any of these sub-fields)
        Then looks across fields in dictionary to find/link/massage all keyspaces of the objects provided
        ------------------------------------------------------------------------------------

            :param fields: list of fields to be "reconciled"

            :param keyspaces_to_resolve: Will "resolve" these keyspaces (if necessary) Then performs keyspace auto_links if/when appropriate
                                NOTE: ALL KEYSPACES WILL BE RECONCILED, not just the keyspaces to be resolved/remapped

                [For example, security_keyspace may be templatized as '<SECMSTR>',
                    in which case this RefLib will be referenced to establish a specific security_keyspace (when present)]

            :param key_grace (boolean): True: Robust to fields not being originally present (will seed with value = None)
                                False: Will throw an error for any unregistered fields

            :param reindex_flag: Imposition of Spanning hyper index

            :param auto_link (boolean): True: Use linking (via RefLibs) in keyspace resolution exercise

            :param suppress_recording (boolean): Flag to control suppression of subsequent get requests within reconciliation


        In the case of non-time keyspaces (e.g., security_keyspaces): Imposes the INTERSECTION of the asset keys within security_keyspace (dates_keyspace)

        In the case of time keyspaces  (e.g., dates_keyspaces), will first convert frequencies accordingly and then impose INTERSECTION of dates
            (including start_date and/or end_date information) across relevant objects

        **OUTPUTS:**

            Returns the following results as tuple: (reconciliation_dict, Span, resolved_keyspaces)

                reconciliation_dict: dictionary w/keys=fields (input arg) and values provide modified field content

                Span: Spanning index (Quble) for reconciled fields such that Span[each unique keyspace] = common target keys

                resolved_keyspaces: list of resolved keyspaces (order and length corresponds to keyspaces_to_resolve input)
        """
        from qubles.io.base.rootlib import RootLib

        root_lib = RootLib()

        if keyspaces_to_resolve is None:
            keyspaces_to_resolve = []

        # ==================================================================
        # STEP #1: Seed Reconciliation Dictionary (reconciliation_dict)
        # ==================================================================
        reconciliation_dict = {}
        if fields is None:
            fields = self.field_index

        if len(fields) == 0:
            return reconciliation_dict

        for field in fields:
            if field not in self.field_index:
                if key_grace:
                    reconciliation_dict[field] = None
                else:
                    raise BuilderError(f"Absent field: {field}")
            else:
                temp1 = self.get(field, suppress_recording=suppress_recording)
                if isinstance(temp1, DataLib):
                    reconciliation_dict[field] = temp1.to_quble(auto_conform=True)
                else:
                    reconciliation_dict[field] = temp1

        # ======================================================
        # STEP #1: Resolve keyspaces & remap when necessary
        # ======================================================
        resolved_keyspaces = []
        for ks_no, ks in enumerate(keyspaces_to_resolve):
            resolved_keyspaces.append(ks)  # <-- Initialize
            if ks is None:
                continue
            elif not isinstance(ks, str):
                raise TypeError(
                    "Invalid arg: string expected for keyspaces_to_resolve[{0}]: {1}".format(
                        ks_no, ks
                    )
                )
            elif (
                (len(ks) > 2) and (ks[0] == "<") and (ks[-1] == ">")
            ):  # <-- Templatized ks provided
                verified = False
            else:
                verified = True  # <-- Do not need to (re)set resolved_keyspaces[ks_no] here (already set within initialization above)

            # Try to verify
            for field in list(reconciliation_dict.keys()):
                if (reconciliation_dict[field] is None) or reconciliation_dict[
                    field
                ].is_empty:
                    continue

                if not verified:
                    try:
                        resolved_keyspaces[ks_no] = reconciliation_dict[
                            field
                        ].hyper_index.keyspace(
                            ks, auto_link=True, grace=False
                        )  # <-- Allows for link conversion. Will throw an error if not supported
                        verified = True
                    except Exception as e:
                        _logger.exception(e)
                        continue

                # If verified, try to apply remap [NOTE: Must be verified=True to get to this line]
                # ----------------------------------------------------
                if (
                    auto_link
                    and (resolved_keyspaces[ks_no] is not None)
                    and (
                        resolved_keyspaces[ks_no]
                        not in reconciliation_dict[field].keyspaces
                    )
                ):  # <-- Seek remapping when appropriate
                    # Will not error if dates_keyspace not mappable (i.e., quble does not support dates)
                    reconciliation_dict[field] = root_lib.remap(
                        reconciliation_dict[field],
                        src_keyspaces=reconciliation_dict[field].keyspaces,
                        tgt_keyspaces=resolved_keyspaces[ks_no],
                        deep_copy=False,
                    )

        # =================================================================================
        # STEP #3: Establish reconciliation frequencies & keys (intesections)
        # ---------------------------------------------------------------------------------
        # NOTE: We expanded this logic to include union of all keyspaces (not just keyspaces_to_be_reconciles: security_keyspaces and dates_keyspaces)
        # =================================================================================
        # Can probably merely perform a join on all Qubles in dictionary
        if len(reconciliation_dict) <= 1:
            # Need atleast two items for non-trivial reconciliation
            pass
        else:
            unjoined_qubles = list(reconciliation_dict.values())

            # Set freq to that of first time_keyspace encountered
            freq = RootLib().get_control("freq")  # <-- Seed w/default
            for unjoined_quble1 in unjoined_qubles:
                freq1 = unjoined_quble1.first_time_keyspace(grace=True)
                if freq1 is not None:
                    freq = freq1
                    break
            # join the Qubles accordingly
            joined_qubles = unjoined_qubles[0].join(
                unjoined_qubles[1:],
                keys_join_op=keys_join_op,
                keyspaces_join_op=keyspaces_join_op,
                valuespaces_join_op=valuespaces_join_op,
                freq=freq,
            )
            # Replace each (unjoined_quble) in reconciliation_dict values the with associated joined_quble
            for field_no, field1 in enumerate(reconciliation_dict.keys()):
                reconciliation_dict[field1] = joined_qubles[field_no]
            # Since joined_qubles are coindexed, we can access the index from any arbitrarily
            if (
                len(joined_qubles) > 0
                and joined_qubles[0] is not None
                and not joined_qubles[0].is_scalar
            ):
                span = joined_qubles[0].index
            else:
                span = None

        return (reconciliation_dict, span, resolved_keyspaces)

    def export_lib(self):
        """
        Exports library-specific properties and field information into a pandas DataFrame, and prepares this DataFrame for CSV export.
        This method compiles data associated with the library itself and its parent, merges this data, and handles serialization errors.
        It also populates additional property fields based on library's content, handles missing values, and removes specific unwanted columns.

        Raises:
            Exception: If there is an error in deserializing property data or if the library data is missing.

        Returns:
            tuple: A tuple containing the pandas DataFrame of field-specific properties and the intended file name for the CSV.
        """
        lib = self
        default_properties = deepcopy(lib.properties.default_properties)
        field_specific_properties = deepcopy(lib.properties.fldspec_properties)
        parent_props = properties_field_select(
            get_parent_id(lib.id, lib.address), lib.id
        )[6]

        for info_type, info_value in default_properties.items():
            if info_type not in field_specific_properties:
                field_specific_properties[info_type] = {}
            field_specific_properties[info_type]["DEFAULT"] = info_value

        for info_type, info_value in parent_props.items():
            if info_type not in field_specific_properties:
                field_specific_properties[info_type] = {}
            field_specific_properties[info_type]["PARENT"] = info_value

        field_specific_properties_df = pd.DataFrame.from_dict(
            field_specific_properties, orient="index"
        ).transpose()
        field_specific_properties_df.reset_index(inplace=True)
        field_specific_properties_df.rename(
            columns={"index": "field_name"}, inplace=True
        )

        if "definition" in field_specific_properties_df.columns:
            dfn_df = field_specific_properties_df[["field_name", "definition"]].dropna()
            field_specific_properties_df["definition_data"] = np.nan
            for itr in dfn_df.index:
                field_specific_properties_df["definition_data"][itr] = (
                    lib.get_definition(dfn_df["field_name"][itr])[0]
                )

        field_specific_properties_df = field_specific_properties_df.fillna(pd.NA)
        remove_column_list = list(
            set(self.PROCUREMENT_PROPERTY_RECORDINGS_LIST)
            & set(field_specific_properties_df.columns)
        )
        field_specific_properties_df.drop(
            remove_column_list, axis="columns", inplace=True
        )

        return field_specific_properties_df, str(lib.name) + ".csv"

    def import_lib(self, data, address):
        from viewer.api.utils.app.snowflake_env import put_node

        definition_flag = False
        csv_reader = csv.DictReader(data)
        df_data = [row for row in csv_reader]
        read_df = pd.DataFrame(df_data)
        remove_column_list = list(
            set(self.PROCUREMENT_PROPERTY_RECORDINGS_LIST) & set(read_df.columns)
        )
        read_df.drop(remove_column_list, axis="columns", inplace=True)
        read_df = read_df.applymap(lambda x: x if len(x) else np.nan)
        if "definition_data" in read_df.columns:
            dfn_df = read_df[["field_name", "definition_data"]].dropna()
            read_df.drop(
                ["definition_data", "definition"], axis="columns", inplace=True
            )
            definition_flag = True

        if "field_index_order" in read_df.columns:
            read_df["field_index_order"] = read_df["field_index_order"].apply(
                lambda x: ast.literal_eval(x) if pd.notna(x) else np.nan
            )
        default_df = read_df[read_df.field_name == "DEFAULT"]
        default_df = default_df.set_index("field_name")
        default_df = default_df.dropna(axis=1)

        if not default_df.shape[0]:
            raise DataLibCSVError(
                "CSV File has No 'DEFAULT' row within column: field_name"
            )
        default_properties = {}
        for col in default_df.columns:
            if len(default_df[col].iloc[0]):
                default_properties[col] = default_df[col].iloc[0]
        parent_df = read_df[read_df.field_name == "PARENT"]
        parent_df = parent_df.dropna(axis=1)

        if not parent_df.shape[0]:
            raise DataLibCSVError(
                "CSV File has No 'PARENT' row within column: field_name"
            )
        parent_properties = parent_df.applymap(
            lambda x: x if pd.notna(x) else None
        ).to_dict("records")[0]

        field_specific_properties_df = read_df[
            (read_df.field_name != "DEFAULT") & (read_df.field_name != "PARENT")
        ]
        fields = field_specific_properties_df["field_name"].to_list()
        field_specific_properties_df = field_specific_properties_df.set_index(
            "field_name"
        )
        field_specific_properties_df.replace("", pd.NA, inplace=True)
        field_specific_properties = {}

        for col in field_specific_properties_df.columns:
            temp_field_specific_properties_df = field_specific_properties_df.dropna(
                subset=[col]
            )
            field_specific_properties[col] = temp_field_specific_properties_df[
                col
            ].to_dict()

        libtype = parent_properties["field_type"]
        kwargs = {}
        if libtype in ("Portfolio", "Backtest", "Forecast"):
            kwargs["reconfigure_flag"] = False
        elif libtype == "RootLib":
            kwargs["force_create_singleton"] = True

        # See the DataLib constructor for an explanation of this setting
        default_properties["validate_properties"] = False
        cls = libtypes.get(libtype)
        new_lib = cls(fields, default_properties, field_specific_properties, **kwargs)
        if definition_flag:
            for itr in dfn_df.index:
                def_id = new_lib.set_definition(
                    code=dfn_df["definition_data"][itr],
                    fieldref=dfn_df["field_name"][itr],
                )
                if def_id is not None:
                    new_lib.set_property(
                        "definition", def_id, dfn_df["field_name"][itr]
                    )
        new_lib.name = address[-1]
        new_lib.address = LibAddress(address)
        put_node(address, new_lib)

        return new_lib

    def import_lib_shared_data(
        self, root_lib, data: StringIO, address: LibAddress, properties=None
    ):
        """
        Utility method to insert libraries into the shared data environment.

        :param root_lib: The rootlib to add the library to
        :param data: a StringIO object containing the library to insert in a CSV format
        :param address: the address of the library to insert
        :param properties: A dictionary containing any additional property values to set on the field such as
        category1, category2, etc.
        :return:
        """
        from viewer.api.utils.app.snowflake_env import put_node

        if properties is None:
            properties = {"category1": "Shared Data"}
        # NOTE: We have to load the root_lib early on for a couple of different reasons:
        # 1. When constructing the LibAddress further down in code, we must have the correct rootlib as a base
        # 2. We must immediately set the category1 and category2 properties before the rootlib will be able to be
        #    loaded again.
        definition_flag = False
        csv_reader = csv.DictReader(data)
        df_data = [row for row in csv_reader]
        read_df = pd.DataFrame(df_data)
        remove_column_list = list(
            set(self.PROCUREMENT_PROPERTY_RECORDINGS_LIST) & set(read_df.columns)
        )
        read_df.drop(remove_column_list, axis="columns", inplace=True)
        read_df = read_df.applymap(lambda x: x if len(x) else np.nan)
        if "definition_data" in read_df.columns:
            dfn_df = read_df[["field_name", "definition_data"]].dropna()
            read_df.drop(
                ["definition_data", "definition"], axis="columns", inplace=True
            )
            definition_flag = True

        if "field_index_order" in read_df.columns:
            read_df["field_index_order"] = read_df["field_index_order"].apply(
                lambda x: ast.literal_eval(x) if pd.notna(x) else np.nan
            )
        default_df = read_df[read_df.field_name == "DEFAULT"]
        default_df = default_df.set_index("field_name")
        default_df = default_df.dropna(axis=1)

        if not default_df.shape[0]:
            raise Exception(" CSV File have No DEFAULT Property ")
        default_properties = {}
        for col in default_df.columns:
            if len(default_df[col].iloc[0]):
                default_properties[col] = default_df[col].iloc[0]
        parent_df = read_df[read_df.field_name == "PARENT"]
        parent_df = parent_df.dropna(axis=1)

        if not parent_df.shape[0]:
            raise Exception(" CSV File have No PARENT Property ")
        parent_properties = parent_df.applymap(
            lambda x: x if pd.notna(x) else None
        ).to_dict("records")[0]

        field_specific_properties_df = read_df[
            (read_df.field_name != "DEFAULT") & (read_df.field_name != "PARENT")
        ]
        fields = field_specific_properties_df["field_name"].to_list()
        field_specific_properties_df = field_specific_properties_df.set_index(
            "field_name"
        )
        field_specific_properties_df.replace("", pd.NA, inplace=True)
        field_specific_properties = {}

        for col in field_specific_properties_df.columns:
            temp_field_specific_properties_df = field_specific_properties_df.dropna(
                subset=[col]
            )
            field_specific_properties[col] = temp_field_specific_properties_df[
                col
            ].to_dict()

        libtype = parent_properties["field_type"]
        kwargs = {}
        if libtype in ("Portfolio", "Backtest", "Forecast"):
            kwargs["reconfigure_flag"] = False
        elif libtype == "RootLib":
            kwargs["force_create_singleton"] = True

        # See the DataLib constructor for an explanation of this setting
        default_properties["validate_properties"] = False
        cls = libtypes.get(libtype)
        new_lib = cls(fields, default_properties, field_specific_properties, **kwargs)
        if definition_flag:
            for itr in dfn_df.index:
                def_id = new_lib.set_definition(
                    code=dfn_df["definition_data"][itr],
                    fieldref=dfn_df["field_name"][itr],
                )
                if def_id is not None:
                    new_lib.set_property(
                        "definition", def_id, dfn_df["field_name"][itr]
                    )
        new_lib.address = address
        new_lib.name = address.to_string()
        put_node(address, new_lib)
        print(new_lib.address.parts())
        for name, value in properties.items():
            root_lib.set_property(name, value, fieldref=address.to_string())
        # Write the updated property information
        root_lib.commit()
        # TODO: Make the nav tree show the new library without restarting django

        return new_lib


def _process_defaults_sheet(
    defaults_book=None,
    defaults_sheet=None,
    defaults_header=None,
    ignore_empty_strings=True,
    verbose=False,
):
    # Returns a dictionary of additional defaults as loaded from the specified default book & sheet
    add_defaults = {}
    if defaults_book is not None:
        # Step 1: Qualify book & sheet
        # --------------------------------
        if not os.path.isfile(defaults_book):
            raise IOError(f"File does not exist: {defaults_book}")
        else:
            wb = load_workbook(defaults_book, data_only=True)

        if defaults_sheet is None:
            raise ValueError("No defaults sheet specified")
        else:
            sheet = wb.get_sheet_by_name(defaults_sheet)

        # Make sure sheet was found within the book...
        if sheet is None:  # or type(sheet) == type(None):
            raise ValueError(f"Specified sheet absent: {defaults_sheet}")

        # Step 2: process columns in the header row of the defaults sheet (ignoring first column)
        # ---------------------------------------------------------------------------------------------
        if defaults_header is None:
            raise ValueError("No defaults_header specified")
        else:
            header_col = None
            col_cntr = 1  # <-- Want to INTENTIONALLY skip the first column (is it is reserved for/assumed to be the property names)
            while True:
                col_cntr += 1  # <-- Should INTENTIONALLY start with 2 the first cycle (so the first/fieldname column is NOT treated as a possible property column)
                col = get_column_letter(col_cntr)
                header_name = sheet[f"{col}1"].value
                if header_name is None:
                    break

                if header_name == defaults_header:
                    header_col = col
                    break

            if header_col is None:
                raise ValueError(
                    f"defaults_header: {defaults_header} not present in columns 2+"
                )

        # Step 3: process Each Row of the Defaults Sheet (skipping hear row)
        # ---------------------------------------------------------------------
        row_cntr = 2  # <-- Assume first row is header row (ignore first row)
        while True:
            # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
            property_name = sheet[
                "A%i" % row_cntr
            ].value  # <-- Assumes that property name is provided in first column

            # Check whether we have reached the last property_name (last row) in the sheet
            # [Note, you can also use syntax such as: sheet.cell(row = 4, column = 2).value]
            # -----------------------------------------------------------------------------------
            if property_name is None:  # or type(property_name) == type(None):
                break

            # Read/interpret the default_property_value
            # ---------------------------------------------------
            # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
            default_property_value = sheet[
                col + "%i" % row_cntr
            ].value  # <-- Assumes that property value is provided in second column

            # Check whether a cell value was provided...
            if (default_property_value is not None) and (
                not ignore_empty_strings or (default_property_value != "")
            ):
                # Convert property_value to numeric type (if applicable)
                if isinstance(default_property_value, str):
                    try:
                        if search("\.", default_property_value):
                            raise ValueError
                        default_property_value = int(default_property_value)
                    except ValueError:
                        try:
                            default_property_value = float(default_property_value)
                        except ValueError:
                            pass

                if isinstance(default_property_value, int) or isinstance(
                    default_property_value, float
                ):
                    pass
                elif default_property_value == "None":
                    default_property_value = None
                elif (default_property_value == "True") or (
                    default_property_value == "TRUE"
                ):
                    default_property_value = True
                elif (default_property_value == "False") or (
                    default_property_value == "FALSE"
                ):
                    default_property_value = False
                elif (
                    isinstance(default_property_value, str)
                    and len(default_property_value.split("/")) > 1
                ):
                    try:
                        default_property_value = datetime.strptime(
                            default_property_value, "%m/%d/%Y:%H:%M:%S"
                        )
                    except Exception as e:
                        _logger.critical(e)
                        try:
                            default_property_value = datetime.strptime(
                                default_property_value, "%m/%d/%Y"
                            )
                        except Exception as e:
                            _logger.critical(e)

                add_defaults[property_name] = default_property_value
            elif verbose:
                _logger.warning(
                    "No default value provided for {0}...skipping record".format(
                        property_name
                    )
                )

            row_cntr += 1

    return add_defaults


def _process_directory_sheet(
    directory_book,
    directory_sheet,
    directory_property_map=None,
    ignore_empty_strings=True,
    verbose=False,
):
    # Returns additional fields (list) & properties (dictionary) as loaded from the specified directory book & sheet
    add_fields = []
    add_properties = {}
    if directory_book is not None:
        if not os.path.isfile(directory_book):
            raise IOError(f"File does not exist: {directory_book}")
        else:
            wb = load_workbook(directory_book, data_only=True)

        if directory_sheet is None:
            raise ValueError("No directory sheet specified")
        else:
            sheet = wb.get_sheet_by_name(directory_sheet)

        # Make sure sheet was found within the book...
        if sheet is None:  # or type(sheet) == type(None):
            raise ValueError(f"Specified sheet absent: {directory_sheet}")

        # Step 1: First, process columns in the header row of the Directory Sheet (ignoring first column)
        # ---------------------------------------------------------------------------------------------------
        _directory_property_columns = {}
        col_cntr = 1  # <-- Want to INTENTIONALLY skip the first column (is it is reserved for/assumed to be the fieldnames)
        while True:
            col_cntr += 1  # <-- Should INTENTIONALLY start with 2 the first cycle (so the first/fieldname column is NOT treated as a possible property column)
            col = get_column_letter(col_cntr)
            # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
            header_name = sheet[f"{col}1"].value

            if header_name is None:  # or type(header_name) == type(None):
                break

            # If no directory_property_map arg provided, treat each header as a direct property name
            # NOTE: this case will load use all columns
            if directory_property_map is None:
                local_property = header_name
                if local_property not in add_properties:
                    add_properties[local_property] = {}
                    _directory_property_columns[local_property] = col
            # If directory_property_map arg provided, then check whether current header name is specified
            # (otherwise this column will be ignored)
            elif header_name in directory_property_map:
                if not isinstance(directory_property_map[header_name], str) or (
                    len(directory_property_map[header_name]) == 0
                ):
                    raise ValueError(
                        f"Empty or invalid: directory_property_map[{header_name}]"
                    )
                local_property = directory_property_map[header_name]
                if local_property not in add_properties:
                    add_properties[local_property] = {}
                    _directory_property_columns[local_property] = col

            # Otherwise, gracefully ignore any headers that are not specified in the directory_property_map dictionary provided

        # Report any properties whose header name were not found in header list
        # ----------------------------------------------------------------------------
        if directory_property_map is not None:
            for property_name, property_header in directory_property_map.items():
                if property_header not in list(_directory_property_columns.keys()):
                    raise ValueError(
                        "Specified header:{0} for property:{1} absent from sheet:{2}".format(
                            property_header, property_name, directory_sheet
                        )
                    )

        # Step 2: Process Each Row of the Directory Sheet
        # ----------------------------------------------------
        # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
        keyspace = sheet["A1"].value

        row_cntr = 2  # <-- Assume first row is header row (ignore first row)
        while True:
            if verbose and ((row_cntr % 100) == 0):
                _logger.debug(f"Processing Sheet:{directory_sheet} @Row: {row_cntr}...")

            # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
            field = sheet[
                "A%i" % row_cntr
            ].value  # <-- Treat first column as field column (regardless of column header name)

            # Check whether we have reached the last property_name (last row) in the sheet
            # [Note, you can also use syntax such as: sheet.cell(row = 4, column = 2).value]
            # ---------------------------------------------------------------------------------
            if field is None:
                break  # <-- Last field

            if field not in add_fields:
                add_fields.append(field)

            # Loop through applicable property names for this row...
            # ---------------------------------------------------------
            for property_name in list(add_properties.keys()):
                # New versions of openpyxl use bracket notation for string references (Ex: sheet['A1'])
                # property_value = sheet.cell('%s%i' % (_directory_property_columns[property_name],row_cntr)).value
                property_value = sheet[
                    "%s%i" % (_directory_property_columns[property_name], row_cntr)
                ].value

                # Check whether a cell value was provided...
                if (property_value is not None) and (
                    not ignore_empty_strings or (property_value != "")
                ):
                    # Add property_name to add_properties dictionary (of dictionaries) if necessary
                    if property_name not in add_properties:
                        add_properties[property_name] = {}

                    # Convert property_value to numeric type (if applicable)
                    if isinstance(property_value, str):
                        try:
                            if search("\.", property_value):
                                raise ValueError
                            property_value = int(property_value)
                        except ValueError:
                            try:
                                property_value = float(property_value)
                            except ValueError:
                                pass

                    if isinstance(property_value, int) or isinstance(
                        property_value, float
                    ):
                        pass
                    elif property_value == "None":
                        property_value = None
                    elif (property_value == "True") or (property_value == "TRUE"):
                        property_value = True
                    elif (property_value == "False") or (property_value == "FALSE"):
                        property_value = False
                    elif (
                        isinstance(property_value, str)
                        and len(property_value.split("/")) > 1
                    ):
                        try:
                            property_value = datetime.strptime(
                                property_value, "%m/%d/%Y:%H:%M:%S"
                            )
                        except Exception as e:
                            _logger.exception(e)
                            try:
                                property_value = datetime.strptime(
                                    property_value, "%m/%d/%Y"
                                )
                            except Exception as e:
                                _logger.exception(e)

                    # Finally, record the field-specific property assignment
                    add_properties[property_name][field] = property_value

            row_cntr += 1

    return (keyspace, add_fields, add_properties)
